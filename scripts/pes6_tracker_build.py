#!/usr/bin/env python3
"""
PES6 Phoenix-style patch over-performer tracker builder
========================================================

Reads a PES6 patch Excel file (Phoenix 2026 format or compatible) and produces
a multi-sheet XLSX tracker identical in structure to the one built for the
Phoenix 2026 draft.

USAGE
-----
    python pes6_tracker_build.py <input_excel> <output_xlsx> [--owned-teams owned.txt]

EXAMPLES
--------
    # Basic: use defaults
    python pes6_tracker_build.py Phoenix_2027.xlsx tracker_2027.xlsx

    # With a custom owned-teams list (plain text, one club per line)
    python pes6_tracker_build.py Phoenix_2027.xlsx tracker_2027.xlsx --owned-teams owned.txt

REQUIREMENTS
------------
    pip install pandas numpy scikit-learn openpyxl

PATCH COMPATIBILITY
-------------------
Expects the source Excel to contain three sheets:
  - Main data sheet: player stats with columns Name, Club, OVR, P (primary position), and the 26 standard attributes (ATTACK, DEFENSE, BALANCE, ... TEAM WORK)
  - Calc sheet: a subset of labelled rows with B1 columns (GK B1, CB/CWP B1, SB B1, WB B1, DMF B1, CMF B1, SMF B1, AMF B1, WF B1, SS B1, CF B1)
  - Abilities sheet: per-player binary flags for special abilities (REACTION, PLAYMAKING, ...)

Default sheet names are the Phoenix 2026 conventions. Override via --main-sheet, --calc-sheet, --abilities-sheet if your patch uses different names.

CONFIGURATION
-------------
Edit the CONFIG block below to change:
  - Banned OVR threshold (default 95)
  - OVR class bands (A/B/C/D)
  - Gap filter threshold (default 3.0)
  - Starred exception threshold (default 1.5)
  - Minimum rows per position × class cell (default 10)
  - Owned teams list
  - Name/club encoding fixes

(c) Written for PES6 Phoenix 2026 draft analysis. Free to reuse.
"""
import argparse
import json
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except AttributeError:
    pass

import numpy as np
import pandas as pd
from sklearn.linear_model import Ridge
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent))
from verify_source import verify_source


# =============================================================================
# CONFIG — edit these to change behaviour for a new patch
# =============================================================================

# Draft rules
BANNED_OVR = 95                     # players with OVR >= this cannot be drafted
PRIMARY_GAP_THRESHOLD = 3.0         # minimum B1 - OVR gap to count as over-performer
STARRED_GAP_THRESHOLD = 1.5         # relaxed threshold for players with abilities/COND
MIN_ROWS_PER_CELL = 10              # minimum rows per (position, class)
B1_MIN, B1_MAX = 80, 100            # B1 range to include (skip near-0 junk)

# OVR class bands (inclusive)
CLASS_BANDS = {
    'A': (90, 94),
    'B': (84, 89),
    'C': (79, 83),
    'D': (0, 78),
}

# Owned teams — replace with your draft's current list
DEFAULT_OWNED_TEAMS = {
    # Liga 1 (13 managers)
    'Aston Villa', 'Como', 'Crvena Zvezda', 'Crystal Palace', 'Fulham',
    'Olympique Lyonnais', "Borussia M'gladbach", 'Fortuna D\ufffdsseldorf',
    'Roma', 'Slavia Praha', 'Torino', 'Bochum', 'Villarreal',
    # Liga 2 (14 managers)
    'Milan', 'Atalanta', 'Athletic', 'Bayer Leverkusen', 'Leipzig',
    'Chelsea', 'Fenerbahce', 'Juventus', 'Napoli', 'Newcastle',
    'Nottingham Forest', 'Olympique Marseille', 'Betis', 'Schalke',
}

# Sheet names (override via CLI if your patch differs)
DEFAULT_MAIN_SHEET = '3. Paste (Value Only)'
DEFAULT_CALC_SHEET = 'Calc'
DEFAULT_ABILITIES_SHEET = '1. Paste cvs'

# Position formulas in Calc sheet
B1_TARGETS = {
    'GK':     'GK B1',
    'CB/CWP': 'CB/CWP B1',
    'SB':     'SB B1',
    'WB':     'WB B1',
    'DMF':    'DMF B1',
    'CMF':    'CMF B1',
    'SMF':    'SMF B1',
    'AMF':    'AMF B1',
    'WF':     'WF B1',
    'SS':     'SS B1',
    'CF':     'CF B1',
}

# Standard 26 attributes used for ridge regression
ATTR_COLS = [
    'ATTACK', 'DEFENSE', 'BALANCE', 'STAMINA', 'TOP SPEED', 'ACCELERATION',
    'RESPONSE', 'AGILITY', 'DRIBBLE ACCURACY', 'DRIBBLE SPEED',
    'SHORT PASS ACCURACY', 'SHORT PASS SPEED', 'LONG PASS ACCURACY',
    'LONG PASS SPEED', 'SHOT ACCURACY', 'SHOT POWER', 'SHOT TECHNIQUE',
    'FREE KICK ACCURACY', 'SWERVE', 'HEADING', 'JUMP', 'TECHNIQUE',
    'AGGRESSION', 'MENTALITY', 'GOAL KEEPING', 'TEAM WORK',
]

# Primary stats displayed per position in the Stats column
PRIMARY_STATS = {
    'GK':     ['RESPONSE', 'GOAL KEEPING', 'DEFENSE', 'JUMP'],
    'CB/CWP': ['DEFENSE', 'BALANCE', 'JUMP', 'HEADING', 'STAMINA'],
    'SB':     ['DEFENSE', 'BALANCE', 'LONG PASS ACCURACY', 'DRIBBLE ACCURACY', 'STAMINA'],
    'WB':     ['DEFENSE', 'BALANCE', 'LONG PASS ACCURACY', 'DRIBBLE ACCURACY', 'STAMINA'],
    'DMF':    ['LONG PASS ACCURACY', 'TECHNIQUE', 'SHORT PASS ACCURACY', 'DRIBBLE ACCURACY', 'STAMINA'],
    'CMF':    ['TECHNIQUE', 'DRIBBLE ACCURACY', 'SHORT PASS ACCURACY', 'LONG PASS ACCURACY', 'STAMINA'],
    'SMF':    ['DRIBBLE ACCURACY', 'TOP SPEED', 'TECHNIQUE', 'AGILITY', 'STAMINA'],
    'AMF':    ['TECHNIQUE', 'DRIBBLE ACCURACY', 'SHORT PASS ACCURACY', 'SHOT ACCURACY', 'LONG PASS ACCURACY'],
    'WF':     ['TECHNIQUE', 'DRIBBLE ACCURACY', 'ATTACK', 'AGILITY', 'TOP SPEED', 'SHOT ACCURACY'],
    'SS':     ['TECHNIQUE', 'DRIBBLE ACCURACY', 'AGILITY', 'SHOT ACCURACY', 'ATTACK'],
    'CF':     ['SHOT ACCURACY', 'TECHNIQUE', 'ATTACK', 'DRIBBLE ACCURACY', 'RESPONSE', 'BALANCE'],
}
SHORT_NAMES = {
    'RESPONSE': 'RSP', 'GOAL KEEPING': 'GK', 'DEFENSE': 'DEF', 'JUMP': 'JMP',
    'BALANCE': 'BAL', 'HEADING': 'HDG', 'STAMINA': 'STA', 'LONG PASS ACCURACY': 'LPA',
    'DRIBBLE ACCURACY': 'DRA', 'TECHNIQUE': 'TEC', 'SHORT PASS ACCURACY': 'SPA',
    'TOP SPEED': 'SPD', 'AGILITY': 'AGI', 'SHOT ACCURACY': 'SHA', 'ATTACK': 'ATK',
}

# Abilities to extract from the abilities sheet (binary flag columns)
ABILITY_LABELS = [
    ('REACTION', 'REACTION'), ('PLAYMAKING', 'PLAYMAKING'), ('PASSING', 'PASSING'),
    ('SCORING', 'SCORING'), ('1-1 SCORING', '1-1 SCORING'), ('POST PLAYER', 'POST PLAYER'),
    ('LINES', 'LINES'), ('MIDDLE SHOOTING', 'MIDDLE SHOOTING'), ('SIDE', 'SIDE'),
    ('CENTRE', 'CENTRE'), ('PENALTIES', 'PENALTIES'), ('1-TOUCH PASS', '1-TOUCH PASS'),
    ('OUTSIDE', 'OUTSIDE'), ('MARKING', 'MARKING'), ('SLIDING', 'SLIDING'),
    ('COVERING', 'COVERING'), ('D-LINE CONTROL', 'D-LINE CONTROL'),
    ('PENALTY STOPPER', 'PENALTY STOPPER'), ('1-ON-1 STOPPER', '1-ON-1 STOPPER'),
    ('LONG THROW', 'LONG THROW'),
]

# Heuristic name/club encoding fixes
NAME_FIXES = {
    '0': 'Free agent',
    'Atl\ufffdtico de Madrid': 'Atlético de Madrid',
    'Bayern M\ufffdnchen': 'Bayern München',
    'Fortuna D\ufffdsseldorf': 'Fortuna Düsseldorf',
    'Deportivo de La Coru\ufffda': 'Deportivo de La Coruña',
    'Alav\ufffds': 'Alavés', 'M\ufffdlaga': 'Málaga',
    # Extend as needed — common players
    'K. Mbapp\ufffd': 'K. Mbappé', 'O. Demb\ufffdl\ufffd': 'O. Dembélé',
    'L. Mart\ufffdnez': 'L. Martínez', 'T. Hern\ufffdndez': 'T. Hernández',
    'L. D\ufffdaz': 'L. Díaz', 'R. Le\ufffdo': 'R. Leão',
    'J. \ufffdlvarez': 'J. Álvarez', 'V. Gy\ufffdkeres': 'V. Gyökeres',
    'G. Magalh\ufffdes': 'G. Magalhães', 'S. Man\ufffd': 'S. Mané',
    'L. San\ufffd': 'L. Sané', 'B. Guimar\ufffdes': 'B. Guimarães',
    'I. Konat\ufffd': 'I. Konaté',
}

# =============================================================================
# PIPELINE FUNCTIONS
# =============================================================================

def load_data(excel_path, main_sheet, calc_sheet, abilities_sheet):
    """Load the three required sheets from the patch Excel."""
    df_main = pd.read_excel(excel_path, sheet_name=main_sheet)
    df_calc = pd.read_excel(excel_path, sheet_name=calc_sheet).dropna(subset=['Name', 'OVR']).reset_index(drop=True)

    # Abilities sheet has no proper header — assign columns based on Phoenix 2026 layout
    df_abil_raw = pd.read_excel(excel_path, sheet_name=abilities_sheet, header=0)
    df_abil_raw.columns = [str(c).strip() for c in df_abil_raw.columns]
    return df_main, df_calc, df_abil_raw


def fit_ridge_models(df_calc):
    """Train ridge regression per position formula. Returns models + R² per position."""
    X = df_calc[ATTR_COLS].apply(pd.to_numeric, errors='coerce').fillna(0).values
    models, r2 = {}, {}
    for pos, col in B1_TARGETS.items():
        if col not in df_calc.columns:
            print(f"  WARNING: '{col}' missing in Calc sheet — skipping {pos}")
            continue
        y = pd.to_numeric(df_calc[col], errors='coerce').values
        mask = ~np.isnan(y)
        m = Ridge(alpha=0.001, fit_intercept=True)
        m.fit(X[mask], y[mask])
        from sklearn.metrics import r2_score
        r2[pos] = r2_score(y[mask], m.predict(X[mask]))
        models[pos] = m
    return models, r2


def predict_b1(df_main, models):
    """Apply trained models to all players in main data sheet."""
    X = df_main[ATTR_COLS].apply(pd.to_numeric, errors='coerce').fillna(0).values
    for pos, model in models.items():
        df_main[B1_TARGETS[pos]] = model.predict(X)
    return df_main


def extract_abilities(df_abil_raw):
    """Extract special abilities as list + count per player, keyed by name+club."""
    # Find the Name and Club columns heuristically
    name_col = next((c for c in df_abil_raw.columns if str(c).upper().startswith('NAME')), df_abil_raw.columns[0])
    club_col = next((c for c in df_abil_raw.columns if 'CLUB' in str(c).upper() and 'NUMBER' not in str(c).upper()), None)
    if club_col is None:
        # Phoenix layout: CLUB is column index 80
        club_col = df_abil_raw.columns[80] if len(df_abil_raw.columns) > 80 else None

    # Build ability-column lookup
    abil_cols = []
    for col_name, label in ABILITY_LABELS:
        matched = next((c for c in df_abil_raw.columns if str(c).strip().upper() == col_name.upper()), None)
        if matched is not None:
            abil_cols.append((matched, label))

    def row_abilities(row):
        active = []
        for col, label in abil_cols:
            v = row.get(col, 0)
            try:
                if pd.notna(v) and int(v) == 1:
                    active.append(label)
            except (ValueError, TypeError):
                pass
        return active

    df_abil_raw['_abilities'] = df_abil_raw.apply(row_abilities, axis=1)
    df_abil_raw['_ability_stars'] = df_abil_raw['_abilities'].apply(len)
    df_abil_raw['_key'] = df_abil_raw[name_col].astype(str) + '|' + df_abil_raw[club_col].astype(str) if club_col else df_abil_raw[name_col].astype(str)
    return df_abil_raw[['_key', '_abilities', '_ability_stars']].drop_duplicates('_key', keep='first')


def fix_text(s):
    if pd.isna(s):
        return s
    s = str(s).replace('_x0007_', 'ć')
    return NAME_FIXES.get(s, s)


def ovr_to_class(ovr):
    for cls, (lo, hi) in CLASS_BANDS.items():
        if lo <= ovr <= hi:
            return cls
    return 'D'


def build_overperformers(df_main, abilities_df, owned_teams):
    """Apply filters and build the over-performer pool."""
    df_main['_key'] = df_main['Name'].astype(str) + '|' + df_main['Club'].astype(str)
    df_main = df_main.merge(abilities_df, on='_key', how='left')
    df_main['_abilities'] = df_main['_abilities'].apply(lambda v: v if isinstance(v, list) else [])
    df_main['_ability_stars'] = df_main['_ability_stars'].fillna(0).astype(int)

    df_main['club_clean'] = df_main['Club'].astype(str).apply(fix_text)
    df_main['name_clean'] = df_main['Name'].apply(fix_text)

    pool = df_main[(~df_main['club_clean'].isin({fix_text(t) for t in owned_teams})) &
                   (~df_main['Club'].astype(str).isin(owned_teams)) &
                   (df_main['OVR'] < BANNED_OVR)].copy()
    pool['class'] = pool['OVR'].apply(ovr_to_class)

    out = []
    for pos, col in B1_TARGETS.items():
        if col not in pool.columns:
            continue
        for cls in CLASS_BANDS:
            sub = pool[pool['class'] == cls].copy()
            sub['B1'] = sub[col]
            sub['Gap'] = sub['B1'] - sub['OVR']
            sub = sub[(sub['B1'] >= B1_MIN) & (sub['B1'] <= B1_MAX)]
            primary = sub[sub['Gap'] >= PRIMARY_GAP_THRESHOLD].copy()
            starred = sub[(sub['Gap'] >= STARRED_GAP_THRESHOLD) &
                          (sub['Gap'] < PRIMARY_GAP_THRESHOLD) &
                          ((sub['_ability_stars'] >= 3) | (sub['Condition/Fitness'] >= 6))].copy()
            primary['_starred'] = False
            starred['_starred'] = True
            combined = pd.concat([primary, starred]).sort_values('Gap', ascending=False)
            if len(combined) < MIN_ROWS_PER_CELL and len(sub) >= MIN_ROWS_PER_CELL:
                extra = sub[~sub.index.isin(combined.index)].sort_values('Gap', ascending=False).head(
                    MIN_ROWS_PER_CELL - len(combined)).copy()
                extra['_starred'] = False
                combined = pd.concat([combined, extra]).sort_values('Gap', ascending=False)

            for _, r in combined.iterrows():
                out.append({
                    'position': pos, 'class': cls, 'name': r['name_clean'],
                    'club': r['club_clean'], 'ovr': round(float(r['OVR']), 1),
                    'b1': round(float(r['B1']), 1), 'gap': round(float(r['Gap']), 2),
                    'cond': int(r['Condition/Fitness']), 'cond_top': int(r['Condition/Fitness']) >= 6,
                    'ability_stars': int(r['_ability_stars']),
                    'abilities': ', '.join(r['_abilities']) if r['_abilities'] else '',
                    'registered_pos': str(r.get('P', '')),
                    'stats': ' '.join(f"{SHORT_NAMES.get(s, s[:3])}:{int(r[s])}"
                                      for s in PRIMARY_STATS[pos] if pd.notna(r.get(s))),
                    'is_starred_exception': bool(r['_starred']),
                })
    return out, pool


# =============================================================================
# XLSX WRITER
# =============================================================================

def build_xlsx(rows, pool, r2_scores, output_path):
    """Build the multi-sheet tracker. Reuses the layout from the Phoenix 2026 build."""
    wb = Workbook()
    BOLD_WHITE = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    HEAD_FILL = PatternFill('solid', start_color='1F4E78')
    GK_FILL = PatternFill('solid', start_color='FFD966')
    CLASS_COLORS = {'A': 'FFE699', 'B': 'D9E1F2', 'C': 'E2EFDA', 'D': 'F4CCCC'}
    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT = Alignment(horizontal='left', vertical='center')
    pos_order = list(B1_TARGETS.keys())
    cls_order = list(CLASS_BANDS.keys())

    def head(ws, row, n):
        for c in range(1, n + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = BOLD_WHITE; cell.fill = HEAD_FILL; cell.alignment = CENTER

    def cls_style(cell, cls):
        cell.fill = PatternFill('solid', start_color=CLASS_COLORS[cls])
        cell.font = Font(bold=True); cell.alignment = CENTER

    # --- README sheet ---
    ws = wb.active; ws.title = 'README'
    ws.cell(row=1, column=1, value='PES6 OVER-PERFORMER DRAFT TRACKER').font = Font(size=16, bold=True, color='1F4E78')
    ws.merge_cells('A1:C1')
    ws.cell(row=3, column=1, value='Source:').font = Font(bold=True)
    ws.cell(row=3, column=2, value=f'{len(pool)} eligible players after filtering')
    ws.cell(row=4, column=1, value='Method:').font = Font(bold=True)
    ws.cell(row=4, column=2, value='Ridge regression on Calc-sheet labels (R² per position shown below)')
    ws.cell(row=5, column=1, value='Primary filter:').font = Font(bold=True)
    ws.cell(row=5, column=2, value=f'B1 {B1_MIN}–{B1_MAX}, Gap ≥ +{PRIMARY_GAP_THRESHOLD}')
    ws.cell(row=6, column=1, value='Starred exception:').font = Font(bold=True)
    ws.cell(row=6, column=2, value=f'Gap ≥ +{STARRED_GAP_THRESHOLD} with ≥3 abilities OR COND ≥ 6')
    ws.cell(row=7, column=1, value='Total over-performer entries:').font = Font(bold=True)
    ws.cell(row=7, column=2, value=len(rows))
    ws.cell(row=9, column=1, value='R² PER POSITION (reverse-engineered patch formula accuracy)').font = Font(bold=True, size=12, color='1F4E78')
    for i, (pos, r2) in enumerate(r2_scores.items()):
        ws.cell(row=10 + i, column=1, value=pos).font = Font(bold=True)
        ws.cell(row=10 + i, column=2, value=f'{r2:.6f}')
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 80

    # --- Master Pool ---
    ws2 = wb.create_sheet('Master Pool')
    headers = ['Position', 'Class', 'Native P', 'Name', 'OVR', 'B1', 'Gap', 'Club',
               'COND', 'COND ★', 'Abil ★', 'Abilities', 'Stats', 'Exception?', 'Interest', 'Notes']
    for i, h in enumerate(headers, 1): ws2.cell(row=1, column=i, value=h)
    head(ws2, 1, len(headers))

    def pos_idx(p): return pos_order.index(p) if p in pos_order else 99
    def cls_idx(c): return cls_order.index(c) if c in cls_order else 99
    rows_sorted = sorted(rows, key=lambda r: (pos_idx(r['position']), cls_idx(r['class']), -r['gap']))

    for r_idx, p in enumerate(rows_sorted, 2):
        cp = ws2.cell(row=r_idx, column=1, value=p['position']); cp.alignment = CENTER
        if p['position'] == 'GK': cp.fill = GK_FILL
        cls_style(ws2.cell(row=r_idx, column=2, value=p['class']), p['class'])
        ws2.cell(row=r_idx, column=3, value=p['registered_pos']).alignment = CENTER
        ws2.cell(row=r_idx, column=4, value=p['name']).alignment = LEFT
        c = ws2.cell(row=r_idx, column=5, value=p['ovr']); c.number_format = '0.0'; c.alignment = CENTER
        c = ws2.cell(row=r_idx, column=6, value=p['b1']); c.number_format = '0.0'; c.alignment = CENTER
        c = ws2.cell(row=r_idx, column=7, value=p['gap']); c.number_format = '+0.00;-0.00'; c.alignment = CENTER
        ws2.cell(row=r_idx, column=8, value=p['club']).alignment = LEFT
        ws2.cell(row=r_idx, column=9, value=p['cond']).alignment = CENTER
        c = ws2.cell(row=r_idx, column=10, value='★' if p['cond_top'] else '')
        c.alignment = CENTER
        if p['cond_top']: c.font = Font(color='C00000', bold=True)
        c = ws2.cell(row=r_idx, column=11, value=p['ability_stars'] if p['ability_stars'] else '')
        c.alignment = CENTER
        if p['ability_stars'] >= 3: c.font = Font(color='C00000', bold=True)
        ws2.cell(row=r_idx, column=12, value=p['abilities']).alignment = LEFT
        ws2.cell(row=r_idx, column=13, value=p['stats']).alignment = LEFT
        ws2.cell(row=r_idx, column=14, value='★' if p['is_starred_exception'] else '').alignment = CENTER

    last = 1 + len(rows_sorted)
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{last}'
    ws2.freeze_panes = 'E2'
    for i, w in enumerate([10, 7, 9, 26, 7, 7, 8, 28, 7, 7, 7, 40, 45, 10, 12, 30], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.conditional_formatting.add(f'G2:G{last}',
        ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                       mid_type='num', mid_value=3.5, mid_color='A9D08E',
                       end_type='num', end_value=12, end_color='375623'))
    dv = DataValidation(type='list', formula1='"Target,Consider,Maybe,Skip,Picked"', allow_blank=True)
    dv.add(f'O2:O{last}'); ws2.add_data_validation(dv)

    # --- Position × Class Grid (top 3 per cell) ---
    ws3 = wb.create_sheet('Position × Class Grid')
    ws3['A1'] = 'TOP 3 OVER-PERFORMERS PER POSITION × CLASS'
    ws3['A1'].font = Font(bold=True, size=13, color='1F4E78')
    ws3.merge_cells('A1:J1')
    grid_h = ['Position', 'Class', '#', 'Player', 'Native P', 'OVR', 'B1', 'Gap', 'Club', 'Flags']
    for i, h in enumerate(grid_h, 1): ws3.cell(row=3, column=i, value=h)
    head(ws3, 3, len(grid_h))
    by_pc = {}
    for r in rows: by_pc.setdefault((r['position'], r['class']), []).append(r)
    row = 4
    for pos in pos_order:
        for cls in cls_order:
            cell_rows = sorted(by_pc.get((pos, cls), []), key=lambda x: -x['gap'])[:3]
            if not cell_rows:
                cp = ws3.cell(row=row, column=1, value=pos); cp.font = Font(bold=True); cp.alignment = CENTER
                if pos == 'GK': cp.fill = GK_FILL
                cls_style(ws3.cell(row=row, column=2, value=cls), cls)
                ws3.cell(row=row, column=4, value='— (no over-performers)').font = Font(italic=True, color='808080')
                row += 1; continue
            for rank, p in enumerate(cell_rows, 1):
                if rank == 1:
                    cp = ws3.cell(row=row, column=1, value=pos); cp.font = Font(bold=True); cp.alignment = CENTER
                    if pos == 'GK': cp.fill = GK_FILL
                    cls_style(ws3.cell(row=row, column=2, value=cls), cls)
                ws3.cell(row=row, column=3, value=rank).alignment = CENTER
                ws3.cell(row=row, column=4, value=p['name'])
                ws3.cell(row=row, column=5, value=p['registered_pos']).alignment = CENTER
                c = ws3.cell(row=row, column=6, value=p['ovr']); c.number_format = '0.0'; c.alignment = CENTER
                c = ws3.cell(row=row, column=7, value=p['b1']); c.number_format = '0.0'; c.alignment = CENTER
                c = ws3.cell(row=row, column=8, value=p['gap']); c.number_format = '+0.00'; c.alignment = CENTER
                ws3.cell(row=row, column=9, value=p['club'])
                flags = []
                if p['cond_top']: flags.append(f"COND {p['cond']}★")
                if p['ability_stars'] >= 3: flags.append(f"{p['ability_stars']}★ abil")
                elif p['ability_stars'] >= 1: flags.append(f"{p['ability_stars']}★")
                if 'Free agent' in p['club']: flags.append('FA')
                ws3.cell(row=row, column=10, value=' · '.join(flags))
                row += 1
    for i, w in enumerate([10, 7, 4, 26, 9, 7, 7, 9, 28, 28], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = 'A4'

    # --- Top 50 Diamonds ---
    ws_d = wb.create_sheet('Top 50 Diamonds')
    ws_d['A1'] = 'TOP 50 MOST UNDER-VALUED PLAYER × POSITION COMBINATIONS'
    ws_d['A1'].font = Font(bold=True, size=13, color='C00000')
    ws_d.merge_cells('A1:J1')
    dh = ['Rank', 'Player', 'Native P', 'Position fit', 'Class', 'OVR', 'B1', 'Gap', 'Club', 'Flags']
    for i, h in enumerate(dh, 1): ws_d.cell(row=3, column=i, value=h)
    head(ws_d, 3, len(dh))
    for rank, p in enumerate(sorted(rows, key=lambda x: -x['gap'])[:50], 1):
        rr = 3 + rank
        ws_d.cell(row=rr, column=1, value=rank).alignment = CENTER
        ws_d.cell(row=rr, column=2, value=p['name']).font = Font(bold=True)
        ws_d.cell(row=rr, column=3, value=p['registered_pos']).alignment = CENTER
        cp = ws_d.cell(row=rr, column=4, value=p['position']); cp.alignment = CENTER
        if p['position'] == 'GK': cp.fill = GK_FILL
        cls_style(ws_d.cell(row=rr, column=5, value=p['class']), p['class'])
        c = ws_d.cell(row=rr, column=6, value=p['ovr']); c.number_format = '0.0'; c.alignment = CENTER
        c = ws_d.cell(row=rr, column=7, value=p['b1']); c.number_format = '0.0'; c.alignment = CENTER
        c = ws_d.cell(row=rr, column=8, value=p['gap']); c.number_format = '+0.00'; c.alignment = CENTER
        c.font = Font(bold=True, color='375623')
        ws_d.cell(row=rr, column=9, value=p['club'])
        flags = []
        if p['cond_top']: flags.append(f"COND {p['cond']}★")
        if p['ability_stars'] >= 3: flags.append(f"{p['ability_stars']}★ abil")
        elif p['ability_stars'] >= 1: flags.append(f"{p['ability_stars']}★")
        if 'Free agent' in p['club']: flags.append('FA')
        ws_d.cell(row=rr, column=10, value=' · '.join(flags))
    for i, w in enumerate([5, 26, 9, 12, 7, 7, 7, 8, 28, 28], 1):
        ws_d.column_dimensions[get_column_letter(i)].width = w
    ws_d.freeze_panes = 'A4'

    # --- GK Picks ---
    gk_pool = pool[pool['P'].astype(str) == 'GK'].copy()
    if len(gk_pool) > 0:
        ws_gk = wb.create_sheet('GK Picks')
        ws_gk['A1'] = 'GOALKEEPERS — TOP PICKS BY B1 + COND HIGHLIGHTS'
        ws_gk['A1'].font = Font(bold=True, size=13, color='1F4E78')
        ws_gk.merge_cells('A1:H1')
        ws_gk['A2'] = 'GK formula is tight — rank by B1 within class; prefer COND ≥ 6.'
        ws_gk['A2'].font = Font(italic=True, color='808080')
        gkh = ['Class', 'Player', 'Native P', 'OVR', 'GK B1', 'Gap', 'COND', 'Club']
        gk_pool['gap'] = gk_pool['GK B1'] - gk_pool['OVR']
        gk_pool['cls'] = gk_pool['OVR'].apply(ovr_to_class)
        row = 4
        ws_gk.cell(row=row, column=1, value='COND ≥ 6 GOALKEEPERS (rare reliable keepers)').font = Font(bold=True, size=12, color='C00000')
        ws_gk.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8); row += 1
        for i, h in enumerate(gkh, 1): ws_gk.cell(row=row, column=i, value=h)
        head(ws_gk, row, len(gkh)); row += 1
        for _, r in gk_pool[gk_pool['Condition/Fitness'] >= 6].sort_values('GK B1', ascending=False).iterrows():
            cls_style(ws_gk.cell(row=row, column=1, value=r['cls']), r['cls'])
            ws_gk.cell(row=row, column=2, value=r['name_clean']).font = Font(bold=True)
            ws_gk.cell(row=row, column=3, value=str(r['P'])).alignment = CENTER
            for idx, val, fmt in [(4, round(r['OVR'], 1), '0.0'), (5, round(r['GK B1'], 1), '0.0'), (6, round(r['gap'], 2), '+0.00')]:
                c = ws_gk.cell(row=row, column=idx, value=val); c.number_format = fmt; c.alignment = CENTER
            c = ws_gk.cell(row=row, column=7, value=int(r['Condition/Fitness']))
            c.font = Font(bold=True, color='C00000'); c.alignment = CENTER
            ws_gk.cell(row=row, column=8, value=r['club_clean']); row += 1
        row += 1
        ws_gk.cell(row=row, column=1, value='TOP 10 GOALKEEPERS BY B1 IN EACH CLASS').font = Font(bold=True, size=12, color='1F4E78')
        ws_gk.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8); row += 1
        for i, h in enumerate(gkh, 1): ws_gk.cell(row=row, column=i, value=h)
        head(ws_gk, row, len(gkh)); row += 1
        for cls in cls_order:
            top = gk_pool[gk_pool['cls'] == cls].nlargest(10, 'GK B1')
            if top.empty:
                cls_style(ws_gk.cell(row=row, column=1, value=cls), cls)
                ws_gk.cell(row=row, column=2, value='— no GKs in this class').font = Font(italic=True, color='808080'); row += 1; continue
            for idx, (_, r) in enumerate(top.iterrows()):
                cls_style(ws_gk.cell(row=row, column=1, value=cls if idx == 0 else ''), cls)
                ws_gk.cell(row=row, column=2, value=r['name_clean'])
                ws_gk.cell(row=row, column=3, value=str(r['P'])).alignment = CENTER
                for i_, val, fmt in [(4, round(r['OVR'], 1), '0.0'), (5, round(r['GK B1'], 1), '0.0'), (6, round(r['gap'], 2), '+0.00')]:
                    c = ws_gk.cell(row=row, column=i_, value=val); c.number_format = fmt; c.alignment = CENTER
                c = ws_gk.cell(row=row, column=7, value=int(r['Condition/Fitness']))
                if r['Condition/Fitness'] >= 6: c.font = Font(bold=True, color='C00000')
                c.alignment = CENTER
                ws_gk.cell(row=row, column=8, value=r['club_clean']); row += 1
        for i, w in enumerate([8, 28, 9, 7, 8, 7, 7, 30], 1):
            ws_gk.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='Build PES6 over-performer draft tracker.')
    parser.add_argument('input', help='Source patch Excel file')
    parser.add_argument('output', help='Output tracker XLSX path')
    parser.add_argument('--main-sheet', default=DEFAULT_MAIN_SHEET)
    parser.add_argument('--calc-sheet', default=DEFAULT_CALC_SHEET)
    parser.add_argument('--abilities-sheet', default=DEFAULT_ABILITIES_SHEET)
    parser.add_argument('--owned-teams', default=None, help='Optional text file, one club per line')
    args = parser.parse_args()

    owned = set(DEFAULT_OWNED_TEAMS)
    if args.owned_teams:
        with open(args.owned_teams, 'r', encoding='utf-8') as f:
            owned = {line.strip() for line in f if line.strip()}

    print(f"Verifying {args.input}...")
    verify_source(args.input)

    print(f"Loading {args.input}...")
    df_main, df_calc, df_abil = load_data(args.input, args.main_sheet, args.calc_sheet, args.abilities_sheet)
    print(f"  Main sheet:       {len(df_main)} players")
    print(f"  Calc sheet:       {len(df_calc)} labelled rows (training data)")
    print(f"  Abilities sheet:  {len(df_abil)} rows")

    print("\nFitting ridge regression per position...")
    models, r2 = fit_ridge_models(df_calc)
    for pos, score in r2.items():
        print(f"  {pos:8s} R² = {score:.6f}")

    print("\nPredicting B1 for all players...")
    df_main = predict_b1(df_main, models)

    print("Extracting abilities...")
    abilities_df = extract_abilities(df_abil)
    print(f"  Players with ≥1 ability: {(abilities_df['_ability_stars'] > 0).sum()} / {len(abilities_df)}")

    print(f"\nApplying filters (OVR < {BANNED_OVR}, {len(owned)} owned teams excluded)...")
    rows, pool = build_overperformers(df_main, abilities_df, owned)
    print(f"  Eligible pool:             {len(pool)}")
    print(f"  Over-performer entries:    {len(rows)}")

    print(f"\nWriting tracker to {args.output}...")
    build_xlsx(rows, pool, r2, args.output)
    print("Done.")


if __name__ == '__main__':
    main()
