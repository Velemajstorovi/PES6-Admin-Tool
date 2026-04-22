"""
PES6 ADMIN TOOL — Full workbook builder.

Creates an Excel workbook with native formulas that:
  1. Takes a pasted CSV export on sheet 'CSV Paste' (from PES Fan Editor)
  2. Computes all derived values (OVR, B1, composite stats, ability bonuses, COND multiplier)
  3. Ranks all players
  4. Ranks all clubs + national teams (top-18 best B1)
  5. Tiers teams A-F based on B1
  6. Shows per-team roster composition (A/B/C/D counts)

No Python dependency after build — everything lives in Excel formulas.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.workbook.defined_name import DefinedName

# ==============================================================
# SHEET LAYOUTS
# ==============================================================
# Sheet 1: "CSV Paste" — raw CSV goes here. 82 columns.
# Sheet 2: "Players"   — derived per-player data, pulls from CSV Paste
# Sheet 3: "Clubs"     — top-18 per club, tiered
# Sheet 4: "Nations"   — top-18 per nation, tiered
# Sheet 5: "Team Composition" — A/B/C/D counts per team
# Sheet 6: "Draft Rules" — reference
# Sheet 7: "README"    — usage guide

# CSV column headers (82 columns) — matches PES Fan Editor export
CSV_HEADERS = [
    'NAME',
    'GK 0','CWP 2','CBT 3','SB 4','DMF 5','WB 6','CMF 7','SMF 8','AMF 9','WF 10','SS 11','CF 12',
    'REGISTERED POSITION','HEIGHT','STRONG FOOT','FAVOURED SIDE',
    'WEAK FOOT ACCURACY','WEAK FOOT FREQUENCY',
    'ATTACK','DEFENSE','BALANCE','STAMINA','TOP SPEED','ACCELERATION','RESPONSE','AGILITY',
    'DRIBBLE ACCURACY','DRIBBLE SPEED','SHORT PASS ACCURACY','SHORT PASS SPEED',
    'LONG PASS ACCURACY','LONG PASS SPEED','SHOT ACCURACY','SHOT POWER','SHOT TECHNIQUE',
    'FREE KICK ACCURACY','SWERVE','HEADING','JUMP','TECHNIQUE','AGGRESSION','MENTALITY',
    'GOAL KEEPING','TEAM WORK',
    'CONSISTENCY','CONDITION / FITNESS',
    'DRIBBLING','TACTICAL DRIBBLE','POSITIONING',
    'REACTION','PLAYMAKING','PASSING','SCORING','1-1 SCORING','POST PLAYER','LINES',
    'MIDDLE SHOOTING','SIDE','CENTRE','PENALTIES','1-TOUCH PASS','OUTSIDE',
    'MARKING','SLIDING','COVERING','D-LINE CONTROL','PENALTY STOPPER','1-ON-1 STOPPER','LONG THROW',
    'INJURY TOLERANCE','DRIBBLE STYLE','FREE KICK STYLE','PK STYLE','DROP KICK STYLE',
    'AGE','WEIGHT','NATIONALITY','INTERNATIONAL NUMBER','CLASSIC NUMBER','CLUB','CLUB NUMBER'
]
MAX_PLAYERS = 6000  # safe upper bound (Phoenix 2026 has 4,783)

# Players-sheet column layout (77 columns)
# 'Effective Club' follows 'Club' - equals Club unless overridden by an entry
# in the Player Changes sheet AND the Apply Changes toggle is YES. Clubs and
# Team Composition sheets group by Effective Club so the applied roster is
# reflected downstream.
PLAYER_COLS = [
    'Name','Nationality','Club','Effective Club','Age','Foot','Registered Pos','OVR','Class',
    # Raw attributes (26)
    'ATTACK','DEFENSE','BALANCE','STAMINA','TOP SPEED','ACCELERATION','RESPONSE','AGILITY',
    'DRIBBLE ACC','DRIBBLE SPD','SHORT PASS ACC','SHORT PASS SPD','LONG PASS ACC','LONG PASS SPD',
    'SHOT ACC','SHOT PWR','SHOT TEC','FREE KICK','SWERVE','HEADING','JUMP','TECHNIQUE',
    'AGGRESSION','MENTALITY','GOAL KEEPING','TEAM WORK',
    'CONDITION',
    # Composites (7)
    'Attacking Prowess','Finishing','Speed','Explosive Power','Kicking Power',
    'Defending Prowess','Ball Winning',
    # Abilities (20)
    'REACTION','PLAYMAKING','PASSING','SCORING','1-1 SCORING','POST PLAYER','LINES',
    'MIDDLE SHOOTING','SIDE','CENTRE','PENALTIES','1-TOUCH PASS','OUTSIDE',
    'MARKING','SLIDING','COVERING','D-LINE CONTROL','PENALTY STOPPER','1-ON-1 STOPPER','LONG THROW',
    'Ability ★',
    # B1 per position (11)
    'GK B1','CB/CWP B1','SB B1','WB B1','DMF B1','CMF B1','SMF B1','AMF B1','WF B1','SS B1','CF B1',
    # Derived
    'Best B1','Best Position','Ability Bonus','COND Multiplier','Effective B1',
]

# Build column letter lookup
def col_letters(names):
    return {name: get_column_letter(i+1) for i, name in enumerate(names)}

CSV = col_letters(CSV_HEADERS)
P = col_letters(PLAYER_COLS)

# ==============================================================
# ABILITY WEIGHTS PER POSITION
# ==============================================================
# These bonuses reflect how much each ability helps at each position.
# Values are multiplicative adjustments to B1 (e.g., +0.02 = +2% boost to Best B1).
# Calibrated from community knowledge of which abilities matter most where.

# Structure: {position: {ability: weight_contribution}}
# Max possible bonus capped at ~15% (i.e., sum of abilities × weight ≤ 0.15)
ABILITY_WEIGHTS = {
    'GK': {
        'PENALTY STOPPER': 0.035, '1-ON-1 STOPPER': 0.035, 'REACTION': 0.02,
        'COVERING': 0.01, 'D-LINE CONTROL': 0.005,
    },
    'CB/CWP': {
        'MARKING': 0.03, 'COVERING': 0.025, 'SLIDING': 0.02,
        'D-LINE CONTROL': 0.02, 'REACTION': 0.015, '1-TOUCH PASS': 0.005,
    },
    'SB': {
        'MARKING': 0.025, 'COVERING': 0.02, 'SLIDING': 0.015, 'REACTION': 0.015,
        'SIDE': 0.015, '1-TOUCH PASS': 0.01, 'LONG THROW': 0.005, 'PASSING': 0.005,
    },
    'WB': {
        'SIDE': 0.025, 'MARKING': 0.02, 'COVERING': 0.015, 'REACTION': 0.015,
        'PASSING': 0.015, '1-TOUCH PASS': 0.01, 'SLIDING': 0.01, 'LONG THROW': 0.005,
    },
    'DMF': {
        'MARKING': 0.025, 'COVERING': 0.02, 'PLAYMAKING': 0.02, 'PASSING': 0.02,
        '1-TOUCH PASS': 0.02, 'REACTION': 0.015, 'MIDDLE SHOOTING': 0.01, 'OUTSIDE': 0.005,
    },
    'CMF': {
        'PLAYMAKING': 0.03, 'PASSING': 0.025, '1-TOUCH PASS': 0.02, 'REACTION': 0.015,
        'LINES': 0.015, 'MIDDLE SHOOTING': 0.01, 'PENALTIES': 0.005, 'OUTSIDE': 0.01,
    },
    'SMF': {
        'SIDE': 0.03, 'PASSING': 0.02, 'OUTSIDE': 0.015, 'REACTION': 0.015,
        '1-TOUCH PASS': 0.01, 'LINES': 0.01, 'CENTRE': 0.01,
    },
    'AMF': {
        'PLAYMAKING': 0.03, 'PASSING': 0.025, 'LINES': 0.02, '1-TOUCH PASS': 0.02,
        'MIDDLE SHOOTING': 0.015, 'REACTION': 0.015, 'PENALTIES': 0.005, 'OUTSIDE': 0.01,
    },
    'WF': {
        'SIDE': 0.025, 'OUTSIDE': 0.02, 'SCORING': 0.015, '1-1 SCORING': 0.015,
        'PASSING': 0.015, 'REACTION': 0.015, 'CENTRE': 0.01, 'MIDDLE SHOOTING': 0.005,
    },
    'SS': {
        'SCORING': 0.025, '1-1 SCORING': 0.025, 'REACTION': 0.02, 'LINES': 0.015,
        'PLAYMAKING': 0.015, 'PASSING': 0.01, '1-TOUCH PASS': 0.01, 'MIDDLE SHOOTING': 0.005,
    },
    'CF': {
        'SCORING': 0.03, '1-1 SCORING': 0.03, 'POST PLAYER': 0.02, 'REACTION': 0.02,
        'CENTRE': 0.015, 'PENALTIES': 0.01, 'LINES': 0.01, 'MIDDLE SHOOTING': 0.005,
    },
}

POSITIONS = ['GK','CB/CWP','SB','WB','DMF','CMF','SMF','AMF','WF','SS','CF']

# ==============================================================
# FORMULA BUILDERS
# ==============================================================

def csvref(col_name, row):
    """Reference a CSV Paste sheet column at given row."""
    return f"'CSV Paste'!{CSV[col_name]}{row}"

def pref(col_name, row):
    """Reference a Players sheet cell (within the same sheet)."""
    return f"{P[col_name]}{row}"

# -- Raw pulls from CSV
def pull_formula(dest_col, src_col, row):
    """Pull a value from CSV, returning blank if empty."""
    src = csvref(src_col, row)
    return f"=IF({src}=\"\",\"\",{src})"

# -- Composite stats (from Calc sheet's formulas, adapted)
def attacking_prowess(row):
    return (f"=IF({pref('ATTACK', row)}=\"\",\"\","
            f"MAX(2/3*{pref('ATTACK', row)}+1/3*{pref('RESPONSE', row)},"
            f"2/3*{pref('ATTACK', row)}+1/3*{pref('AGGRESSION', row)}))")

def finishing(row):
    return (f"=IF({pref('SHOT ACC', row)}=\"\",\"\","
            f"{pref('SHOT ACC', row)}*0.84+{pref('SHOT TEC', row)}*0.16)")

def speed(row):
    return (f"=IF({pref('TOP SPEED', row)}=\"\",\"\","
            f"{pref('DRIBBLE SPD', row)}*(1-(((1/1.3)+0.5)/2))"
            f"+{pref('TOP SPEED', row)}*(((1/1.3)+0.5)/2))")

def explosive_power(row):
    return (f"=IF({pref('ACCELERATION', row)}=\"\",\"\","
            f"(2*{pref('AGILITY', row)}+{pref('ACCELERATION', row)})/3)")

def kicking_power(row):
    """=MAX($AC2,$AA2,$Y2)*0.4+$AC2*(0.6/3)+$AA2*(0.6/3)+$Y2*(0.6/3)
       where AC=SHOT PWR, AA=LONG PASS SPD, Y=SHORT PASS SPD"""
    return (f"=IF({pref('SHOT PWR', row)}=\"\",\"\","
            f"MAX({pref('SHOT PWR', row)},{pref('LONG PASS SPD', row)},{pref('SHORT PASS SPD', row)})*0.4"
            f"+{pref('SHOT PWR', row)}*(0.6/3)+{pref('LONG PASS SPD', row)}*(0.6/3)"
            f"+{pref('SHORT PASS SPD', row)}*(0.6/3))")

def defending_prowess(row):
    return (f"=IF({pref('DEFENSE', row)}=\"\",\"\","
            f"MAX(2/3*{pref('DEFENSE', row)}+1/3*{pref('RESPONSE', row)},"
            f"2/3*{pref('DEFENSE', row)}+1/3*{pref('TEAM WORK', row)}))")

def ball_winning(row):
    return (f"=IF({pref('DEFENSE', row)}=\"\",\"\","
            f"MAX(2/3*{pref('DEFENSE', row)}+1/3*{pref('BALANCE', row)},"
            f"2/3*{pref('DEFENSE', row)}+1/3*{pref('MENTALITY', row)}))")

# -- B1 formulas per position (using Players sheet columns)
def b1(position, row):
    """Exact patch-maker B1 formula, expressed via Players-sheet columns."""
    r = str(row)
    s = lambda col: pref(col, row)
    if position == 'GK':
        return (f"=IF({pref('GOAL KEEPING', row)}=\"\",\"\","
                f"SUM((({s('GOAL KEEPING')}*2+{s('DEFENSE')})/3-25)*0.52,"
                f"({s('RESPONSE')}-25)*0.52,({s('BALANCE')}-25)*0.12,"
                f"({s('JUMP')}-25)*0.12,7))")
    if position == 'CB/CWP':
        return (f"=IF({pref('DEFENSE', row)}=\"\",\"\","
                f"SUM(({s('HEADING')}-25)*0.2,({s('Defending Prowess')}-25)*0.27,"
                f"({s('Ball Winning')}-25)*0.27,({s('Speed')}-25)*0.11,"
                f"({s('BALANCE')}-25)*0.21,({s('JUMP')}-25)*0.21,"
                f"({s('STAMINA')}-25)*0.1,7))")
    if position in ('SB', 'WB'):
        return (f"=IF({pref('DEFENSE', row)}=\"\",\"\","
                f"SUM(({s('Attacking Prowess')}-25)*0.06,({s('TECHNIQUE')}-25)*0.1,"
                f"({s('DRIBBLE ACC')}-25)*0.15,({s('LONG PASS ACC')}-25)*0.15,"
                f"({s('Defending Prowess')}-25)*0.15,({s('Ball Winning')}-25)*0.14,"
                f"({s('Speed')}-25)*0.15,({s('Explosive Power')}-25)*0.15,"
                f"({s('BALANCE')}-25)*0.12,({s('JUMP')}-25)*0.12,"
                f"({s('STAMINA')}-25)*0.13,8))")
    if position == 'DMF':
        return (f"=IF({pref('DEFENSE', row)}=\"\",\"\","
                f"SUM(({s('Attacking Prowess')}-25)*0.07,({s('TECHNIQUE')}-25)*0.19,"
                f"({s('DRIBBLE ACC')}-25)*0.16,({s('SHORT PASS ACC')}-25)*0.19,"
                f"({s('LONG PASS ACC')}-25)*0.2,({s('SWERVE')}-25)*0.13,"
                f"({s('Defending Prowess')}-25)*0.07,({s('Ball Winning')}-25)*0.03,"
                f"({s('Speed')}-25)*0.03,({s('Explosive Power')}-25)*0.03,"
                f"({s('BALANCE')}-25)*0.14,({s('JUMP')}-25)*0.05,"
                f"({s('STAMINA')}-25)*0.15,8))")
    if position == 'CMF':
        return (f"=IF({pref('DEFENSE', row)}=\"\",\"\","
                f"SUM(({s('Attacking Prowess')}-25)*0.05,({s('TECHNIQUE')}-25)*0.25,"
                f"({s('DRIBBLE ACC')}-25)*0.25,({s('SHORT PASS ACC')}-25)*0.25,"
                f"({s('LONG PASS ACC')}-25)*0.22,({s('Defending Prowess')}-25)*0.03,"
                f"({s('Speed')}-25)*0.04,({s('Explosive Power')}-25)*0.06,"
                f"({s('BALANCE')}-25)*0.05,({s('STAMINA')}-25)*0.18,7))")
    if position == 'SMF':
        return (f"=IF({pref('DEFENSE', row)}=\"\",\"\","
                f"SUM(({s('Attacking Prowess')}-25)*0.07,({s('TECHNIQUE')}-25)*0.16,"
                f"({s('DRIBBLE ACC')}-25)*0.26,({s('SHORT PASS ACC')}-25)*0.07,"
                f"({s('LONG PASS ACC')}-25)*0.13,({s('SWERVE')}-25)*0.04,"
                f"({s('Speed')}-25)*0.26,({s('Explosive Power')}-25)*0.23,"
                f"({s('STAMINA')}-25)*0.14,7))")
    if position == 'AMF':
        return (f"=IF({pref('DEFENSE', row)}=\"\",\"\","
                f"SUM(({s('Attacking Prowess')}-25)*0.15,({s('TECHNIQUE')}-25)*0.23,"
                f"({s('DRIBBLE ACC')}-25)*0.23,({s('SHORT PASS ACC')}-25)*0.23,"
                f"({s('LONG PASS ACC')}-25)*0.15,({s('Finishing')}-25)*0.18,"
                f"({s('Speed')}-25)*0.05,({s('Explosive Power')}-25)*0.07,"
                f"({s('BALANCE')}-25)*0.05,({s('STAMINA')}-25)*0.03,7))")
    if position == 'WF':
        return (f"=IF({pref('DEFENSE', row)}=\"\",\"\","
                f"SUM(({s('Attacking Prowess')}-25)*0.18,({s('TECHNIQUE')}-25)*0.22,"
                f"({s('DRIBBLE ACC')}-25)*0.22,({s('SHORT PASS ACC')}-25)*0.05,"
                f"({s('LONG PASS ACC')}-25)*0.1,({s('Finishing')}-25)*0.12,"
                f"({s('Kicking Power')}-25)*0.05,({s('Speed')}-25)*0.16,"
                f"({s('Explosive Power')}-25)*0.16,({s('BALANCE')}-25)*0.06,"
                f"({s('STAMINA')}-25)*0.06,9))")
    if position == 'SS':
        return (f"=IF({pref('DEFENSE', row)}=\"\",\"\","
                f"SUM(({s('Attacking Prowess')}-25)*0.16,({s('TECHNIQUE')}-25)*0.2,"
                f"({s('DRIBBLE ACC')}-25)*0.2,({s('SHORT PASS ACC')}-25)*0.1,"
                f"({s('LONG PASS ACC')}-25)*0.1,({s('Finishing')}-25)*0.15,"
                f"({s('Kicking Power')}-25)*0.06,({s('Speed')}-25)*0.1,"
                f"({s('Explosive Power')}-25)*0.2,({s('BALANCE')}-25)*0.07,"
                f"({s('STAMINA')}-25)*0.04,7))")
    if position == 'CF':
        return (f"=IF({pref('DEFENSE', row)}=\"\",\"\","
                f"SUM(({s('Attacking Prowess')}-25)*0.33,({s('TECHNIQUE')}-25)*0.25,"
                f"({s('DRIBBLE ACC')}-25)*0.15,({s('Finishing')}-25)*0.38,"
                f"({s('HEADING')}-25)*0.03,({s('Speed')}-25)*0.05,"
                f"({s('Explosive Power')}-25)*0.05,({s('BALANCE')}-25)*0.1,"
                f"({s('JUMP')}-25)*0.03,7))")
    return "0"

def ovr_formula(row):
    """OVR = max B1 across all positions."""
    b1_refs = ",".join(pref(f'{p} B1', row) for p in POSITIONS)
    return f"=IF({pref('Name', row)}=\"\",\"\",ROUND(MAX({b1_refs}),1))"

def best_b1(row):
    refs = ",".join(pref(f'{p} B1', row) for p in POSITIONS)
    return f"=IF({pref('Name', row)}=\"\",\"\",ROUND(MAX({refs}),2))"

def best_position(row):
    """Find which B1 column had the max value.

    Nested IF over the 11 B1 columns. Previously attempted with
    INDEX/MATCH on an inline array literal, but Excel treats
    ``(ref1,ref2,...)`` as a function-args list, not an array,
    so MATCH fails. The IF-chain is robust across Excel / Calc.
    """
    b1_cols = [pref(f'{p} B1', row) for p in POSITIONS]
    max_expr = f"MAX({','.join(b1_cols)})"
    expr = '""'
    for pos, col in reversed(list(zip(POSITIONS, b1_cols))):
        expr = f'IF({col}={max_expr},"{pos}",{expr})'
    return f"=IF({pref('Name', row)}=\"\",\"\",{expr})"

def ability_count(row):
    """Count of active abilities (sum of 20 binary flags)."""
    abilities = ['REACTION','PLAYMAKING','PASSING','SCORING','1-1 SCORING','POST PLAYER','LINES',
                 'MIDDLE SHOOTING','SIDE','CENTRE','PENALTIES','1-TOUCH PASS','OUTSIDE',
                 'MARKING','SLIDING','COVERING','D-LINE CONTROL','PENALTY STOPPER','1-ON-1 STOPPER','LONG THROW']
    refs = [pref(a, row) for a in abilities]
    sum_expr = "+".join(f"IF({r}=\"\",0,{r})" for r in refs)
    return f"=IF({pref('Name', row)}=\"\",\"\",{sum_expr})"

def ability_bonus(row):
    """Sum of (ability * weight) for the player's BEST position.

    For each position, emit IF(best_pos="POS", sum_of_terms, 0), then
    sum across positions. Depends on Best Position having been computed.
    """
    best_pos_ref = pref('Best Position', row)
    per_pos = []
    for pos in POSITIONS:
        terms = []
        for ability, weight in ABILITY_WEIGHTS.get(pos, {}).items():
            ab_ref = pref(ability, row)
            terms.append(f"IF({ab_ref}=\"\",0,{ab_ref})*{weight}")
        if terms:
            per_pos.append(f'IF({best_pos_ref}="{pos}",({"+".join(terms)}),0)')
    return f"=IF({pref('Name', row)}=\"\",\"\",({'+'.join(per_pos)}))"

def cond_multiplier(row):
    """Condition impact: 
    COND 7 = 1.06 (6% boost), COND 6 = 1.03, COND 5 = 1.00, COND 4 = 0.98, COND 3 = 0.96, COND ≤2 = 0.94
    Based on patch-guide info: higher COND = higher probability of positive arrow."""
    c = pref('CONDITION', row)
    return (f"=IF({pref('Name', row)}=\"\",\"\","
            f"IF({c}=\"\",1,"
            f"IF({c}>=7,1.06,"
            f"IF({c}>=6,1.03,"
            f"IF({c}>=5,1.00,"
            f"IF({c}>=4,0.98,"
            f"IF({c}>=3,0.96,0.94)))))))")

def effective_b1(row):
    """Best B1 × Ability bonus × Condition multiplier.
    This is the TRUE player-value metric used for rankings."""
    bb1 = pref('Best B1', row)
    ab = pref('Ability Bonus', row)
    cm = pref('COND Multiplier', row)
    return (f"=IF({pref('Name', row)}=\"\",\"\","
            f"ROUND({bb1}*(1+{ab})*{cm},2))")

def player_class(row):
    """A/B/C/D based on OVR."""
    o = pref('OVR', row)
    return (f"=IF({pref('Name', row)}=\"\",\"\","
            f"IF({o}>=90,\"A\","
            f"IF({o}>=84,\"B\","
            f"IF({o}>=78,\"C\",\"D\"))))")


def effective_club_formula(row, log_first=7, log_last=206):
    """Effective Club = override from Player Changes log if one exists AND
    the Apply Changes toggle is YES; otherwise the original CSV Club.

    Uses LOOKUP(2, 1/(cond), range) so the LAST matching row wins when a
    player has multiple log entries (later changes supersede earlier ones).
    Portable across Excel and LibreOffice (no CSE or _xlfn. prefix needed).
    """
    name = pref('Name', row)
    club = pref('Club', row)
    apply_ref = "'Player Changes'!$B$3"
    log_player = f"'Player Changes'!$C${log_first}:$C${log_last}"
    log_toclub = f"'Player Changes'!$E${log_first}:$E${log_last}"
    lookup = (
        f'IFERROR(LOOKUP(2,1/(({log_player}={name})*({log_toclub}<>"")),'
        f'{log_toclub}),{club})'
    )
    return (
        f'=IF({name}="","",'
        f'IF({apply_ref}<>"YES",{club},{lookup}))'
    )
