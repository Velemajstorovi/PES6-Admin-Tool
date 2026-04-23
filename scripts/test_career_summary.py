"""Seed Season Results with a few seasons of data across several managers,
run LibreOffice headless recalc, then read Career Summary back and verify
the formulas compute expected aggregates.
"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except AttributeError:
    pass

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
SRC = REPO / 'outputs' / 'pes6_admin_tool.xlsx'
TEST_DIR = REPO / 'outputs' / '_career_test'
TEST = TEST_DIR / 'career_test.xlsx'
RECALC_DIR = REPO / 'outputs' / '_career_recalc'
TEST_RECALC = RECALC_DIR / 'career_test.xlsx'

# (Season, Manager, Finish, W, D, L, GF, GA, YC, RC,
#  PL, Cup, Europa, Conference, Friendly Cup, SP-World Cup, Outcome)
SEED_ROWS = [
    # Manager M-L1-12 (Bochum) - 3 seasons
    (1, 'M-L1-12', 3,  16, 6, 2,  48, 22, 45, 2,  'Winner', 'Semi',   '', '', '', '', 'Stayed'),
    (2, 'M-L1-12', 1,  20, 3, 1,  62, 18, 38, 1,  'Final',  'Winner', '', '', '', '', 'Stayed'),
    (3, 'M-L1-12', 8,  10, 6, 8,  35, 34, 52, 3,  'Group',  'QF',     '', '', '', '', 'Stayed'),
    # Manager M-L2-14 (Schalke) - 2 seasons, gets promoted
    (1, 'M-L2-14', 1,  22, 3, 1,  58, 15, 40, 1,  '',       'R16',    '', '', '', '', 'Promoted'),
    (2, 'M-L2-14', 12, 6,  5, 15, 22, 48, 60, 4,  'DNQ',    'Group',  '', '', '', '', 'Relegated'),
    # Manager M-L1-01 (Aston Villa) - 1 season
    (1, 'M-L1-01', 5,  13, 7, 4,  40, 30, 50, 2,  'QF',     'Final',  '', '', '', '', 'Stayed'),
]


def seed() -> None:
    TEST_DIR.mkdir(parents=True, exist_ok=True)
    RECALC_DIR.mkdir(parents=True, exist_ok=True)
    # Clean previous runs
    for f in (TEST, TEST_RECALC):
        if f.exists():
            f.unlink()
    wb = load_workbook(SRC)
    sr = wb['Season Results']
    # Column layout: A=Season, B=Manager, C=Finish, D-J=stats(7), K-P=cups(6), Q=Outcome
    for i, row in enumerate(SEED_ROWS, start=5):
        for col_idx, val in enumerate(row, start=1):
            if val != '':
                sr.cell(row=i, column=col_idx, value=val)
    wb.save(TEST)
    print(f"Seeded {len(SEED_ROWS)} rows to {TEST}")


def recalc() -> None:
    """Convert via LibreOffice headless to trigger full formula recalc."""
    soffice = r'C:\Program Files\LibreOffice\program\soffice.exe'
    if not Path(soffice).exists():
        soffice = 'soffice'
    cmd = [
        soffice, '--headless', '--calc',
        '--convert-to', 'xlsx',
        '--outdir', str(RECALC_DIR),
        str(TEST),
    ]
    print(f"Running: {' '.join(cmd)}")
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    print(result.stdout)
    if result.stderr:
        print("stderr:", result.stderr)
    if not TEST_RECALC.exists():
        raise RuntimeError(f"LibreOffice recalc output missing: {TEST_RECALC}")
    print(f"Recalc output: {TEST_RECALC}")


def verify() -> None:
    wb = load_workbook(TEST_RECALC, data_only=True)
    cs = wb['Career Summary']

    # Find the managers we seeded and check their computed values.
    # Row layout: 5=M-L1-01, 12=M-L1-12, 27=M-L2-14 (alphabetical in enum order? no,
    # declaration order from MANAGERS list: M-L1-01 row 5, M-L1-12 row 16, M-L2-14 row 32)
    # Let me look it up by scanning column A.
    mgr_row = {}
    for r in range(5, 5 + 27):
        name = cs.cell(row=r, column=1).value
        if name:
            mgr_row[name] = r

    print("\nManager rows:", mgr_row)

    def dump(name):
        r = mgr_row[name]
        hdrs = [cs.cell(row=4, column=c).value for c in range(1, 26)]
        vals = [cs.cell(row=r, column=c).value for c in range(1, 26)]
        print(f"\n=== {name} (row {r}) ===")
        for h, v in zip(hdrs, vals):
            print(f"  {h}: {v}")

    dump('M-L1-12')
    dump('M-L2-14')
    dump('M-L1-01')
    # Also dump one manager with zero seasons to confirm blanks
    for n, r in mgr_row.items():
        if n not in ('M-L1-12', 'M-L2-14', 'M-L1-01'):
            dump(n)
            break

    # Assertions for M-L1-12 (3 seasons)
    r = mgr_row['M-L1-12']
    seasons = cs.cell(row=r, column=5).value
    promos = cs.cell(row=r, column=6).value
    relegs = cs.cell(row=r, column=7).value
    avg = cs.cell(row=r, column=8).value
    best = cs.cell(row=r, column=9).value
    worst = cs.cell(row=r, column=10).value
    wins = cs.cell(row=r, column=11).value
    draws = cs.cell(row=r, column=12).value
    losses = cs.cell(row=r, column=13).value
    gf = cs.cell(row=r, column=14).value
    ga = cs.cell(row=r, column=15).value
    gd = cs.cell(row=r, column=16).value
    yc = cs.cell(row=r, column=17).value
    rc = cs.cell(row=r, column=18).value
    pl_titles = cs.cell(row=r, column=19).value
    cup_titles = cs.cell(row=r, column=20).value
    total_titles = cs.cell(row=r, column=25).value

    print("\n--- Assertions for M-L1-12 (3 seasons: finishes 3,1,8) ---")
    assert seasons == 3, f"seasons: {seasons}"
    assert promos == 0, f"promos: {promos}"
    assert relegs == 0, f"relegs: {relegs}"
    assert abs((avg or 0) - 4.0) < 0.05, f"avg: {avg}"
    assert best == 1, f"best: {best}"
    assert worst == 8, f"worst: {worst}"
    assert wins == 46, f"wins: {wins} (expect 16+20+10=46)"
    assert draws == 15, f"draws: {draws} (expect 6+3+6=15)"
    assert losses == 11, f"losses: {losses} (expect 2+1+8=11)"
    assert gf == 145, f"gf: {gf} (expect 48+62+35=145)"
    assert ga == 74, f"ga: {ga} (expect 22+18+34=74)"
    assert gd == 71, f"gd: {gd}"
    assert yc == 135, f"yc: {yc} (expect 45+38+52=135)"
    assert rc == 6, f"rc: {rc} (expect 2+1+3=6)"
    assert pl_titles == 1, f"pl_titles: {pl_titles} (season 1 PL Winner)"
    assert cup_titles == 1, f"cup_titles: {cup_titles} (season 2 Cup Winner)"
    assert total_titles == 2, f"total_titles: {total_titles}"
    print("  M-L1-12 OK")

    # Assertions for M-L2-14 (2 seasons, 1 promo, 1 releg)
    r = mgr_row['M-L2-14']
    seasons = cs.cell(row=r, column=5).value
    promos = cs.cell(row=r, column=6).value
    relegs = cs.cell(row=r, column=7).value
    best = cs.cell(row=r, column=9).value
    worst = cs.cell(row=r, column=10).value
    total_titles = cs.cell(row=r, column=25).value
    print("\n--- Assertions for M-L2-14 (2 seasons, promo then releg, 0 titles) ---")
    assert seasons == 2, f"seasons: {seasons}"
    assert promos == 1, f"promos: {promos}"
    assert relegs == 1, f"relegs: {relegs}"
    assert best == 1, f"best: {best}"
    assert worst == 12, f"worst: {worst}"
    assert total_titles == 0, f"total_titles: {total_titles}"
    print("  M-L2-14 OK")

    # Assertion for unseeded manager - should be blank or 0
    for n, r in mgr_row.items():
        if n not in ('M-L1-12', 'M-L2-14', 'M-L1-01'):
            seasons = cs.cell(row=r, column=5).value
            avg = cs.cell(row=r, column=8).value
            assert seasons == 0, f"unseeded {n} seasons: {seasons}"
            # Avg should be "" (blank) for zero-season managers
            assert avg in (None, ''), f"unseeded {n} avg: {avg!r}"
            print(f"  Zero-season {n}: seasons=0, avg=blank -- OK")
            break

    print("\nAll assertions passed.")


if __name__ == '__main__':
    seed()
    recalc()
    verify()
