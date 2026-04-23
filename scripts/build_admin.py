"""Build outputs/pes6_admin_tool.xlsx in a single pass.

Consolidates four earlier steps (build_admin_v2.py, add_remaining_sheets.py,
make_live_sheets.py, final_fix.py) into one script. All derived sheets
(Clubs, Nations, Team Composition) are written with live Excel formulas from
the start - no intermediate data_only reads, no LibreOffice recalc needed.

Usage:
    python scripts/build_admin.py
"""
from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except AttributeError:
    pass

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula

SCRIPTS_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPTS_DIR.parent
sys.path.insert(0, str(SCRIPTS_DIR))

from admin_tool_builder import (
    ABILITY_WEIGHTS, CSV_HEADERS, P, PLAYER_COLS, POSITIONS,
    ability_bonus, ability_count, attacking_prowess, b1, ball_winning,
    best_b1, best_position, cond_multiplier, defending_prowess,
    effective_b1, effective_club_formula, explosive_power, finishing,
    kicking_power, ovr_formula, player_class, pull_formula, speed,
)
from verify_source import verify_source

SRC = REPO_ROOT / 'data' / 'source.xlsx'
OUT = REPO_ROOT / 'outputs' / 'pes6_admin_tool.xlsx'

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

ROWS = 4900  # Players / CSV Paste row capacity

MIN_CLUB_POOL = 11
MIN_NATION_POOL = 16

# Tier thresholds for Top-18 Effective B1 (tighter than raw OVR because of
# ability and condition multipliers applied to Effective B1).
TIER_THRESHOLDS = [
    ('BANNED', 92.5),
    ('A', 89.0),
    ('B', 86.0),
    ('C', 84.0),
    ('D', 82.0),
    ('E', 79.5),
    # F: anything below
]

TIER_COLORS = {
    'BANNED': '000000', 'A': 'C00000', 'B': 'ED7D31', 'C': 'FFC000',
    'D': '70AD47', 'E': '5B9BD5', 'F': '808080',
}
TIER_TEXT = {
    'BANNED': 'FFFFFF', 'A': 'FFFFFF', 'B': 'FFFFFF', 'C': '000000',
    'D': 'FFFFFF', 'E': 'FFFFFF', 'F': 'FFFFFF',
}

# ---------------------------------------------------------------------------
# MANAGER ROSTER / COMPETITIONS / TIER RULES
# ---------------------------------------------------------------------------
# Single source of truth for the draft league. Edit these lists and rebuild
# when roster / competitions / rules change.

# (manager_name, team, league [1 or 2], initial_tier [A/B/C/D])
# Manager names are placeholders - replace with real names.
MANAGERS: list[tuple[str, str, int, str]] = [
    # Liga 1 (13 managers)
    ('M-L1-01', 'Aston Villa',          1, 'C'),
    ('M-L1-02', 'Como',                 1, 'C'),
    ('M-L1-03', 'Crvena Zvezda',        1, 'C'),
    ('M-L1-04', 'Crystal Palace',       1, 'C'),
    ('M-L1-05', 'Fulham',               1, 'C'),
    ('M-L1-06', 'Olympique Lyonnais',   1, 'C'),
    ('M-L1-07', "Borussia M'gladbach",  1, 'C'),
    ('M-L1-08', 'Fortuna D�sseldorf', 1, 'C'),
    ('M-L1-09', 'Roma',                 1, 'C'),
    ('M-L1-10', 'Slavia Praha',         1, 'C'),
    ('M-L1-11', 'Torino',               1, 'C'),
    ('M-L1-12', 'Bochum',               1, 'C'),
    ('M-L1-13', 'Villarreal',           1, 'C'),
    # Liga 2 (14 managers)
    ('M-L2-01', 'Milan',                2, 'C'),
    ('M-L2-02', 'Atalanta',             2, 'C'),
    ('M-L2-03', 'Athletic',             2, 'C'),
    ('M-L2-04', 'Bayer Leverkusen',     2, 'C'),
    ('M-L2-05', 'Leipzig',              2, 'C'),
    ('M-L2-06', 'Chelsea',              2, 'C'),
    ('M-L2-07', 'Fenerbahce',           2, 'C'),
    ('M-L2-08', 'Juventus',             2, 'C'),
    ('M-L2-09', 'Napoli',               2, 'C'),
    ('M-L2-10', 'Newcastle',            2, 'C'),
    ('M-L2-11', 'Nottingham Forest',    2, 'C'),
    ('M-L2-12', 'Olympique Marseille',  2, 'C'),
    ('M-L2-13', 'Betis',                2, 'C'),
    ('M-L2-14', 'Schalke',              2, 'C'),
]

# Derived from MANAGERS - used to highlight owned clubs in Clubs / Team Composition.
OWNED_TEAMS = {team for _, team, _, _ in MANAGERS}

# (competition_name, coefficient, active_this_season)
# Hierarchy: League (1.00) > PL / Cup (primary cups) > Europa / Conference
# (25% less than PL) > Friendly / SP-WC (25% less than Europa).
# Admins toggle Active per season when a competition is actually played.
COMPETITIONS: list[tuple[str, float, bool]] = [
    ('Liga 1',       1.00, True),
    ('Liga 2',       1.00, True),
    ('PL',           0.80, True),
    ('Cup',          0.70, True),
    ('Europa',       0.60, False),
    ('Conference',   0.60, False),
    ('Friendly Cup', 0.45, False),
    ('SP-World Cup', 0.45, False),
]

# Cup result vocab, best -> worst. Used as the dropdown list for Season Results
# result columns and for indexing relative achievement strength.
RESULT_LEVELS = ['Winner', 'Final', 'Semi', 'QF', 'R16', 'Group', 'DNQ']

TIERS = ['A', 'B', 'C', 'D']

# Outcome-based league finishing bands. Order here = worst-to-best index 0..3.
# (Band index 0 = best = Champion zone.)
BAND_LABELS = ['Champion', 'Playoff', 'Mid-safe', 'Playout']

# Tier -> suggested band index, per league. L2 is weaker overall so each tier's
# expected finish is bumped up (floored at Champion = 0).
#   L1: A -> Champion, B -> Playoff, C -> Mid-safe, D -> Playout (direct)
#   L2: A -> Champion, B -> Champion, C -> Playoff, D -> Mid-safe (shifted up 1)
TIER_BAND_BY_LEAGUE = {
    1: {'A': 0, 'B': 1, 'C': 2, 'D': 3},
    2: {'A': 0, 'B': 0, 'C': 1, 'D': 2},
}

# Tier -> minimum (1-based) RESULT_LEVELS index needed in a cup for that cup's
# coefficient to contribute to the bump score. Lower index = better result.
#   B: Winner only                 (index 1)
#   C: Semi or better              (index <= 3)
#   D: R16 or better               (index <= 5)
# Tier A managers cannot be bumped up (already at top).
CUP_BUMP_MAX_IDX_1B = {'B': 1, 'C': 3, 'D': 5}

# Cup bump triggers when SUM of eligible cups' coefficients >= this floor.
# 0.50 means: a single secondary cup (>=0.60) triggers a bump; two tertiary
# cups (0.45 + 0.45 = 0.90) trigger; one tertiary alone does not.
BUMP_SCORE_FLOOR = 0.50

# Season Results "Outcome" dropdown vocab. Auto tier shift:
#   Promoted  -> +1 tier shift (in addition to league delta + cup bump)
#   Relegated -> -1 tier shift
#   Stayed / blank -> no shift
OUTCOME_LEVELS = ['Stayed', 'Promoted', 'Relegated']
OUTCOME_SHIFT = {'Promoted': 1, 'Relegated': -1, 'Stayed': 0}

# Draft lottery weights - weight of getting #1 pick for the NEW tier after
# tier movement. Worse tier (D) has much higher weight (NBA lottery style).
LOTTERY_WEIGHTS = {'A': 1, 'B': 2, 'C': 4, 'D': 8}

# Blank rows to seed in Season Results for admin entry.
SEASON_RESULTS_BLANK_ROWS = 200

# Per-season team stats columns in Season Results, between League Finish and
# the cup columns. Integer values entered by admin each season. Career Summary
# aggregates these across all seasons.
SEASON_STATS_COLS = ['W', 'D', 'L', 'GF', 'GA', 'YC', 'RC']

# Blank rows to seed in Player Changes log for admin entry.
# Log data rows: 7 to 6 + PLAYER_CHANGES_LOG_ROWS. Must match the hardcoded
# range in effective_club_formula() in admin_tool_builder.py.
PLAYER_CHANGES_LOG_ROWS = 200
PLAYER_CHANGES_LOG_FIRST = 7
PLAYER_CHANGES_LOG_LAST = 6 + PLAYER_CHANGES_LOG_ROWS

# ---------------------------------------------------------------------------
# TACTICAL MENTOR CONFIG
# ---------------------------------------------------------------------------
# Formation catalog: each formation's 11 slot positions in order
# (GK, then defenders, then midfielders, then attackers).
FORMATIONS: dict[str, list[str]] = {
    '4-4-2':   ['GK','CB','CB','LB','RB','CMF','CMF','LMF','RMF','CF','CF'],
    '4-3-3':   ['GK','CB','CB','LB','RB','DMF','CMF','CMF','LWF','RWF','CF'],
    '4-5-1':   ['GK','CB','CB','LB','RB','DMF','CMF','CMF','LMF','RMF','CF'],
    '3-5-2':   ['GK','CWP','CB','CB','DMF','CMF','CMF','LMF','RMF','CF','CF'],
    '5-3-2':   ['GK','CWP','CB','CB','LWB','RWB','CMF','CMF','CMF','CF','CF'],
    '3-4-3':   ['GK','CWP','CB','CB','CMF','CMF','LMF','RMF','LWF','RWF','CF'],
    '4-2-3-1': ['GK','CB','CB','LB','RB','DMF','DMF','AMF','LMF','RMF','CF'],
}

# Each slot position -> which Players-sheet B1 column to score from.
POS_TO_B1COL: dict[str, str] = {
    'GK':  'GK B1',
    'CB':  'CB/CWP B1', 'CWP': 'CB/CWP B1',
    'LB':  'SB B1',     'RB':  'SB B1',
    'LWB': 'WB B1',     'RWB': 'WB B1',
    'DMF': 'DMF B1',
    'CMF': 'CMF B1',
    'LMF': 'SMF B1',    'RMF': 'SMF B1',
    'AMF': 'AMF B1',
    'LWF': 'WF B1',     'RWF': 'WF B1',
    'SS':  'SS B1',
    'CF':  'CF B1',
}

TACTICAL_MENTOR_SQUAD_ROWS = 30
TACTICAL_MENTOR_SUBS_ROWS = 5

# Full 16-position set admins can pick from in the Manual XI section.
POSITIONS_FULL = [
    'GK', 'CB', 'CWP', 'LB', 'RB', 'LWB', 'RWB',
    'DMF', 'CMF', 'LMF', 'RMF', 'AMF',
    'LWF', 'RWF', 'SS', 'CF',
]

# Manual XI position constraints (hard limits). Positions absent from the map
# have no per-position cap beyond the category cap.
POSITION_SINGLE_SLOT = {
    'GK', 'CWP', 'LB', 'RB', 'LWB', 'RWB',
    'LMF', 'RMF', 'LWF', 'RWF',
}
CATEGORY_DEFENDERS = {'CB', 'CWP', 'LB', 'RB', 'LWB', 'RWB'}
CATEGORY_MIDFIELDERS = {'DMF', 'CMF', 'LMF', 'RMF', 'AMF'}
CATEGORY_ATTACKERS = {'LWF', 'RWF', 'SS', 'CF'}
CATEGORY_MAX = 6

# PES6 team-style preset names (dropdown options for tactics selection).
TACTIC_PRESETS = [
    'Counter Attack',
    'Frontline Pressure',
    'Offside Trap',
    'Possession Game',
    'All-Out Attack',
    'All-Out Defence',
    'Side Attack',
    'Centre Attack',
    'Quick Passing',
    'Long Pass',
]

# Position category headers used by the B1 modifier table.
POSITION_CATEGORIES = [
    'GK', 'CB', 'LB/RB', 'LWB/RWB', 'DMF', 'CMF',
    'LMF/RMF', 'AMF', 'LWF/RWF', 'SS', 'CF',
]

# Map each of the 16 formation positions to its category column in the
# modifier table. (CB and CWP share "CB"; SB variants share "LB/RB" etc.)
POSITION_TO_CATEGORY = {
    'GK': 'GK',
    'CB': 'CB', 'CWP': 'CB',
    'LB': 'LB/RB', 'RB': 'LB/RB',
    'LWB': 'LWB/RWB', 'RWB': 'LWB/RWB',
    'DMF': 'DMF',
    'CMF': 'CMF',
    'LMF': 'LMF/RMF', 'RMF': 'LMF/RMF',
    'AMF': 'AMF',
    'LWF': 'LWF/RWF', 'RWF': 'LWF/RWF',
    'SS': 'SS',
    'CF': 'CF',
}

# Percent B1 modifier per (tactic, position category). Synthetic - no public
# PES6 formula exists - tuned to reward positions each tactic relies on and
# mildly penalise positions that become less useful. Admins can tune in the
# workbook's Tactic Modifier Table section; formulas read live from that
# table so edits take effect on recalc.
TACTIC_MODIFIERS: dict[str, dict[str, int]] = {
    'Counter Attack':     {'GK':0,'CB':-2,'LB/RB':0,'LWB/RWB':3,'DMF':-2,'CMF':-3,'LMF/RMF':4,'AMF':-2,'LWF/RWF':5,'SS':5,'CF':4},
    'Frontline Pressure': {'GK':0,'CB':2,'LB/RB':3,'LWB/RWB':4,'DMF':5,'CMF':4,'LMF/RMF':3,'AMF':2,'LWF/RWF':2,'SS':3,'CF':1},
    'Offside Trap':       {'GK':2,'CB':5,'LB/RB':3,'LWB/RWB':2,'DMF':0,'CMF':0,'LMF/RMF':0,'AMF':0,'LWF/RWF':0,'SS':0,'CF':0},
    'Possession Game':    {'GK':0,'CB':0,'LB/RB':0,'LWB/RWB':0,'DMF':2,'CMF':4,'LMF/RMF':2,'AMF':5,'LWF/RWF':2,'SS':2,'CF':1},
    'All-Out Attack':     {'GK':0,'CB':-3,'LB/RB':-2,'LWB/RWB':0,'DMF':-2,'CMF':-1,'LMF/RMF':2,'AMF':3,'LWF/RWF':4,'SS':5,'CF':5},
    'All-Out Defence':    {'GK':2,'CB':4,'LB/RB':3,'LWB/RWB':2,'DMF':3,'CMF':0,'LMF/RMF':-1,'AMF':-2,'LWF/RWF':-3,'SS':-4,'CF':-3},
    'Side Attack':        {'GK':0,'CB':0,'LB/RB':2,'LWB/RWB':3,'DMF':0,'CMF':0,'LMF/RMF':4,'AMF':1,'LWF/RWF':5,'SS':1,'CF':2},
    'Centre Attack':      {'GK':0,'CB':0,'LB/RB':0,'LWB/RWB':0,'DMF':1,'CMF':3,'LMF/RMF':1,'AMF':4,'LWF/RWF':1,'SS':4,'CF':4},
    'Quick Passing':      {'GK':0,'CB':0,'LB/RB':1,'LWB/RWB':1,'DMF':2,'CMF':3,'LMF/RMF':2,'AMF':3,'LWF/RWF':1,'SS':2,'CF':1},
    'Long Pass':          {'GK':1,'CB':2,'LB/RB':1,'LWB/RWB':0,'DMF':1,'CMF':1,'LMF/RMF':1,'AMF':1,'LWF/RWF':1,'SS':3,'CF':3},
}

# Slider modifier per tactic: each tactic shifts the recommended 0-20 slider
# value. Baseline is 10 (neutral); tactics add/subtract. Slider name -> tactic
# -> delta. Final slider = clamp(10 + sum of chosen-tactic deltas, 0, 20).
TACTIC_SLIDER_DELTAS: dict[str, dict[str, int]] = {
    'Defensive Line':  {'Frontline Pressure':6,'Offside Trap':5,'All-Out Attack':3,'All-Out Defence':-4,'Counter Attack':-5,'Possession Game':2},
    'Pressing':        {'Frontline Pressure':8,'All-Out Attack':3,'All-Out Defence':-4,'Counter Attack':-3,'Possession Game':1},
    'Compactness':     {'Offside Trap':4,'All-Out Defence':4,'Counter Attack':3,'Possession Game':2,'All-Out Attack':-3},
    'Width':           {'Side Attack':7,'Centre Attack':-7,'Quick Passing':-2,'Long Pass':2},
    'Player Support':  {'Possession Game':5,'Quick Passing':4,'Counter Attack':-4,'Long Pass':-3,'All-Out Attack':2},
    'Position Switch': {'Possession Game':4,'Centre Attack':2,'Side Attack':2,'All-Out Defence':-3},
}
TACTICAL_MENTOR_SUBS_MANUAL_ROWS = 7

# Which CSV column of `1. Paste cvs` holds what.
SRC_CLUB_COL = CSV_HEADERS.index('CLUB')          # 80 (0-indexed)
SRC_NATION_COL = CSV_HEADERS.index('NATIONALITY') # 77

# Source.xlsx's actual header for the club column (differs from our internal
# CSV_HEADERS). Used only when reading via pandas by column name.
SRC_CLUB_HEADER = 'CLUB TEAM'

# ---------------------------------------------------------------------------
# STYLING HELPERS
# ---------------------------------------------------------------------------

BOLD_W = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEAD_FILL = PatternFill('solid', start_color='1F4E78')
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
WRAP_TOP = Alignment(wrap_text=True, vertical='top')


def head_row(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_W
        cell.fill = HEAD_FILL
        cell.alignment = CENTER


# ---------------------------------------------------------------------------
# FORMULA HELPERS
# ---------------------------------------------------------------------------

def players_range(col_name: str) -> str:
    """Absolute reference to a column of the Players sheet, rows 2..ROWS+1."""
    col = P[col_name]
    return f"Players!${col}$2:${col}${ROWS + 1}"


def top18_formula(value_range: str, group_range: str, group_ref: str) -> str:
    """Average of the top-18 values in value_range where group_range = group_ref.

    Uses N(+range) to coerce empty-string cells to 0 - without this, LARGE and
    SUMPRODUCT return #VALUE! because the Players sheet leaves unused rows as
    empty strings (the ``IF(Name="","",...)`` pattern).
    """
    count = f"MIN(18,COUNTIF({group_range},{group_ref}))"
    large_sum = (
        f"SUMPRODUCT(LARGE(({group_range}={group_ref})"
        f"*N(+{value_range}),ROW($1:$18)))"
    )
    return f"=IFERROR(ROUND({large_sum}/{count},2),0)"


def tier_formula(value_cell: str) -> str:
    """Nested IF returning tier label (BANNED / A-F) from a numeric cell."""
    expr = f'"F"'
    for tier, cutoff in reversed(TIER_THRESHOLDS):
        expr = f'IF({value_cell}>={cutoff},"{tier}",{expr})'
    return f"={expr}"


def apply_tier_coloring(ws, cell_range: str) -> None:
    for tier, fill_color in TIER_COLORS.items():
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator='equal',
                formula=[f'"{tier}"'],
                fill=PatternFill('solid', start_color=fill_color),
                font=Font(bold=True, color=TIER_TEXT[tier]),
            ),
        )


def apply_effb1_colorscale(ws, cell_range: str) -> None:
    ws.conditional_formatting.add(
        cell_range,
        ColorScaleRule(
            start_type='min', start_color='F8D7DA',
            mid_type='percentile', mid_value=50, mid_color='FFFFFF',
            end_type='max', end_color='375623',
        ),
    )


def tier_legend_row(ws, row: int) -> None:
    ws.cell(row=row, column=1, value='Tiers:').font = Font(bold=True)
    legend = TIER_THRESHOLDS + [('F', None)]
    labels = {'BANNED': '>=92.5', 'A': '89.0+', 'B': '86.0+', 'C': '84.0+',
              'D': '82.0+', 'E': '79.5+', 'F': '<79.5'}
    for i, (tier, _) in enumerate(legend):
        cell = ws.cell(row=row, column=2 + i, value=f"{tier}: {labels[tier]}")
        cell.fill = PatternFill('solid', start_color=TIER_COLORS[tier])
        cell.font = Font(bold=True, color=TIER_TEXT[tier])
        cell.alignment = CENTER


# ---------------------------------------------------------------------------
# SHEET BUILDERS
# ---------------------------------------------------------------------------

def build_readme(ws) -> None:
    ws['A1'] = 'PES6 ADMIN TOOL'
    ws['A1'].font = Font(size=18, bold=True, color='1F4E78')
    ws.merge_cells('A1:C1')

    sections = [
        ('', ''),
        ('WORKFLOW', 'h'),
        ('Step 1:', 'Export player data from PES Fan Editor as CSV.'),
        ('Step 2:', 'Open this workbook, go to sheet "CSV Paste".'),
        ('Step 3:', 'Delete old data (keep row 1 headers). Paste new CSV at A2.'),
        ('Step 4:', 'Press F9 to recalculate. All other sheets refresh automatically.'),
        ('', ''),
        ('KEY METRICS', 'h'),
        ('OVR', 'Max B1 across all 11 positions. Official overall rating.'),
        ('B1 per position', "Position-specific rating using the patch-maker's exact formula."),
        ('Ability Bonus', 'Multiplicative bonus to Best B1 based on special abilities, weighted per position. Max ~15%.'),
        ('COND Multiplier', 'COND 7 -> x1.06, 6 -> x1.03, 5 -> x1.00, 4 -> x0.98, 3 -> x0.96, <=2 -> x0.94.'),
        ('Effective B1', 'Best B1 * (1 + Ability Bonus) * COND Multiplier. The true player-value metric.'),
        ('Class', 'A (>=90), B (84-89), C (78-83), D (<78), based on OVR.'),
        ('Tier (teams)', 'A-F, based on top-18 average Effective B1.'),
        ('', ''),
        ('SHEETS', 'h'),
        ('1. CSV Paste', 'Raw CSV data. Paste new patch exports here.'),
        ('2. Players', 'Per-player analysis (formulas pulling from CSV Paste).'),
        ('3. Clubs', 'Club rankings and tiers (live formulas).'),
        ('4. Nations', 'National team rankings (live formulas).'),
        ('5. Team Composition', 'Class distribution (A/B/C/D counts) per club (live formulas).'),
        ('6. Draft Rules', 'Reference sheet: class/group/tier rules.'),
        ('7. Managers', 'Roster: manager <-> team <-> league <-> initial tier. Source of truth.'),
        ('8. Competitions', 'League, PL, Cup, Europa/Conference, Friendly, SP-World Cup with importance.'),
        ('9. Expectations', 'Per-manager expected league finish (quartile), admin-overridable.'),
        ('10. Season Results', 'ADMIN DATA ENTRY: one row per manager per season. League Finish + per-season team stats (W / D / L / GF / GA / Yellow / Red) + cup results + Outcome. Append-only; feeds Tier Movement and Career Summary.'),
        ('11. Tier Movement', 'Auto-computed: compares actual vs expected, outputs new tier for next draft.'),
        ('12. Draft Order', 'NBA-style lottery weights by new tier. Admin runs draw, fills Actual Pick.'),
        ('13. Career Summary', 'Multi-season aggregates per manager (seasons, promotions, avg/best/worst finish, totals for W/D/L/GF/GA/GD/YC/RC, titles per cup). All formula-driven from Season Results.'),
        ('14. Player Changes', 'ADMIN OVERRIDE: stage transfers / drafts / arrangements. Apply Changes toggle makes Clubs and Team Composition reflect the new rosters.'),
        ('15. Tactical Mentor', 'Dynamic scouting + planner. Top section: type team + formation, see squad, greedy Best XI, subs, tactical slider suggestions. Lower section: manually pick your XI (player + PES6 position each), choose up to 4 tactic presets, get PES6-style 0-20 slider values and a per-player B1 efficiency panel.'),
        ('', ''),
        ('TIER MOVEMENT LOGIC', 'h'),
        ('Bands', 'Outcome-based: Champion (top 3), Playoff (4-8), Mid-safe (middle), Playout (bottom 4).'),
        ('Expected (L1)', 'Tier A -> Champion, B -> Playoff, C -> Mid-safe, D -> Playout.'),
        ('Expected (L2)', 'One band stricter: A -> Champion, B -> Champion, C -> Playoff, D -> Mid-safe.'),
        ('League Delta', 'Actual Band vs Expected: better = +1, worse = -1, same = 0.'),
        ('Cup Bump', '+1 if SUM of eligible-cup coefficients >= 0.50 (eligible = Active AND result meets tier threshold: B=Winner, C=Semi+, D=R16+). A: no bump.'),
        ('Coefficients', 'League 1.00, PL 0.80, Cup 0.70, Europa / Conference 0.60, Friendly / SP-World Cup 0.45. Admin-editable in Competitions sheet.'),
        ('Anti-gaming', 'Cup bump is ignored when league delta is negative (winning a friendly does not save a tanked league).'),
        ('Outcome Shift', '+1 if Promoted, -1 if Relegated, 0 if Stayed. Entered by admin in Season Results after playoff / playout KOs resolve.'),
        ('Final Delta', 'League component capped at +-1; Outcome Shift adds on top. Total range: -2 (tanked and relegated) to +2 (dominant and promoted).'),
        ('', ''),
        ('ABILITY WEIGHTS PER POSITION', 'h'),
    ]
    r = 2
    for label, val in sections:
        if val == 'h':
            cell = ws.cell(row=r, column=1, value=label)
            cell.font = Font(size=12, bold=True, color='1F4E78')
            cell.fill = PatternFill('solid', start_color='DDEBF7')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        elif label or val:
            if label:
                ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            if val:
                ws.cell(row=r, column=2, value=val).alignment = WRAP_TOP
        r += 1

    ws.cell(row=r, column=1, value='Position').font = Font(bold=True)
    ws.cell(row=r, column=2, value='Top weighted abilities').font = Font(bold=True)
    head_row(ws, r, 2)
    r += 1
    for pos in POSITIONS:
        weights = sorted(ABILITY_WEIGHTS.get(pos, {}).items(), key=lambda kv: -kv[1])[:6]
        ws.cell(row=r, column=1, value=pos).font = Font(bold=True)
        ws.cell(row=r, column=2, value=', '.join(f"{k} ({v:.3f})" for k, v in weights))
        r += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 110


def build_csv_paste(ws, df_source: pd.DataFrame) -> None:
    for i, header in enumerate(CSV_HEADERS, 1):
        ws.cell(row=1, column=i, value=header)
    head_row(ws, 1, len(CSV_HEADERS))

    for row_idx, row in enumerate(df_source.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            if col_idx > len(CSV_HEADERS):
                break
            if pd.notna(val):
                if hasattr(val, 'item'):
                    val = val.item()
                ws.cell(row=row_idx, column=col_idx, value=val)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(CSV_HEADERS))}1"
    ws.freeze_panes = 'B2'


def build_players(ws) -> None:
    # Header row
    for i, header in enumerate(PLAYER_COLS, 1):
        ws.cell(row=1, column=i, value=header)
    head_row(ws, 1, len(PLAYER_COLS))

    # Mapping from Players column name to CSV_HEADERS source column.
    csv_to_p_map = {
        'Name': 'NAME', 'Nationality': 'NATIONALITY', 'Club': 'CLUB',
        'Age': 'AGE', 'Foot': 'STRONG FOOT', 'Registered Pos': 'REGISTERED POSITION',
        'ATTACK': 'ATTACK', 'DEFENSE': 'DEFENSE', 'BALANCE': 'BALANCE',
        'STAMINA': 'STAMINA', 'TOP SPEED': 'TOP SPEED', 'ACCELERATION': 'ACCELERATION',
        'RESPONSE': 'RESPONSE', 'AGILITY': 'AGILITY',
        'DRIBBLE ACC': 'DRIBBLE ACCURACY', 'DRIBBLE SPD': 'DRIBBLE SPEED',
        'SHORT PASS ACC': 'SHORT PASS ACCURACY', 'SHORT PASS SPD': 'SHORT PASS SPEED',
        'LONG PASS ACC': 'LONG PASS ACCURACY', 'LONG PASS SPD': 'LONG PASS SPEED',
        'SHOT ACC': 'SHOT ACCURACY', 'SHOT PWR': 'SHOT POWER', 'SHOT TEC': 'SHOT TECHNIQUE',
        'FREE KICK': 'FREE KICK ACCURACY', 'SWERVE': 'SWERVE', 'HEADING': 'HEADING',
        'JUMP': 'JUMP', 'TECHNIQUE': 'TECHNIQUE', 'AGGRESSION': 'AGGRESSION',
        'MENTALITY': 'MENTALITY', 'GOAL KEEPING': 'GOAL KEEPING',
        'TEAM WORK': 'TEAM WORK', 'CONDITION': 'CONDITION / FITNESS',
        'REACTION': 'REACTION', 'PLAYMAKING': 'PLAYMAKING', 'PASSING': 'PASSING',
        'SCORING': 'SCORING', '1-1 SCORING': '1-1 SCORING', 'POST PLAYER': 'POST PLAYER',
        'LINES': 'LINES', 'MIDDLE SHOOTING': 'MIDDLE SHOOTING', 'SIDE': 'SIDE',
        'CENTRE': 'CENTRE', 'PENALTIES': 'PENALTIES', '1-TOUCH PASS': '1-TOUCH PASS',
        'OUTSIDE': 'OUTSIDE', 'MARKING': 'MARKING', 'SLIDING': 'SLIDING',
        'COVERING': 'COVERING', 'D-LINE CONTROL': 'D-LINE CONTROL',
        'PENALTY STOPPER': 'PENALTY STOPPER', '1-ON-1 STOPPER': '1-ON-1 STOPPER',
        'LONG THROW': 'LONG THROW',
    }
    composite_fn = {
        'Attacking Prowess': attacking_prowess, 'Finishing': finishing,
        'Speed': speed, 'Explosive Power': explosive_power,
        'Kicking Power': kicking_power, 'Defending Prowess': defending_prowess,
        'Ball Winning': ball_winning,
    }
    b1_map = {f'{pos} B1': pos for pos in POSITIONS}

    for row_idx in range(2, ROWS + 2):
        for pcol_name in PLAYER_COLS:
            col = P[pcol_name]
            cell = ws[f"{col}{row_idx}"]
            if pcol_name in csv_to_p_map:
                cell.value = pull_formula(pcol_name, csv_to_p_map[pcol_name], row_idx)
            elif pcol_name in composite_fn:
                cell.value = composite_fn[pcol_name](row_idx)
            elif pcol_name in b1_map:
                cell.value = b1(b1_map[pcol_name], row_idx)
            elif pcol_name == 'OVR':
                cell.value = ovr_formula(row_idx)
            elif pcol_name == 'Best B1':
                cell.value = best_b1(row_idx)
            elif pcol_name == 'Best Position':
                cell.value = best_position(row_idx)
            elif pcol_name == 'Ability ★':
                cell.value = ability_count(row_idx)
            elif pcol_name == 'Ability Bonus':
                cell.value = ability_bonus(row_idx)
            elif pcol_name == 'COND Multiplier':
                cell.value = cond_multiplier(row_idx)
            elif pcol_name == 'Effective B1':
                cell.value = effective_b1(row_idx)
            elif pcol_name == 'Class':
                cell.value = player_class(row_idx)
            elif pcol_name == 'Effective Club':
                cell.value = effective_club_formula(
                    row_idx,
                    PLAYER_CHANGES_LOG_FIRST,
                    PLAYER_CHANGES_LOG_LAST,
                )
        if row_idx % 1000 == 0:
            print(f"  Players rows: {row_idx}/{ROWS + 1}")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(PLAYER_COLS))}1"
    ws.freeze_panes = 'B2'

    # Number formats
    for col_name in ['OVR', 'Best B1', 'Effective B1'] + [f'{p} B1' for p in POSITIONS]:
        col = P[col_name]
        for row in range(2, ROWS + 2):
            ws[f"{col}{row}"].number_format = '0.00'
    for col_name, fmt in [('Ability Bonus', '0.000'), ('COND Multiplier', '0.00')]:
        col = P[col_name]
        for row in range(2, ROWS + 2):
            ws[f"{col}{row}"].number_format = fmt

    col_widths = {
        'Name': 22, 'Nationality': 14, 'Club': 22, 'Effective Club': 22,
        'Registered Pos': 10, 'OVR': 7, 'Class': 6, 'Best B1': 9,
        'Effective B1': 10, 'Best Position': 11, 'Ability ★': 7,
        'CONDITION': 9, 'Ability Bonus': 10, 'COND Multiplier': 10,
    }
    for name, w in col_widths.items():
        ws.column_dimensions[P[name]].width = w


def build_clubs(ws, clubs: list[str]) -> None:
    ws['A1'] = 'CLUB RANKINGS - top-18 Effective B1, auto-refreshes from Players'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:K1')
    ws['A2'] = ('All values are live Excel formulas. Groups players by their '
                'Effective Club (overridden via Player Changes when Apply '
                'Changes is YES). Press F9 to recalculate after edits.')
    ws['A2'].font = Font(italic=True, color='808080')

    tier_legend_row(ws, 3)

    headers = ['Rank', 'Club', 'Tier', 'Top-18 Eff B1', 'Top-18 OVR', 'Pool',
               'A (#)', 'B (#)', 'C (#)', 'D (#)', 'Owned']
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)
    head_row(ws, 5, len(headers))

    club_rng = players_range('Effective Club')
    effb1_rng = players_range('Effective B1')
    ovr_rng = players_range('OVR')
    class_rng = players_range('Class')

    first_row, last_row = 6, 5 + len(clubs)
    rank_range = f"$D${first_row}:$D${last_row}"

    for idx, club in enumerate(clubs):
        row = first_row + idx
        club_ref = f'B{row}'

        ws.cell(row=row, column=1, value=f'=RANK(D{row},{rank_range})').alignment = CENTER
        club_cell = ws.cell(row=row, column=2, value=club)
        club_cell.alignment = LEFT
        if club in OWNED_TEAMS:
            club_cell.font = Font(bold=True)
            club_cell.fill = PatternFill('solid', start_color='FFF2CC')

        ws.cell(row=row, column=3, value=tier_formula(f'D{row}')).alignment = CENTER
        v = ws.cell(row=row, column=4, value=top18_formula(effb1_rng, club_rng, club_ref))
        v.number_format = '0.00'
        v.alignment = CENTER
        v = ws.cell(row=row, column=5, value=top18_formula(ovr_rng, club_rng, club_ref))
        v.number_format = '0.00'
        v.alignment = CENTER
        ws.cell(row=row, column=6, value=f'=COUNTIF({club_rng},{club_ref})').alignment = CENTER
        for i, cls in enumerate(['A', 'B', 'C', 'D']):
            ws.cell(row=row, column=7 + i,
                    value=f'=COUNTIFS({club_rng},{club_ref},{class_rng},"{cls}")').alignment = CENTER
        ws.cell(row=row, column=11,
                value='YES' if club in OWNED_TEAMS else '').alignment = CENTER

    apply_tier_coloring(ws, f'C{first_row}:C{last_row}')
    apply_effb1_colorscale(ws, f'D{first_row}:D{last_row}')

    ws.auto_filter.ref = f"A5:{get_column_letter(len(headers))}{last_row}"
    ws.freeze_panes = 'C6'
    for i, w in enumerate([6, 28, 9, 14, 14, 7, 7, 7, 7, 7, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_nations(ws, nations: list[str]) -> None:
    ws['A1'] = 'NATIONAL TEAM RANKINGS - top-18 Effective B1 (>=16 players)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:G1')

    tier_legend_row(ws, 3)

    headers = ['Rank', 'Nation', 'Tier', 'Top-18 Eff B1', 'Top-18 OVR', 'Pool']
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)
    head_row(ws, 5, len(headers))

    nat_rng = players_range('Nationality')
    effb1_rng = players_range('Effective B1')
    ovr_rng = players_range('OVR')

    first_row, last_row = 6, 5 + len(nations)
    rank_range = f"$D${first_row}:$D${last_row}"

    for idx, nat in enumerate(nations):
        row = first_row + idx
        nat_ref = f'B{row}'

        ws.cell(row=row, column=1, value=f'=RANK(D{row},{rank_range})').alignment = CENTER
        ws.cell(row=row, column=2, value=nat).alignment = LEFT
        ws.cell(row=row, column=3, value=tier_formula(f'D{row}')).alignment = CENTER
        v = ws.cell(row=row, column=4, value=top18_formula(effb1_rng, nat_rng, nat_ref))
        v.number_format = '0.00'
        v.alignment = CENTER
        v = ws.cell(row=row, column=5, value=top18_formula(ovr_rng, nat_rng, nat_ref))
        v.number_format = '0.00'
        v.alignment = CENTER
        ws.cell(row=row, column=6, value=f'=COUNTIF({nat_rng},{nat_ref})').alignment = CENTER

    apply_tier_coloring(ws, f'C{first_row}:C{last_row}')
    apply_effb1_colorscale(ws, f'D{first_row}:D{last_row}')

    ws.auto_filter.ref = f"A5:{get_column_letter(len(headers))}{last_row}"
    ws.freeze_panes = 'C6'
    for i, w in enumerate([6, 26, 9, 14, 14, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_team_composition(ws, clubs: list[str]) -> None:
    ws['A1'] = 'TEAM COMPOSITION - class breakdown per club (auto-refresh)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:L1')
    ws['A2'] = ("For admin use: audit draft picks, verify manager A-caps, "
                "edit teams after draft. Groups players by Effective Club "
                "(reflects Player Changes when Apply Changes is YES).")
    ws['A2'].font = Font(italic=True, color='808080')

    headers = ['Rank', 'Club', 'Tier', 'Top-18 Eff B1', 'Total',
               'A (90+)', 'B (84-89)', 'C (78-83)', 'D (<78)', 'A%', 'B%', 'Owned']
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    head_row(ws, 4, len(headers))

    club_rng = players_range('Effective Club')
    effb1_rng = players_range('Effective B1')
    class_rng = players_range('Class')

    first_row, last_row = 5, 4 + len(clubs)
    rank_range = f"$D${first_row}:$D${last_row}"

    for idx, club in enumerate(clubs):
        row = first_row + idx
        club_ref = f'B{row}'

        ws.cell(row=row, column=1, value=f'=RANK(D{row},{rank_range})').alignment = CENTER
        club_cell = ws.cell(row=row, column=2, value=club)
        if club in OWNED_TEAMS:
            club_cell.font = Font(bold=True)
            club_cell.fill = PatternFill('solid', start_color='FFF2CC')
        ws.cell(row=row, column=3, value=tier_formula(f'D{row}')).alignment = CENTER
        v = ws.cell(row=row, column=4, value=top18_formula(effb1_rng, club_rng, club_ref))
        v.number_format = '0.00'
        v.alignment = CENTER
        ws.cell(row=row, column=5, value=f'=COUNTIF({club_rng},{club_ref})').alignment = CENTER
        for i, cls in enumerate(['A', 'B', 'C', 'D']):
            ws.cell(row=row, column=6 + i,
                    value=f'=COUNTIFS({club_rng},{club_ref},{class_rng},"{cls}")').alignment = CENTER
        ws.cell(row=row, column=10, value=f'=IFERROR(F{row}/E{row},0)').number_format = '0%'
        ws.cell(row=row, column=10).alignment = CENTER
        ws.cell(row=row, column=11, value=f'=IFERROR(G{row}/E{row},0)').number_format = '0%'
        ws.cell(row=row, column=11).alignment = CENTER
        ws.cell(row=row, column=12,
                value='YES' if club in OWNED_TEAMS else '').alignment = CENTER

    apply_tier_coloring(ws, f'C{first_row}:C{last_row}')
    # Highlight A-class counts >= 3 (admin attention)
    ws.conditional_formatting.add(
        f'F{first_row}:F{last_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=['3'],
                   fill=PatternFill('solid', start_color='FFE699'),
                   font=Font(bold=True, color='C00000')),
    )

    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last_row}"
    ws.freeze_panes = 'C5'
    for i, w in enumerate([6, 28, 9, 14, 8, 9, 10, 10, 9, 7, 7, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_draft_rules(ws) -> None:
    ws['A1'] = 'DRAFT RULES & SYSTEM REFERENCE'
    ws['A1'].font = Font(bold=True, size=16, color='1F4E78')
    ws.merge_cells('A1:D1')

    rules = [
        ('', ''),
        ('PLAYER CLASSES (based on OVR)', 'h'),
        ('A', 'OVR 90-94'),
        ('B', 'OVR 84-89'),
        ('C', 'OVR 78-83'),
        ('D', 'OVR <= 77'),
        ('Banned', 'OVR >= 95 (cannot be drafted)'),
        ('', ''),
        ('MANAGER GROUPS', 'h'),
        ('Group A (top)', 'Max 1 A lifetime | Per season: 1A + 2B + 1C + 1D'),
        ('Group B (mid-strong)', 'Max 2 A lifetime | Per season: 2A + 1B + 1C + 1D'),
        ('Group C (mid)', 'Max 3 A lifetime | Per season: 2A + 2B + 1 (C or D)'),
        ('Group D (new)', 'Max 3 A lifetime | Per season: 2A + 3 (B/C/D free)'),
        ('', ''),
        ('POSITION RULE', 'h'),
        ('', 'No 2 A-class players in the same color position '
             '(e.g., not 2 A-class CFs).'),
        ('', 'Color groups: BLUE = defenders, GREEN = midfielders, '
             'RED = attackers, GK.'),
        ('', ''),
        ('BONUSES FOR WEAK TEAMS', 'h'),
        ('Tier F team', '+1 A + 1 B bonus player'),
        ('Tier E team', '+2 B bonus players'),
        ('', ''),
        ('PROMOTION / RELEGATION', 'h'),
        ('', 'After each season, managers move between groups by success. '
             'Best -> harder group.'),
        ('', "If a manager's A-count exceeds their new group's cap, "
             'swap an A -> B at season end.'),
        ('', ''),
        ('TEAM TIER THRESHOLDS (Top-18 Effective B1)', 'h'),
        ('BANNED', '>= 92.5'),
        ('A', '>= 89.0'),
        ('B', '>= 86.0'),
        ('C', '>= 84.0'),
        ('D', '>= 82.0'),
        ('E', '>= 79.5'),
        ('F', '< 79.5'),
    ]

    r = 2
    for label, val in rules:
        if val == 'h':
            cell = ws.cell(row=r, column=1, value=label)
            cell.font = Font(size=12, bold=True, color='1F4E78')
            cell.fill = PatternFill('solid', start_color='DDEBF7')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        elif label or val:
            if label:
                ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            if val:
                ws.cell(row=r, column=2, value=val).alignment = WRAP_TOP
        r += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 90


def build_managers(ws) -> None:
    ws['A1'] = 'MANAGER ROSTER'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:E1')
    ws['A2'] = ("Source of truth for manager <-> team assignments. "
                "To change roster, edit the MANAGERS constant in "
                "scripts/build_admin.py and rebuild.")
    ws['A2'].font = Font(italic=True, color='808080')
    ws.merge_cells('A2:E2')

    headers = ['Manager', 'Team', 'League', 'Initial Tier', 'Active?']
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    head_row(ws, 4, len(headers))

    first_row = 5
    last_row = 4 + len(MANAGERS)
    for idx, (name, team, league, tier) in enumerate(MANAGERS):
        row = first_row + idx
        ws.cell(row=row, column=1, value=name).alignment = LEFT
        team_cell = ws.cell(row=row, column=2, value=team)
        team_cell.alignment = LEFT
        team_cell.font = Font(bold=True)
        team_cell.fill = PatternFill('solid', start_color='FFF2CC')
        ws.cell(row=row, column=3, value=league).alignment = CENTER
        ws.cell(row=row, column=4, value=tier).alignment = CENTER
        ws.cell(row=row, column=5, value='YES').alignment = CENTER

    apply_tier_coloring(ws, f'D{first_row}:D{last_row}')

    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last_row}"
    ws.freeze_panes = 'A5'
    for col, width in enumerate([14, 28, 8, 14, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_competitions(ws) -> None:
    ws['A1'] = 'COMPETITIONS'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:C1')
    ws['A2'] = ("Coefficient is each cup's weight in the tier-movement cup-bump "
                "score. Set Active to YES when the competition is played this "
                "season - inactive cups contribute 0 to the bump regardless of "
                "result. League rows are here for reference; they drive tier "
                "movement directly, not via the cup-bump score.")
    ws['A2'].font = Font(italic=True, color='808080')
    ws.merge_cells('A2:C2')

    headers = ['Competition', 'Coefficient', 'Active This Season?']
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    head_row(ws, 4, len(headers))

    for idx, (name, coef, active) in enumerate(COMPETITIONS):
        row = 5 + idx
        ws.cell(row=row, column=1, value=name).alignment = LEFT
        c = ws.cell(row=row, column=2, value=coef)
        c.alignment = CENTER
        c.number_format = '0.00'
        ws.cell(row=row, column=3,
                value='YES' if active else 'NO').alignment = CENTER

    # Active-flag dropdown so admins can toggle without typos.
    active_dv = DataValidation(type='list', formula1='"YES,NO"', allow_blank=False)
    ws.add_data_validation(active_dv)
    active_dv.add(f'C5:C{4 + len(COMPETITIONS)}')

    ws.freeze_panes = 'A5'
    for col, width in enumerate([18, 14, 24], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_expectations(ws) -> None:
    ws['A1'] = 'EXPECTATIONS - per-manager expected league finish'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:F1')
    ws['A2'] = ("Bands are outcome-based: Champion (top 3), Playoff (4-8), "
                "Mid-safe (middle), Playout (bottom 4). Suggested Band uses "
                "both Tier AND League - L2 expectations are one band stricter "
                "than L1 at the same tier (since L2 is overall weaker). "
                "Override Band is admin-typed. Effective Band = Override if "
                "set, otherwise Suggested.")
    ws['A2'].font = Font(italic=True, color='808080')
    ws.merge_cells('A2:F2')

    headers = ['Manager', 'Current Tier', 'Current League',
               'Suggested Band', 'Override Band', 'Effective Band']
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    head_row(ws, 4, len(headers))

    n_mgr = len(MANAGERS)
    mgr_first = 5
    mgr_last = 4 + n_mgr
    mgrs_name_rng = f'Managers!$A$5:$A${mgr_last}'
    mgrs_tier_rng = f'Managers!$D$5:$D${mgr_last}'
    mgrs_league_rng = f'Managers!$C$5:$C${mgr_last}'

    # Build the suggested-band IF chain from TIER_BAND_BY_LEAGUE.
    # Outer: IF(league=1, <L1 block>, IF(league=2, <L2 block>, ""))
    # Each block: IF(tier="A", label_A, IF(tier="B", ...))
    def tier_to_label_chain(tier_cell: str, league: int) -> str:
        mapping = TIER_BAND_BY_LEAGUE[league]
        expr = '""'
        for t in reversed(TIERS):
            label = BAND_LABELS[mapping[t]]
            expr = f'IF({tier_cell}="{t}","{label}",{expr})'
        return expr

    for idx, (name, _team, _lg, _tier) in enumerate(MANAGERS):
        row = mgr_first + idx
        mgr_ref = f'A{row}'
        tier_ref = f'B{row}'
        league_ref = f'C{row}'
        sug_ref = f'D{row}'
        ovr_ref = f'E{row}'

        ws.cell(row=row, column=1, value=name).alignment = LEFT

        # Current Tier from Managers
        ws.cell(row=row, column=2, value=(
            f'=IFERROR(INDEX({mgrs_tier_rng},'
            f'MATCH({mgr_ref},{mgrs_name_rng},0)),"")'
        )).alignment = CENTER

        # Current League from Managers
        ws.cell(row=row, column=3, value=(
            f'=IFERROR(INDEX({mgrs_league_rng},'
            f'MATCH({mgr_ref},{mgrs_name_rng},0)),"")'
        )).alignment = CENTER

        # Suggested band: depends on both tier and league
        l1_chain = tier_to_label_chain(tier_ref, 1)
        l2_chain = tier_to_label_chain(tier_ref, 2)
        ws.cell(row=row, column=4, value=(
            f'=IF({league_ref}=1,{l1_chain},'
            f'IF({league_ref}=2,{l2_chain},""))'
        )).alignment = CENTER

        # Override: blank for admin input
        ws.cell(row=row, column=5, value=None).alignment = CENTER

        # Effective
        ws.cell(row=row, column=6, value=(
            f'=IF({ovr_ref}="",{sug_ref},{ovr_ref})'
        )).alignment = CENTER

    # Override accepts BAND_LABELS only
    band_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(BAND_LABELS) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(band_dv)
    band_dv.add(f'E{mgr_first}:E{mgr_last}')

    apply_tier_coloring(ws, f'B{mgr_first}:B{mgr_last}')

    ws.freeze_panes = 'A5'
    for col, width in enumerate([14, 13, 14, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def build_season_results(ws) -> None:
    ws['A1'] = 'SEASON RESULTS - admin data entry (append-only log)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:K1')
    ws['A2'] = ("One row per manager per season. League Finish = final rank "
                "after playoff/playout. Per-season team stats (W/D/L/GF/GA/YC/"
                "RC) are integers entered by admin; they feed Career Summary. "
                "Cup columns use the dropdown list. Outcome = Promoted / "
                "Relegated / Stayed (after playoff / playout KO resolves). "
                "Tier Movement auto-uses the LATEST Season for each manager.")
    ws['A2'].font = Font(italic=True, color='808080')
    ws.merge_cells('A2:K2')

    cup_names = [c for c, _, _ in COMPETITIONS if c not in ('Liga 1', 'Liga 2')]
    # Layout: Season | Manager | League Finish | <stats cols> | <cup cols> | Outcome
    headers = (
        ['Season', 'Manager', 'League Finish']
        + list(SEASON_STATS_COLS)
        + cup_names
        + ['Outcome']
    )
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    head_row(ws, 4, len(headers))

    data_first = 5
    data_last = 4 + SEASON_RESULTS_BLANK_ROWS
    mgr_last = 4 + len(MANAGERS)
    stats_count = len(SEASON_STATS_COLS)
    first_stats_col = 4                                  # col D
    first_cup_col = first_stats_col + stats_count        # col K (first cup)
    outcome_col = first_cup_col + len(cup_names)         # column index for Outcome

    # Manager dropdown from Managers sheet
    mgr_dv = DataValidation(
        type='list',
        formula1=f'=Managers!$A$5:$A${mgr_last}',
        allow_blank=True,
    )
    ws.add_data_validation(mgr_dv)
    mgr_dv.add(f'B{data_first}:B{data_last}')

    # Cup result dropdown from RESULT_LEVELS
    result_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(RESULT_LEVELS) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(result_dv)
    for ci in range(first_cup_col, first_cup_col + len(cup_names)):
        col = get_column_letter(ci)
        result_dv.add(f'{col}{data_first}:{col}{data_last}')

    # Outcome dropdown
    outcome_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(OUTCOME_LEVELS) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(outcome_dv)
    out_letter = get_column_letter(outcome_col)
    outcome_dv.add(f'{out_letter}{data_first}:{out_letter}{data_last}')

    # Integer number format for stats cells
    for ci in range(first_stats_col, first_stats_col + stats_count):
        col = get_column_letter(ci)
        for r in range(data_first, data_last + 1):
            ws.cell(row=r, column=ci).number_format = '0'

    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{data_last}"
    ws.freeze_panes = 'C5'
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 13
    for ci in range(first_stats_col, first_stats_col + stats_count):
        ws.column_dimensions[get_column_letter(ci)].width = 6
    for ci in range(first_cup_col, first_cup_col + len(cup_names)):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.column_dimensions[out_letter].width = 12


def build_tier_movement(ws) -> None:
    ws['A1'] = 'TIER MOVEMENT - computed from latest Season Results'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:M1')
    ws['A2'] = ("Uses each manager's MOST RECENT season in Season Results. "
                "League Delta: +1 if Actual Band better than Expected, -1 if "
                "worse, 0 if same. Cup Bump: +1 if SUM of eligible-cup "
                "coefficients >= 0.5 (eligible = Active AND result reaches the "
                "tier threshold: B=Winner, C=Semi+, D=R16+). Cup bump is "
                "suppressed when league delta is negative. League component "
                "capped at +-1. Outcome Shift adds +1 for Promoted / -1 for "
                "Relegated on top of league component; combined Final Delta "
                "can be +-2 in promotion/relegation cases.")
    ws['A2'].font = Font(italic=True, color='808080')
    ws.merge_cells('A2:M2')

    headers = ['Manager', 'League', 'Current Tier', 'Latest Season',
               'League Finish', 'League Size', 'Actual Band', 'Expected Band',
               'League Delta', 'Cup Bump', 'Outcome Shift',
               'Final Delta', 'New Tier']
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    head_row(ws, 4, len(headers))

    n_mgr = len(MANAGERS)
    mgr_first = 5
    mgr_last = 4 + n_mgr
    mgrs_name_rng = f'Managers!$A$5:$A${mgr_last}'
    mgrs_tier_rng = f'Managers!$D$5:$D${mgr_last}'
    mgrs_league_rng = f'Managers!$C$5:$C${mgr_last}'
    exp_name_rng = f'Expectations!$A$5:$A${mgr_last}'
    exp_eff_rng = f'Expectations!$F$5:$F${mgr_last}'  # Effective Band is col F now

    sr_first = 5
    sr_last = 4 + SEASON_RESULTS_BLANK_ROWS
    sr_season = f"'Season Results'!$A${sr_first}:$A${sr_last}"
    sr_mgr = f"'Season Results'!$B${sr_first}:$B${sr_last}"
    sr_finish = f"'Season Results'!$C${sr_first}:$C${sr_last}"

    cup_names = [c for c, _, _ in COMPETITIONS if c not in ('Liga 1', 'Liga 2')]
    # Season Results layout: A=Season, B=Manager, C=Finish, then stats
    # columns (W/D/L/GF/GA/YC/RC), then cup columns, then Outcome.
    sr_first_cup_col = 3 + len(SEASON_STATS_COLS) + 1  # 1-based index
    sr_cup_rngs = []
    for i in range(len(cup_names)):
        col = get_column_letter(sr_first_cup_col + i)
        sr_cup_rngs.append(
            f"'Season Results'!${col}${sr_first}:${col}${sr_last}"
        )

    # Outcome column in Season Results (after the cup columns).
    outcome_col_idx = sr_first_cup_col + len(cup_names)
    outcome_col_letter = get_column_letter(outcome_col_idx)
    sr_outcome = f"'Season Results'!${outcome_col_letter}${sr_first}:${outcome_col_letter}${sr_last}"

    # Competitions sheet rows map (for Active + Coefficient lookups).
    comp_row_map = {n: 5 + i for i, (n, _, _) in enumerate(COMPETITIONS)}

    # Array literals for MATCH-based lookups.
    result_levels_literal = '{' + ','.join(f'"{r}"' for r in RESULT_LEVELS) + '}'
    band_labels_literal = '{' + ','.join(f'"{b}"' for b in BAND_LABELS) + '}'

    for idx, (name, _team, _lg, _tier) in enumerate(MANAGERS):
        row = mgr_first + idx
        mgr_ref = f'A{row}'
        league_ref = f'B{row}'
        tier_ref = f'C{row}'
        latest_ref = f'D{row}'
        finish_ref = f'E{row}'
        size_ref = f'F{row}'
        actband_ref = f'G{row}'
        expband_ref = f'H{row}'
        ldelta_ref = f'I{row}'
        bump_ref = f'J{row}'
        oshift_ref = f'K{row}'
        fdelta_ref = f'L{row}'

        ws.cell(row=row, column=1, value=name).alignment = LEFT

        ws.cell(row=row, column=2, value=(
            f'=IFERROR(INDEX({mgrs_league_rng},'
            f'MATCH({mgr_ref},{mgrs_name_rng},0)),"")'
        )).alignment = CENTER

        ws.cell(row=row, column=3, value=(
            f'=IFERROR(INDEX({mgrs_tier_rng},'
            f'MATCH({mgr_ref},{mgrs_name_rng},0)),"")'
        )).alignment = CENTER

        # Latest Season via SUMPRODUCT(MAX((cond)*range)) - portable.
        ws.cell(row=row, column=4, value=(
            f'=IF(COUNTIF({sr_mgr},{mgr_ref})=0,"",'
            f'SUMPRODUCT(MAX(({sr_mgr}={mgr_ref})*{sr_season})))'
        )).alignment = CENTER

        # League Finish for that latest season.
        ws.cell(row=row, column=5, value=(
            f'=IF({latest_ref}="","",'
            f'SUMIFS({sr_finish},{sr_season},{latest_ref},{sr_mgr},{mgr_ref}))'
        )).alignment = CENTER

        # League Size.
        ws.cell(row=row, column=6, value=(
            f'=COUNTIF({mgrs_league_rng},{league_ref})'
        )).alignment = CENTER

        # Actual Band: outcome-based label. Playout wins ties (bottom-4 priority).
        ws.cell(row=row, column=7, value=(
            f'=IF(OR({finish_ref}="",{finish_ref}=0),"",'
            f'IF({finish_ref}>={size_ref}-3,"Playout",'
            f'IF({finish_ref}<=3,"Champion",'
            f'IF({finish_ref}<=8,"Playoff","Mid-safe"))))'
        )).alignment = CENTER

        # Expected Band from Expectations.Effective Band.
        ws.cell(row=row, column=8, value=(
            f'=IFERROR(INDEX({exp_eff_rng},'
            f'MATCH({mgr_ref},{exp_name_rng},0)),"")'
        )).alignment = CENTER

        # League Delta via band-index comparison (lower index = better).
        act_idx = f'MATCH({actband_ref},{band_labels_literal},0)'
        exp_idx = f'MATCH({expband_ref},{band_labels_literal},0)'
        ws.cell(row=row, column=9, value=(
            f'=IF(OR({actband_ref}="",{expband_ref}=""),"",'
            f'IF({act_idx}<{exp_idx},1,'
            f'IF({act_idx}>{exp_idx},-1,0)))'
        )).alignment = CENTER

        # Cup Bump: coefficient-weighted sum across active cups where result
        # meets tier threshold. +1 if sum >= BUMP_SCORE_FLOOR, else 0.
        contrib_terms = []
        for cup_name, cup_rng in zip(cup_names, sr_cup_rngs):
            comp_row = comp_row_map[cup_name]
            active_cell = f'Competitions!$C${comp_row}'
            coef_cell = f'Competitions!$B${comp_row}'
            cell_pull = (
                f'IFERROR(LOOKUP(2,1/(({sr_season}={latest_ref})*'
                f'({sr_mgr}={mgr_ref})),{cup_rng}),"")'
            )
            match_idx = (
                f'IFERROR(MATCH({cell_pull},{result_levels_literal},0),999)'
            )
            # Tier-dependent eligibility (B<=1, C<=3, D<=5, A never).
            b_thr = CUP_BUMP_MAX_IDX_1B['B']
            c_thr = CUP_BUMP_MAX_IDX_1B['C']
            d_thr = CUP_BUMP_MAX_IDX_1B['D']
            eligible = (
                f'IF({tier_ref}="B",{match_idx}<={b_thr},'
                f'IF({tier_ref}="C",{match_idx}<={c_thr},'
                f'IF({tier_ref}="D",{match_idx}<={d_thr},FALSE)))'
            )
            # Contribution = (active=YES) * (eligible) * coefficient.
            contrib = f'(({active_cell}="YES")*({eligible})*{coef_cell})'
            contrib_terms.append(contrib)
        bump_score = '(' + '+'.join(contrib_terms) + ')'
        ws.cell(row=row, column=10, value=(
            f'=IF({latest_ref}="",0,'
            f'IF({bump_score}>={BUMP_SCORE_FLOOR},1,0))'
        )).alignment = CENTER

        # Outcome Shift from Season Results.Outcome for this season.
        outcome_pull = (
            f'IFERROR(LOOKUP(2,1/(({sr_season}={latest_ref})*'
            f'({sr_mgr}={mgr_ref})),{sr_outcome}),"")'
        )
        ws.cell(row=row, column=11, value=(
            f'=IF({latest_ref}="",0,'
            f'IF({outcome_pull}="Promoted",1,'
            f'IF({outcome_pull}="Relegated",-1,0)))'
        )).alignment = CENTER

        # Final Delta: clamp(league_delta + cup_bump_if_eligible, -1, +1)
        # THEN add outcome shift (total can be +-2 for promotion/relegation).
        ws.cell(row=row, column=12, value=(
            f'=IF({ldelta_ref}="",{oshift_ref},'
            f'MAX(-1,MIN(1,{ldelta_ref}+'
            f'IF({ldelta_ref}<0,0,{bump_ref})))+{oshift_ref})'
        )).alignment = CENTER

        # New Tier: A=1..D=4; +1 delta = tier UP (lower index). Clamp 1..4.
        old_idx = (
            f'IF({tier_ref}="A",1,IF({tier_ref}="B",2,'
            f'IF({tier_ref}="C",3,IF({tier_ref}="D",4,0))))'
        )
        new_idx = f'MAX(1,MIN(4,{old_idx}-{fdelta_ref}))'
        ws.cell(row=row, column=13, value=(
            f'=IF(AND({fdelta_ref}=0,{tier_ref}<>""),{tier_ref},'
            f'IF({tier_ref}="","",'
            f'CHOOSE({new_idx},"A","B","C","D")))'
        )).alignment = CENTER

    apply_tier_coloring(ws, f'C{mgr_first}:C{mgr_last}')
    apply_tier_coloring(ws, f'M{mgr_first}:M{mgr_last}')

    ws.freeze_panes = 'A5'
    widths = [14, 8, 10, 13, 13, 11, 12, 13, 12, 10, 13, 12, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_draft_order(ws) -> None:
    ws['A1'] = 'DRAFT ORDER - NBA-style lottery weights by new tier'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:G1')
    ws['A2'] = ("Lower tier = better draft odds. Lottery Weight is each manager's "
                "share of the #1-pick probability. Reverse Rank shows the straight "
                "reverse-order pick (ties = same rank). Run the draw externally "
                "and record the outcome in Actual Pick.")
    ws['A2'].font = Font(italic=True, color='808080')
    ws.merge_cells('A2:G2')

    headers = ['Manager', 'Team', 'League', 'New Tier',
               'Lottery Weight', 'Reverse Rank', 'Actual Pick']
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    head_row(ws, 4, len(headers))

    n_mgr = len(MANAGERS)
    mgr_first = 5
    mgr_last = 4 + n_mgr
    tm_name_rng = f"'Tier Movement'!$A$5:$A${mgr_last}"
    tm_newtier_rng = f"'Tier Movement'!$M$5:$M${mgr_last}"

    # Build nested IF for lottery weight: IF(tier="A",w_A, IF(tier="B",w_B, ...))
    def weight_formula(tier_cell: str) -> str:
        expr = '0'
        for tier in reversed(TIERS):
            expr = f'IF({tier_cell}="{tier}",{LOTTERY_WEIGHTS[tier]},{expr})'
        return expr

    weight_range = f'$E${mgr_first}:$E${mgr_last}'

    for idx, (name, team, league, _tier) in enumerate(MANAGERS):
        row = mgr_first + idx
        mgr_ref = f'A{row}'
        tier_ref = f'D{row}'
        weight_ref = f'E{row}'

        ws.cell(row=row, column=1, value=name).alignment = LEFT
        team_cell = ws.cell(row=row, column=2, value=team)
        team_cell.alignment = LEFT
        team_cell.font = Font(bold=True)
        team_cell.fill = PatternFill('solid', start_color='FFF2CC')
        ws.cell(row=row, column=3, value=league).alignment = CENTER

        ws.cell(row=row, column=4, value=(
            f'=IFERROR(INDEX({tm_newtier_rng},'
            f'MATCH({mgr_ref},{tm_name_rng},0)),"")'
        )).alignment = CENTER

        ws.cell(row=row, column=5, value=(
            f'=IF({tier_ref}="","",{weight_formula(tier_ref)})'
        )).alignment = CENTER

        ws.cell(row=row, column=6, value=(
            f'=IF({tier_ref}="","",RANK({weight_ref},{weight_range}))'
        )).alignment = CENTER

        ws.cell(row=row, column=7, value=None).alignment = CENTER

    apply_tier_coloring(ws, f'D{mgr_first}:D{mgr_last}')

    ws.freeze_panes = 'A5'
    widths = [14, 26, 8, 10, 14, 13, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_career_summary(ws) -> None:
    ws['A1'] = 'CAREER SUMMARY - multi-season aggregates per manager'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:J1')
    ws['A2'] = ("All-time totals computed live from Season Results. No manual "
                "entry. Blank where no seasons are logged. Grows automatically "
                "as admin appends new seasons - use this to track trajectory "
                "across the full patch run.")
    ws['A2'].font = Font(italic=True, color='808080')
    ws.merge_cells('A2:J2')

    cup_names = [c for c, _, _ in COMPETITIONS if c not in ('Liga 1', 'Liga 2')]

    # Season Results column letters (matches build_season_results layout):
    # A=Season, B=Manager, C=Finish, D..=stats, then cups, then Outcome.
    sr_first = 5
    sr_last = 4 + SEASON_RESULTS_BLANK_ROWS
    stats_first_col = 4  # D
    stats_letters = [
        get_column_letter(stats_first_col + i)
        for i in range(len(SEASON_STATS_COLS))
    ]
    # stats map: W=D, D=E, L=F, GF=G, GA=H, YC=I, RC=J
    sr = lambda col: f"'Season Results'!${col}${sr_first}:${col}${sr_last}"
    sr_mgr = sr('B')
    sr_fin = sr('C')
    sr_w, sr_d, sr_l, sr_gf, sr_ga, sr_yc, sr_rc = (sr(c) for c in stats_letters)
    first_cup_col_idx = stats_first_col + len(SEASON_STATS_COLS)
    cup_letters = [get_column_letter(first_cup_col_idx + i) for i in range(len(cup_names))]
    sr_cup_rngs = [sr(c) for c in cup_letters]
    outcome_col_idx = first_cup_col_idx + len(cup_names)
    sr_outcome = sr(get_column_letter(outcome_col_idx))

    headers = [
        'Manager', 'Team', 'League', 'Initial Tier',
        'Seasons', 'Promoted', 'Relegated',
        'Avg Finish', 'Best Finish', 'Worst Finish',
        'W', 'D', 'L', 'GF', 'GA', 'GD', 'YC', 'RC',
    ] + [f'{c} Titles' for c in cup_names] + ['Total Titles']

    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    head_row(ws, 4, len(headers))

    first_title_col = 19  # S (column after RC)
    total_titles_col = first_title_col + len(cup_names)

    for idx, (name, team, league, tier) in enumerate(MANAGERS):
        row = 5 + idx
        mgr = f'$A{row}'
        seasons_ref = f'E{row}'

        ws.cell(row=row, column=1, value=name).alignment = LEFT
        ws.cell(row=row, column=2, value=team).alignment = LEFT
        ws.cell(row=row, column=3, value=league).alignment = CENTER
        ws.cell(row=row, column=4, value=tier).alignment = CENTER

        # Seasons Played
        ws.cell(row=row, column=5,
                value=f'=COUNTIF({sr_mgr},{mgr})').alignment = CENTER

        # Promotions / Relegations
        ws.cell(row=row, column=6,
                value=f'=SUMPRODUCT(({sr_mgr}={mgr})*({sr_outcome}="Promoted"))'
                ).alignment = CENTER
        ws.cell(row=row, column=7,
                value=f'=SUMPRODUCT(({sr_mgr}={mgr})*({sr_outcome}="Relegated"))'
                ).alignment = CENTER

        # Avg / Best / Worst Finish - blank when no seasons logged
        avg_f = (
            f'=IF({seasons_ref}=0,"",'
            f'IFERROR(AVERAGEIFS({sr_fin},{sr_mgr},{mgr},{sr_fin},">0"),""))'
        )
        c = ws.cell(row=row, column=8, value=avg_f)
        c.alignment = CENTER
        c.number_format = '0.0'

        best_f = (
            f'=IF({seasons_ref}=0,"",'
            f'IFERROR(_xlfn.MINIFS({sr_fin},{sr_mgr},{mgr},{sr_fin},">0"),""))'
        )
        ws.cell(row=row, column=9, value=best_f).alignment = CENTER

        worst_f = (
            f'=IF({seasons_ref}=0,"",'
            f'IFERROR(_xlfn.MAXIFS({sr_fin},{sr_mgr},{mgr}),""))'
        )
        ws.cell(row=row, column=10, value=worst_f).alignment = CENTER

        # W / D / L / GF / GA
        stat_ranges = [sr_w, sr_d, sr_l, sr_gf, sr_ga]
        for off, rng in enumerate(stat_ranges):
            f = f'=IF({seasons_ref}=0,"",SUMIFS({rng},{sr_mgr},{mgr}))'
            ws.cell(row=row, column=11 + off, value=f).alignment = CENTER

        # GD = GF - GA (recomputed to avoid "" arithmetic)
        gd_f = (
            f'=IF({seasons_ref}=0,"",'
            f'SUMIFS({sr_gf},{sr_mgr},{mgr})-SUMIFS({sr_ga},{sr_mgr},{mgr}))'
        )
        ws.cell(row=row, column=16, value=gd_f).alignment = CENTER

        # YC / RC
        ws.cell(row=row, column=17,
                value=f'=IF({seasons_ref}=0,"",SUMIFS({sr_yc},{sr_mgr},{mgr}))'
                ).alignment = CENTER
        ws.cell(row=row, column=18,
                value=f'=IF({seasons_ref}=0,"",SUMIFS({sr_rc},{sr_mgr},{mgr}))'
                ).alignment = CENTER

        # Titles per cup
        for i, cup_rng in enumerate(sr_cup_rngs):
            title_f = f'=SUMPRODUCT(({sr_mgr}={mgr})*({cup_rng}="Winner"))'
            ws.cell(row=row, column=first_title_col + i,
                    value=title_f).alignment = CENTER

        # Total Titles = sum of per-cup title cells on this row
        first_letter = get_column_letter(first_title_col)
        last_letter = get_column_letter(first_title_col + len(cup_names) - 1)
        ws.cell(row=row, column=total_titles_col,
                value=f'=SUM({first_letter}{row}:{last_letter}{row})'
                ).alignment = CENTER

    # Auto-filter + freeze + widths
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A4:{last_col_letter}{4 + len(MANAGERS)}"
    ws.freeze_panes = 'B5'

    widths = [14, 22, 8, 10, 9, 11, 11, 11, 12, 12]
    widths += [6] * 8   # W, D, L, GF, GA, GD, YC, RC
    widths += [13] * len(cup_names)
    widths += [12]      # Total Titles
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_player_changes(ws) -> None:
    ws['A1'] = 'PLAYER CHANGES - transfers, drafts, arrangements'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:H1')
    ws['A2'] = (
        "Admin tool for staging player movement between seasons. Entries "
        "here override the Players sheet's Club column (used by Clubs and "
        "Team Composition) ONLY when Apply Changes is YES. Toggle OFF to "
        "preview original rosters; toggle ON to see post-change squads. "
        "Player names must match Players sheet exactly."
    )
    ws['A2'].font = Font(italic=True, color='808080')
    ws.merge_cells('A2:H2')

    # --- Apply Changes toggle (B3) ---
    lbl = ws.cell(row=3, column=1, value='Apply Changes:')
    lbl.font = Font(bold=True)
    lbl.alignment = Alignment(horizontal='right', vertical='center')

    toggle = ws.cell(row=3, column=2, value='NO')
    toggle.alignment = CENTER
    toggle.font = Font(bold=True)
    toggle.fill = PatternFill('solid', start_color='FFE699')

    apply_dv = DataValidation(type='list', formula1='"YES,NO"', allow_blank=False)
    ws.add_data_validation(apply_dv)
    apply_dv.add('B3:B3')

    hint = ws.cell(
        row=3, column=3,
        value='Set to YES to override team squads based on the change log below.'
    )
    hint.font = Font(italic=True, color='808080')
    ws.merge_cells('C3:H3')

    # --- Change Log section ---
    section = ws.cell(row=5, column=1, value='CHANGE LOG')
    section.font = Font(bold=True, size=12, color='1F4E78')
    section.fill = PatternFill('solid', start_color='DDEBF7')
    ws.merge_cells('A5:H5')

    log_headers = ['Season', 'Type', 'Player', 'From Club', 'To Club', 'Notes']
    for i, h in enumerate(log_headers, 1):
        ws.cell(row=6, column=i, value=h)
    head_row(ws, 6, len(log_headers))

    log_first = PLAYER_CHANGES_LOG_FIRST
    log_last = PLAYER_CHANGES_LOG_LAST

    type_dv = DataValidation(
        type='list',
        formula1='"Transfer,Draft-In,Draft-Out,Arrangement"',
        allow_blank=True,
    )
    ws.add_data_validation(type_dv)
    type_dv.add(f'B{log_first}:B{log_last}')

    ws.auto_filter.ref = f"A6:{get_column_letter(len(log_headers))}{log_last}"

    # --- Per-club summary ---
    sum_head_row = log_last + 2
    sum_cols_row = sum_head_row + 1
    sum_first = sum_cols_row + 1
    sum_last = sum_first + len(MANAGERS) - 1

    sh = ws.cell(row=sum_head_row, column=1, value='PER-CLUB SUMMARY')
    sh.font = Font(bold=True, size=12, color='1F4E78')
    sh.fill = PatternFill('solid', start_color='DDEBF7')
    ws.merge_cells(start_row=sum_head_row, start_column=1,
                   end_row=sum_head_row, end_column=8)

    sum_headers = ['Manager', 'Club', 'Transfers In', 'Transfers Out',
                   'Drafts In', 'Drafts Out', 'Arrangements', 'Net']
    for i, h in enumerate(sum_headers, 1):
        ws.cell(row=sum_cols_row, column=i, value=h)
    head_row(ws, sum_cols_row, len(sum_headers))

    log_type_rng = f'$B${log_first}:$B${log_last}'
    log_from_rng = f'$D${log_first}:$D${log_last}'
    log_to_rng = f'$E${log_first}:$E${log_last}'

    for idx, (name, team, league, _tier) in enumerate(MANAGERS):
        row = sum_first + idx
        ws.cell(row=row, column=1, value=name).alignment = LEFT
        club_cell = ws.cell(row=row, column=2, value=team)
        club_cell.alignment = LEFT
        club_cell.font = Font(bold=True)
        club_cell.fill = PatternFill('solid', start_color='FFF2CC')

        club_ref = f'B{row}'

        ws.cell(row=row, column=3, value=(
            f'=COUNTIFS({log_type_rng},"Transfer",{log_to_rng},{club_ref})'
        )).alignment = CENTER
        ws.cell(row=row, column=4, value=(
            f'=COUNTIFS({log_type_rng},"Transfer",{log_from_rng},{club_ref})'
        )).alignment = CENTER
        ws.cell(row=row, column=5, value=(
            f'=COUNTIFS({log_type_rng},"Draft-In",{log_to_rng},{club_ref})'
        )).alignment = CENTER
        ws.cell(row=row, column=6, value=(
            f'=COUNTIFS({log_type_rng},"Draft-Out",{log_from_rng},{club_ref})'
        )).alignment = CENTER
        ws.cell(row=row, column=7, value=(
            f'=COUNTIFS({log_type_rng},"Arrangement",{log_to_rng},{club_ref})+'
            f'COUNTIFS({log_type_rng},"Arrangement",{log_from_rng},{club_ref})'
        )).alignment = CENTER
        ws.cell(row=row, column=8, value=(
            f'=(C{row}+E{row})-(D{row}+F{row})'
        )).alignment = CENTER

    ws.freeze_panes = 'A7'
    widths = [10, 16, 24, 22, 22, 30, 14, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_tactical_mentor(ws) -> None:
    """Dynamic team-scouting tab: pick a team + formation, see squad, best XI,
    and tactical slider suggestions.  All formulas - changing Team Name or
    Formation triggers live recalc.
    """
    INPUT_FILL = PatternFill('solid', start_color='FFF2CC')
    SECTION_FONT = Font(bold=True, size=12, color='1F4E78')
    SECTION_FILL = PatternFill('solid', start_color='DDEBF7')

    # --- Row layout anchors -------------------------------------------------
    R_INPUT = 4            # Team Type at row 4
    R_NAME = 5             # Team Name at row 5
    R_FORM = 6             # Formation at row 6
    R_SIZE = 7             # Squad size display

    R_SQUAD_HDR = 9
    R_SQUAD_COL = 10
    SQUAD_ROWS = TACTICAL_MENTOR_SQUAD_ROWS
    R_SQUAD_FIRST = 11
    R_SQUAD_LAST = R_SQUAD_FIRST + SQUAD_ROWS - 1   # 40

    R_XI_HDR = R_SQUAD_LAST + 2    # 42
    R_XI_COL = R_XI_HDR + 1        # 43
    R_XI_FIRST = R_XI_COL + 1      # 44
    R_XI_LAST = R_XI_FIRST + 10    # 54 (11 slots)

    R_SUBS_HDR = R_XI_LAST + 2     # 56
    R_SUBS_COL = R_SUBS_HDR + 1    # 57
    R_SUBS_FIRST = R_SUBS_COL + 1  # 58
    SUBS_ROWS = TACTICAL_MENTOR_SUBS_ROWS
    R_SUBS_LAST = R_SUBS_FIRST + SUBS_ROWS - 1  # 62

    R_TAC_HDR = R_SUBS_LAST + 2    # 64
    R_TAC_COL = R_TAC_HDR + 1      # 65
    R_TAC_FIRST = R_TAC_COL + 1    # 66
    R_TAC_LAST = R_TAC_FIRST + 4   # 70 (5 metrics)

    R_FORM_HDR = R_TAC_LAST + 2    # 72
    R_FORM_COL = R_FORM_HDR + 1    # 73
    R_FORM_FIRST = R_FORM_COL + 1  # 74
    R_FORM_LAST = R_FORM_FIRST + len(FORMATIONS) - 1  # 80

    # --- Cell refs ----------------------------------------------------------
    TEAM_TYPE = '$B$4'
    TEAM_NAME = '$B$5'
    FORMATION = '$B$6'

    # --- Players sheet ranges ----------------------------------------------
    eff_club_rng = players_range('Effective Club')
    nat_rng = players_range('Nationality')
    name_rng = players_range('Name')
    ovr_rng = players_range('OVR')
    eff_b1_rng = players_range('Effective B1')
    best_pos_rng = players_range('Best Position')
    class_rng = players_range('Class')
    age_rng = players_range('Age')
    ability_rng = players_range('Ability ★')
    topspeed_rng = players_range('TOP SPEED')
    accel_rng = players_range('ACCELERATION')
    stamina_rng = players_range('STAMINA')
    aggression_rng = players_range('AGGRESSION')
    response_rng = players_range('RESPONSE')
    teamwork_rng = players_range('TEAM WORK')
    dline_rng = players_range('D-LINE CONTROL')

    # --- Helper expressions -------------------------------------------------
    # Team filter mask (array of 0/1 per player row).
    team_mask = (
        f'(({TEAM_TYPE}="Club")*({eff_club_rng}={TEAM_NAME})+'
        f'({TEAM_TYPE}="Nation")*({nat_rng}={TEAM_NAME}))'
    )

    def b1_slot_array_expr(position_cell: str) -> str:
        """Each player's B1 at whatever position `position_cell` contains.

        Groups the 16 formation positions by their shared B1 column so the
        expression stays as short as possible.
        """
        col_to_positions: dict[str, list[str]] = {}
        for pos, b1_col in POS_TO_B1COL.items():
            col_to_positions.setdefault(b1_col, []).append(pos)
        parts = []
        for b1_col, positions in col_to_positions.items():
            pos_match = '+'.join(f'({position_cell}="{p}")' for p in positions)
            parts.append(f'({pos_match})*N(+{players_range(b1_col)})')
        return '(' + '+'.join(parts) + ')'

    def excluded_mask_expr(prior_cells: list[str]) -> str:
        """1 where player's name is NOT in any of `prior_cells`, else 0."""
        if not prior_cells:
            return '1'
        return '(' + '*'.join(f'({name_rng}<>{c})' for c in prior_cells) + ')'

    # =======================================================================
    # TITLE
    # =======================================================================
    ws['A1'] = 'TACTICAL MENTOR'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:T1')
    ws['A2'] = (
        "Pick a club or nation + formation; tool pulls the squad, builds a "
        "greedy Best XI (top B1 per slot, each player used once), and "
        "suggests tactical slider values based on squad strengths. All "
        "dynamic - edit Team Name or Formation and everything recalcs."
    )
    ws['A2'].font = Font(italic=True, color='808080')
    ws.merge_cells('A2:T2')

    # =======================================================================
    # INPUT SECTION
    # =======================================================================
    def input_label(row, text):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='right', vertical='center')

    input_label(R_INPUT, 'Team Type:')
    ws.cell(row=R_INPUT, column=2, value='Club').alignment = CENTER
    ws.cell(row=R_INPUT, column=2).font = Font(bold=True)
    ws.cell(row=R_INPUT, column=2).fill = INPUT_FILL
    type_dv = DataValidation(type='list', formula1='"Club,Nation"', allow_blank=False)
    ws.add_data_validation(type_dv)
    type_dv.add(f'B{R_INPUT}:B{R_INPUT}')
    ws.cell(row=R_INPUT, column=3,
            value='(Club -> Effective Club; Nation -> Nationality)').font = \
        Font(italic=True, color='808080')
    ws.merge_cells(start_row=R_INPUT, start_column=3, end_row=R_INPUT, end_column=20)

    input_label(R_NAME, 'Team Name:')
    nm = ws.cell(row=R_NAME, column=2, value='')
    nm.alignment = LEFT
    nm.font = Font(bold=True)
    nm.fill = INPUT_FILL
    ws.cell(row=R_NAME, column=3,
            value='(Type exactly: e.g. "Bochum", "Bayern Munchen", "Germany")').font = \
        Font(italic=True, color='808080')
    ws.merge_cells(start_row=R_NAME, start_column=3, end_row=R_NAME, end_column=20)

    input_label(R_FORM, 'Formation:')
    fm = ws.cell(row=R_FORM, column=2, value='4-4-2')
    fm.alignment = CENTER
    fm.font = Font(bold=True)
    fm.fill = INPUT_FILL
    form_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(FORMATIONS.keys()) + '"',
        allow_blank=False,
    )
    ws.add_data_validation(form_dv)
    form_dv.add(f'B{R_FORM}:B{R_FORM}')
    ws.cell(row=R_FORM, column=3,
            value='(Changing reshuffles Best XI slots below)').font = \
        Font(italic=True, color='808080')
    ws.merge_cells(start_row=R_FORM, start_column=3, end_row=R_FORM, end_column=20)

    input_label(R_SIZE, 'Squad size:')
    sz = ws.cell(row=R_SIZE, column=2, value=f'=SUMPRODUCT({team_mask})')
    sz.alignment = CENTER
    sz.font = Font(bold=True)

    # =======================================================================
    # SQUAD SECTION
    # =======================================================================
    sh = ws.cell(row=R_SQUAD_HDR, column=1,
                 value='SQUAD - sorted by Effective B1 descending')
    sh.font = SECTION_FONT
    sh.fill = SECTION_FILL
    ws.merge_cells(start_row=R_SQUAD_HDR, start_column=1,
                   end_row=R_SQUAD_HDR, end_column=20)

    squad_headers = [
        'Rank', 'Name', 'Best Pos', 'OVR', 'Eff B1', 'Class', 'Age', 'Foot', 'Ability *',
        'GK', 'CB/CWP', 'SB', 'WB', 'DMF', 'CMF', 'SMF', 'AMF', 'WF', 'SS', 'CF',
    ]
    for i, h in enumerate(squad_headers, 1):
        ws.cell(row=R_SQUAD_COL, column=i, value=h)
    head_row(ws, R_SQUAD_COL, len(squad_headers))

    foot_rng = players_range('Foot')

    # First cell of the players range (Players!$BQ$2 or similar) - used for
    # converting an absolute row number to a 1-based position in the range.
    first_cell = eff_b1_rng.split(':')[0]  # e.g. 'Players!$BQ$2'

    # Each squad row uses hidden helper cells in columns U and V:
    #   U = Nth largest Effective B1 in team (via AGGREGATE LARGE)
    #   V = position within players range where that value lives (via
    #       AGGREGATE SMALL over ROW()/(mask=nth) which ignores div-by-zero)
    # Display cells (B..T) INDEX into Players using the V helper.
    # AGGREGATE(14/15, 6, ..., k) is the portable non-CSE way to run
    # LARGE/SMALL over a computed array in Excel 2010+/LibreOffice 4.4+.
    HELPER_NTH = 21   # column U
    HELPER_POS = 22   # column V

    for i in range(SQUAD_ROWS):
        r = R_SQUAD_FIRST + i
        rank = i + 1
        masked_b1 = f'({team_mask}*N(+{eff_b1_rng}))'
        nth_ref = f'$U${r}'
        pos_ref = f'$V${r}'

        ws.cell(row=r, column=1, value=rank).alignment = CENTER

        # Helper U: Nth largest B1 in team.
        # CSE ArrayFormula - needed so LARGE sees the computed array as an
        # array (LibreOffice pre-Calc-7.5-dynamic otherwise returns scalar 0).
        u_cell = f'U{r}'
        ws[u_cell] = ArrayFormula(
            u_cell,
            f'=IFERROR(LARGE({masked_b1},{rank}),0)',
        )
        # Helper V: position in players range where nth_ref lives.
        v_cell = f'V{r}'
        ws[v_cell] = ArrayFormula(
            v_cell,
            f'=IFERROR(IF({nth_ref}<=0,"",'
            f'MATCH({nth_ref},{masked_b1},0)),"")',
        )

        def cell_from(col_idx, src_rng, fmt=None):
            c = ws.cell(row=r, column=col_idx,
                        value=f'=IFERROR(INDEX({src_rng},{pos_ref}),"")')
            c.alignment = LEFT if col_idx == 2 else CENTER
            if fmt:
                c.number_format = fmt
            return c

        cell_from(2, name_rng)              # Name
        cell_from(3, best_pos_rng)          # Best Pos
        cell_from(4, ovr_rng, '0.0')        # OVR
        cell_from(5, eff_b1_rng, '0.00')    # Eff B1
        cell_from(6, class_rng)             # Class
        cell_from(7, age_rng)               # Age
        cell_from(8, foot_rng)              # Foot
        cell_from(9, ability_rng)           # Ability *

        # 11 B1 per position columns
        for j, pos_col in enumerate(POSITIONS):
            b1_range = players_range(f'{pos_col} B1')
            cell_from(10 + j, b1_range, '0.00')

    # Highlight Class column A/B/C/D
    apply_tier_coloring(ws, f'F{R_SQUAD_FIRST}:F{R_SQUAD_LAST}')
    # OVR color scale for quick visual ranking
    apply_effb1_colorscale(ws, f'E{R_SQUAD_FIRST}:E{R_SQUAD_LAST}')

    # =======================================================================
    # BEST XI SECTION
    # =======================================================================
    xh = ws.cell(row=R_XI_HDR, column=1,
                 value='BEST XI - greedy (highest B1 per slot; no duplicates)')
    xh.font = SECTION_FONT
    xh.fill = SECTION_FILL
    ws.merge_cells(start_row=R_XI_HDR, start_column=1,
                   end_row=R_XI_HDR, end_column=10)

    xi_headers = ['Slot', 'Position', 'Player', 'Slot B1', 'OVR', 'Class',
                  'Best Pos', 'Note']
    for i, h in enumerate(xi_headers, 1):
        ws.cell(row=R_XI_COL, column=i, value=h)
    head_row(ws, R_XI_COL, len(xi_headers))

    # Formation reference: each XI row pulls position from the formation table.
    form_name_rng = f'$A${R_FORM_FIRST}:$A${R_FORM_LAST}'
    form_slots_rng = f'$B${R_FORM_FIRST}:$L${R_FORM_LAST}'

    # Track the player-name cells as we go, for the excluded-mask on later slots.
    xi_name_cells: list[str] = []

    # Hidden helpers: I = best B1 at slot, J = match position in players range.
    XI_HELPER_B1 = 9    # column I
    XI_HELPER_POS = 10  # column J

    for i in range(11):
        r = R_XI_FIRST + i
        slot_num = i + 1

        ws.cell(row=r, column=1, value=slot_num).alignment = CENTER

        # Position for this slot (INDEX into the formation table)
        pos_formula = (
            f'=IFERROR(INDEX({form_slots_rng},'
            f'MATCH({FORMATION},{form_name_rng},0),{slot_num}),"")'
        )
        pos_cell = ws.cell(row=r, column=2, value=pos_formula)
        pos_cell.alignment = CENTER
        pos_cell.font = Font(bold=True)

        # Greedy pick helpers via AGGREGATE (array-safe without CSE).
        pos_ref = f'$B${r}'
        b1_arr = b1_slot_array_expr(pos_ref)
        excl = excluded_mask_expr(xi_name_cells)
        masked_slot_b1 = f'({team_mask}*{excl}*{b1_arr})'

        b1_ref = f'$I${r}'
        pos_match_ref = f'$J${r}'

        # Helper I: best available B1 at this slot. CSE array formula.
        i_cell = f'I{r}'
        ws[i_cell] = ArrayFormula(
            i_cell,
            f'=IFERROR(MAX({masked_slot_b1}),0)',
        )
        # Helper J: position in players range where that B1 lives.
        j_cell = f'J{r}'
        ws[j_cell] = ArrayFormula(
            j_cell,
            f'=IFERROR(IF({b1_ref}<=0,"",'
            f'MATCH({b1_ref},{masked_slot_b1},0)),"")',
        )

        ws.cell(row=r, column=3, value=(
            f'=IF({pos_match_ref}="","[no player]",'
            f'IFERROR(INDEX({name_rng},{pos_match_ref}),"[no player]"))'
        )).alignment = LEFT
        slot_b1_cell = ws.cell(
            row=r, column=4,
            value=f'=IF({b1_ref}<=0,"",{b1_ref})'
        )
        slot_b1_cell.alignment = CENTER
        slot_b1_cell.number_format = '0.00'

        ovr_cell = ws.cell(
            row=r, column=5,
            value=f'=IFERROR(INDEX({ovr_rng},{pos_match_ref}),"")'
        )
        ovr_cell.alignment = CENTER
        ovr_cell.number_format = '0.0'

        ws.cell(row=r, column=6,
                value=f'=IFERROR(INDEX({class_rng},{pos_match_ref}),"")').alignment = CENTER
        ws.cell(row=r, column=7,
                value=f'=IFERROR(INDEX({best_pos_rng},{pos_match_ref}),"")').alignment = CENTER

        # "Note" flag: warn when player's natural Best Position doesn't match
        # the slot's position (common with greedy picks).
        note_formula = (
            f'=IFERROR(IF(INDEX({best_pos_rng},{pos_match_ref})="",'
            f'"",'
            f'IF({pos_ref}=INDEX({best_pos_rng},{pos_match_ref}),"",'
            f'"Out of position (best at "&INDEX({best_pos_rng},{pos_match_ref})&")"'
            f')),"")'
        )
        ws.cell(row=r, column=8, value=note_formula).font = \
            Font(italic=True, color='808080')

        xi_name_cells.append(f'$C${r}')

    apply_tier_coloring(ws, f'F{R_XI_FIRST}:F{R_XI_LAST}')

    # =======================================================================
    # SUBSTITUTES SECTION
    # =======================================================================
    subh = ws.cell(row=R_SUBS_HDR, column=1,
                   value='SUBSTITUTES - next 5 by Effective B1, excluding XI')
    subh.font = SECTION_FONT
    subh.fill = SECTION_FILL
    ws.merge_cells(start_row=R_SUBS_HDR, start_column=1,
                   end_row=R_SUBS_HDR, end_column=10)

    sub_headers = ['#', 'Name', 'Best Pos', 'OVR', 'Eff B1', 'Class']
    for i, h in enumerate(sub_headers, 1):
        ws.cell(row=R_SUBS_COL, column=i, value=h)
    head_row(ws, R_SUBS_COL, len(sub_headers))

    # Hidden helpers: H = best available B1, I = match pos in players range.
    SUBS_HELPER_B1 = 8
    SUBS_HELPER_POS = 9

    for i in range(SUBS_ROWS):
        r = R_SUBS_FIRST + i
        rank = i + 1
        # prior sub names already picked
        prior_subs = [f'$B${R_SUBS_FIRST + k}' for k in range(i)]
        excl = excluded_mask_expr(xi_name_cells + prior_subs)
        masked_b1 = f'({team_mask}*{excl}*N(+{eff_b1_rng}))'
        b1_ref = f'$H${r}'
        pos_ref_sub = f'$I${r}'

        h_cell = f'H{r}'
        ws[h_cell] = ArrayFormula(
            h_cell,
            f'=IFERROR(MAX({masked_b1}),0)',
        )
        i_cell_sub = f'I{r}'
        ws[i_cell_sub] = ArrayFormula(
            i_cell_sub,
            f'=IFERROR(IF({b1_ref}<=0,"",'
            f'MATCH({b1_ref},{masked_b1},0)),"")',
        )

        ws.cell(row=r, column=1, value=f'S{rank}').alignment = CENTER
        ws.cell(row=r, column=2,
                value=f'=IFERROR(INDEX({name_rng},{pos_ref_sub}),"")').alignment = LEFT
        ws.cell(row=r, column=3,
                value=f'=IFERROR(INDEX({best_pos_rng},{pos_ref_sub}),"")').alignment = CENTER
        o = ws.cell(row=r, column=4,
                    value=f'=IFERROR(INDEX({ovr_rng},{pos_ref_sub}),"")')
        o.alignment = CENTER
        o.number_format = '0.0'
        b = ws.cell(row=r, column=5,
                    value=f'=IFERROR(INDEX({eff_b1_rng},{pos_ref_sub}),"")')
        b.alignment = CENTER
        b.number_format = '0.00'
        ws.cell(row=r, column=6,
                value=f'=IFERROR(INDEX({class_rng},{pos_ref_sub}),"")').alignment = CENTER

    apply_tier_coloring(ws, f'F{R_SUBS_FIRST}:F{R_SUBS_LAST}')

    # =======================================================================
    # TACTICAL SUGGESTIONS
    # =======================================================================
    th = ws.cell(row=R_TAC_HDR, column=1,
                 value='TACTICAL SUGGESTIONS - based on squad aggregates')
    th.font = SECTION_FONT
    th.fill = SECTION_FILL
    ws.merge_cells(start_row=R_TAC_HDR, start_column=1,
                   end_row=R_TAC_HDR, end_column=10)

    tac_headers = ['Metric', 'Squad Value', 'Recommendation', 'Rationale']
    for i, h in enumerate(tac_headers, 1):
        ws.cell(row=R_TAC_COL, column=i, value=h)
    head_row(ws, R_TAC_COL, len(tac_headers))

    def defender_mask_expr() -> str:
        """1 if player's Best Position is CB/CWP, SB, or WB; else 0."""
        return (
            f'(({best_pos_rng}="CB/CWP")+({best_pos_rng}="SB")+'
            f'({best_pos_rng}="WB"))'
        )

    def attacker_b1_sum_expr() -> str:
        """Sum of Effective B1 over attackers (best pos CF/SS/WF) in team."""
        mask = (
            f'(({best_pos_rng}="CF")+({best_pos_rng}="SS")+({best_pos_rng}="WF"))'
        )
        return f'SUMPRODUCT({team_mask}*{mask}*N(+{eff_b1_rng}))'

    def defender_b1_sum_expr() -> str:
        mask = defender_mask_expr()
        return f'SUMPRODUCT({team_mask}*{mask}*N(+{eff_b1_rng}))'

    def wing_count_expr() -> str:
        """Number of players with wide best positions (WF / SMF)."""
        return (
            f'SUMPRODUCT({team_mask}*'
            f'(({best_pos_rng}="WF")+({best_pos_rng}="SMF")))'
        )

    def team_avg_expr(stat_rng: str) -> str:
        return (
            f'IFERROR(SUMPRODUCT({team_mask}*N(+{stat_rng}))/'
            f'SUMPRODUCT({team_mask}),0)'
        )

    def defender_avg_expr(stat_rng: str) -> str:
        dm = defender_mask_expr()
        return (
            f'IFERROR(SUMPRODUCT({team_mask}*{dm}*N(+{stat_rng}))/'
            f'SUMPRODUCT({team_mask}*{dm}),0)'
        )

    # --- Row 1: Defensive Line ---
    r = R_TAC_FIRST
    ws.cell(row=r, column=1, value='Defensive Line').font = Font(bold=True)
    def_speed_avg = (
        f'({defender_avg_expr(topspeed_rng)}+{defender_avg_expr(accel_rng)})/2'
    )
    v = ws.cell(row=r, column=2, value=f'={def_speed_avg}')
    v.number_format = '0.0'
    v.alignment = CENTER
    speed_cell = f'B{r}'
    ws.cell(row=r, column=3, value=(
        f'=IF(SUMPRODUCT({team_mask})=0,"-",'
        f'IF({speed_cell}>=85,"High (70-80)",'
        f'IF({speed_cell}>=78,"Balanced (50-60)",'
        f'"Deep (20-40)")))'
    )).alignment = CENTER
    ws.cell(row=r, column=4, value=(
        f'=IF(SUMPRODUCT({team_mask})=0,"",'
        f'"Avg defender Top Speed+Acceleration = "&ROUND({speed_cell},0)&". "&'
        f'IF({speed_cell}>=85,"Fast backs can recover; push the line up.",'
        f'IF({speed_cell}>=78,"Moderate pace; hold a normal line.",'
        f'"Slower backs; sit deep to avoid being exposed on the break.")))'
    )).alignment = LEFT

    # --- Row 2: Pressing ---
    r += 1
    ws.cell(row=r, column=1, value='Pressing').font = Font(bold=True)
    press_score = f'({team_avg_expr(stamina_rng)}+{team_avg_expr(aggression_rng)})/2'
    v = ws.cell(row=r, column=2, value=f'={press_score}')
    v.number_format = '0.0'
    v.alignment = CENTER
    press_cell = f'B{r}'
    ws.cell(row=r, column=3, value=(
        f'=IF(SUMPRODUCT({team_mask})=0,"-",'
        f'IF({press_cell}>=80,"High (75-85)",'
        f'IF({press_cell}>=72,"Medium (50-60)",'
        f'"Low (30-40)")))'
    )).alignment = CENTER
    ws.cell(row=r, column=4, value=(
        f'=IF(SUMPRODUCT({team_mask})=0,"",'
        f'"Avg Stamina+Aggression = "&ROUND({press_cell},0)&". "&'
        f'IF({press_cell}>=80,"Fit and fierce - press high to force turnovers.",'
        f'IF({press_cell}>=72,"Moderate engine - selective pressing in own half.",'
        f'"Low tank - conserve energy; press only around the ball.")))'
    )).alignment = LEFT

    # --- Row 3: Offside Trap ---
    r += 1
    ws.cell(row=r, column=1, value='Offside Trap').font = Font(bold=True)
    trap_score = (
        f'({defender_avg_expr(response_rng)}+'
        f'{defender_avg_expr(teamwork_rng)}+'
        f'{defender_avg_expr(dline_rng)})/3'
    )
    v = ws.cell(row=r, column=2, value=f'={trap_score}')
    v.number_format = '0.0'
    v.alignment = CENTER
    trap_cell = f'B{r}'
    ws.cell(row=r, column=3, value=(
        f'=IF(SUMPRODUCT({team_mask})=0,"-",'
        f'IF({trap_cell}>=82,"ON (aggressive)",'
        f'IF({trap_cell}>=75,"Situational",'
        f'"OFF")))'
    )).alignment = CENTER
    ws.cell(row=r, column=4, value=(
        f'=IF(SUMPRODUCT({team_mask})=0,"",'
        f'"Defender Response+Teamwork+D-Line Ctrl avg = "&ROUND({trap_cell},0)&". "&'
        f'IF({trap_cell}>=82,"Cohesive line can step up in unison.",'
        f'IF({trap_cell}>=75,"Use only against predictable attackers.",'
        f'"Risk too high - leave off.")))'
    )).alignment = LEFT

    # --- Row 4: Attack Strategy ---
    r += 1
    ws.cell(row=r, column=1, value='Attack Strategy').font = Font(bold=True)
    att_sum = attacker_b1_sum_expr()
    def_sum = defender_b1_sum_expr()
    # bias = (att_sum - def_sum) normalised; use raw diff for simplicity
    # Protect divide-by-zero with IFERROR.
    bias = f'IFERROR(({att_sum}-{def_sum})/SUMPRODUCT({team_mask}),0)'
    v = ws.cell(row=r, column=2, value=f'={bias}')
    v.number_format = '0.00'
    v.alignment = CENTER
    bias_cell = f'B{r}'
    ws.cell(row=r, column=3, value=(
        f'=IF(SUMPRODUCT({team_mask})=0,"-",'
        f'IF({bias_cell}>=2,"All-Out Attack",'
        f'IF({bias_cell}<=-2,"Counter Attack",'
        f'"Possession / Balanced")))'
    )).alignment = CENTER
    ws.cell(row=r, column=4, value=(
        f'=IF(SUMPRODUCT({team_mask})=0,"",'
        f'"Attacker vs defender B1 bias = "&ROUND({bias_cell},1)&". "&'
        f'IF({bias_cell}>=2,"Attack is your strength - commit bodies forward.",'
        f'IF({bias_cell}<=-2,"Defence is your strength - sit and spring.",'
        f'"Even side - keep the ball and pick your moments.")))'
    )).alignment = LEFT

    # --- Row 5: Attack Width ---
    r += 1
    ws.cell(row=r, column=1, value='Attack Width').font = Font(bold=True)
    wing_cnt = wing_count_expr()
    v = ws.cell(row=r, column=2, value=f'={wing_cnt}')
    v.number_format = '0'
    v.alignment = CENTER
    wing_cell = f'B{r}'
    ws.cell(row=r, column=3, value=(
        f'=IF(SUMPRODUCT({team_mask})=0,"-",'
        f'IF({wing_cell}>=4,"Wide",'
        f'IF({wing_cell}>=2,"Balanced","Narrow")))'
    )).alignment = CENTER
    ws.cell(row=r, column=4, value=(
        f'=IF(SUMPRODUCT({team_mask})=0,"",'
        f'"Wingers (WF/SMF best pos) in squad = "&{wing_cell}&". "&'
        f'IF({wing_cell}>=4,"Plenty of wide options - stretch the pitch.",'
        f'IF({wing_cell}>=2,"Some width - use flanks situationally.",'
        f'"Few wingers - attack through the middle.")))'
    )).alignment = LEFT

    # =======================================================================
    # FORMATION REFERENCE (used by Best XI INDEX lookup)
    # =======================================================================
    fh = ws.cell(row=R_FORM_HDR, column=1,
                 value='FORMATION REFERENCE - slot-by-slot position list')
    fh.font = SECTION_FONT
    fh.fill = SECTION_FILL
    ws.merge_cells(start_row=R_FORM_HDR, start_column=1,
                   end_row=R_FORM_HDR, end_column=12)

    form_col_headers = ['Formation'] + [f'Slot {i}' for i in range(1, 12)]
    for i, h in enumerate(form_col_headers, 1):
        ws.cell(row=R_FORM_COL, column=i, value=h)
    head_row(ws, R_FORM_COL, len(form_col_headers))

    for i, (fname, slots) in enumerate(FORMATIONS.items()):
        r = R_FORM_FIRST + i
        ws.cell(row=r, column=1, value=fname).font = Font(bold=True)
        for j, slot_pos in enumerate(slots):
            c = ws.cell(row=r, column=2 + j, value=slot_pos)
            c.alignment = CENTER

    # =======================================================================
    # MANUAL STARTING XI - admin picks player + position for each slot
    # =======================================================================
    R_MXI_HDR = R_FORM_LAST + 2           # 82
    R_MXI_COL = R_MXI_HDR + 2             # 84
    R_MXI_FIRST = R_MXI_COL + 1           # 85
    R_MXI_LAST = R_MXI_FIRST + 10         # 95
    R_FORM_DETECT = R_MXI_LAST + 2        # 97
    R_MXI_COUNTS = R_FORM_DETECT + 1      # 98

    mh = ws.cell(row=R_MXI_HDR, column=1,
                 value='MANUAL STARTING XI - admin picks player and position')
    mh.font = SECTION_FONT
    mh.fill = SECTION_FILL
    ws.merge_cells(start_row=R_MXI_HDR, start_column=1,
                   end_row=R_MXI_HDR, end_column=10)
    ws.cell(row=R_MXI_HDR + 1, column=1, value=(
        'Player dropdown pulls from current squad. Position dropdown = 16 '
        'PES6 positions. Slot B1 = player\'s B1 at the picked position. '
        'Validation flags duplicates, missing fields, and PES6 position '
        'limits (1 GK, 1 CWP, 1 of each LB/RB/LWB/RWB/LMF/RMF/LWF/RWF, max '
        '6 per category).'
    )).font = Font(italic=True, color='808080')
    ws.merge_cells(start_row=R_MXI_HDR + 1, start_column=1,
                   end_row=R_MXI_HDR + 1, end_column=10)

    mxi_headers = ['Slot', 'Player', 'Position', 'Slot B1', 'OVR', 'Class',
                   'Best Pos', 'Validation']
    for i, h in enumerate(mxi_headers, 1):
        ws.cell(row=R_MXI_COL, column=i, value=h)
    head_row(ws, R_MXI_COL, len(mxi_headers))

    # Squad name range (source of player dropdown).
    squad_names_rng = f'$B${R_SQUAD_FIRST}:$B${R_SQUAD_LAST}'
    # Full players Name range (for MATCH to get player row in Players sheet).
    mxi_player_rng = f'$B${R_MXI_FIRST}:$B${R_MXI_LAST}'
    mxi_pos_rng = f'$C${R_MXI_FIRST}:$C${R_MXI_LAST}'

    player_dv = DataValidation(
        type='list',
        formula1=f'={squad_names_rng}',
        allow_blank=True,
    )
    ws.add_data_validation(player_dv)
    player_dv.add(f'B{R_MXI_FIRST}:B{R_MXI_LAST}')

    position_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(POSITIONS_FULL) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(position_dv)
    position_dv.add(f'C{R_MXI_FIRST}:C{R_MXI_LAST}')

    # Helper to compute slot B1 for a given player_row and position cell.
    def slot_b1_for_player_expr(player_row_cell: str, position_cell: str) -> str:
        col_to_positions: dict[str, list[str]] = {}
        for pos, b1_col in POS_TO_B1COL.items():
            col_to_positions.setdefault(b1_col, []).append(pos)
        parts = []
        for b1_col, positions in col_to_positions.items():
            pos_match = '+'.join(f'({position_cell}="{p}")' for p in positions)
            idx = f'IFERROR(INDEX({players_range(b1_col)},{player_row_cell}),0)'
            parts.append(f'({pos_match})*{idx}')
        return '(' + '+'.join(parts) + ')'

    # Single-side positions (LB/RB/LWB/RWB/LMF/RMF/LWF/RWF) - for "two on
    # same side" check.
    side_positions = sorted(POSITION_SINGLE_SLOT - {'GK', 'CWP'})
    side_check_or = '+'.join(f'({{pos}}="{p}")' for p in side_positions)

    for i in range(11):
        r = R_MXI_FIRST + i
        slot_num = i + 1
        player_ref = f'$B${r}'
        pos_ref = f'$C${r}'

        ws.cell(row=r, column=1, value=slot_num).alignment = CENTER
        # Player/Position cells start blank; dropdowns are already attached.
        ws.cell(row=r, column=2, value=None).alignment = LEFT
        ws.cell(row=r, column=3, value=None).alignment = CENTER

        # Player row in Players sheet (used by multiple lookups).
        player_row_expr = f'IFERROR(MATCH({player_ref},{name_rng},0),"")'

        # Slot B1 at chosen position.
        b1_expr = slot_b1_for_player_expr(player_row_expr, pos_ref)
        b1_cell = ws.cell(row=r, column=4, value=(
            f'=IF(OR({player_ref}="",{pos_ref}=""),"",{b1_expr})'
        ))
        b1_cell.alignment = CENTER
        b1_cell.number_format = '0.00'

        # OVR, Class, Best Pos (natural - independent of position slot).
        ws.cell(row=r, column=5, value=(
            f'=IF({player_ref}="","",IFERROR(INDEX({ovr_rng},{player_row_expr}),""))'
        )).alignment = CENTER
        ws.cell(row=r, column=5).number_format = '0.0'
        ws.cell(row=r, column=6, value=(
            f'=IF({player_ref}="","",IFERROR(INDEX({class_rng},{player_row_expr}),""))'
        )).alignment = CENTER
        ws.cell(row=r, column=7, value=(
            f'=IF({player_ref}="","",IFERROR(INDEX({best_pos_rng},{player_row_expr}),""))'
        )).alignment = CENTER

        # Validation: first failure wins.
        side_or = side_check_or.replace('{pos}', pos_ref)
        val_formula = (
            f'=IF({player_ref}="","",'
            f'IF(COUNTIF({squad_names_rng},{player_ref})=0,"Not in squad",'
            f'IF(COUNTIF({mxi_player_rng},{player_ref})>1,"Duplicate player",'
            f'IF({pos_ref}="","Position missing",'
            f'IF(AND({pos_ref}="GK",COUNTIF({mxi_pos_rng},"GK")>1),"Multiple GKs",'
            f'IF(AND({pos_ref}="CWP",COUNTIF({mxi_pos_rng},"CWP")>1),"Multiple CWPs",'
            f'IF(AND(({side_or})>0,COUNTIF({mxi_pos_rng},{pos_ref})>1),'
            f'"Two on same side ("&{pos_ref}&")",'
            f'"OK")))))))'
        )
        ws.cell(row=r, column=8, value=val_formula).alignment = CENTER

    apply_tier_coloring(ws, f'F{R_MXI_FIRST}:F{R_MXI_LAST}')

    # --- Detected Formation + category counts ---
    def_count_expr = (
        f'(COUNTIF({mxi_pos_rng},"CB")+COUNTIF({mxi_pos_rng},"CWP")+'
        f'COUNTIF({mxi_pos_rng},"LB")+COUNTIF({mxi_pos_rng},"RB")+'
        f'COUNTIF({mxi_pos_rng},"LWB")+COUNTIF({mxi_pos_rng},"RWB"))'
    )
    dmf_count_expr = f'COUNTIF({mxi_pos_rng},"DMF")'
    mid_count_expr = (
        f'(COUNTIF({mxi_pos_rng},"CMF")+COUNTIF({mxi_pos_rng},"LMF")+'
        f'COUNTIF({mxi_pos_rng},"RMF"))'
    )
    amf_count_expr = f'COUNTIF({mxi_pos_rng},"AMF")'
    att_count_expr = (
        f'(COUNTIF({mxi_pos_rng},"LWF")+COUNTIF({mxi_pos_rng},"RWF")+'
        f'COUNTIF({mxi_pos_rng},"SS")+COUNTIF({mxi_pos_rng},"CF"))'
    )
    gk_count_expr = f'COUNTIF({mxi_pos_rng},"GK")'
    all_def_expr = def_count_expr
    all_mid_expr = f'({dmf_count_expr}+{mid_count_expr}+{amf_count_expr})'
    all_att_expr = att_count_expr

    # Formation string: join non-zero segments with "-"
    # Layers: backline | DMF | Mid | AMF | Fwd
    def seg(expr):
        return f'IF({expr}>0,{expr}&"-","")'
    segs = (
        f'{seg(def_count_expr)}&{seg(dmf_count_expr)}&'
        f'{seg(mid_count_expr)}&{seg(amf_count_expr)}&'
        f'{seg(att_count_expr)}'
    )
    # Strip trailing "-" if non-empty.
    form_string_formula = (
        f'=IF(LEN({segs})=0,"",LEFT({segs},LEN({segs})-1))'
    )

    lbl = ws.cell(row=R_FORM_DETECT, column=1, value='Detected Formation:')
    lbl.font = Font(bold=True)
    lbl.alignment = Alignment(horizontal='right')
    fdetect = ws.cell(row=R_FORM_DETECT, column=2, value=form_string_formula)
    fdetect.font = Font(bold=True, color='1F4E78', size=12)
    fdetect.alignment = LEFT

    lbl2 = ws.cell(row=R_MXI_COUNTS, column=1, value='Category counts:')
    lbl2.font = Font(bold=True)
    lbl2.alignment = Alignment(horizontal='right')
    ws.cell(row=R_MXI_COUNTS, column=2, value=f'=\"GK: \"&{gk_count_expr}').alignment = LEFT
    ws.cell(row=R_MXI_COUNTS, column=3, value=f'=\"Defenders: \"&{all_def_expr}').alignment = LEFT
    ws.cell(row=R_MXI_COUNTS, column=4, value=f'=\"Midfielders: \"&{all_mid_expr}').alignment = LEFT
    ws.cell(row=R_MXI_COUNTS, column=5, value=f'=\"Attackers: \"&{all_att_expr}').alignment = LEFT
    ws.cell(row=R_MXI_COUNTS, column=6, value=(
        f'=IF({gk_count_expr}>1,"[Multi GK]",'
        f'IF({all_def_expr}>{CATEGORY_MAX},"[Too many defenders]",'
        f'IF({all_mid_expr}>{CATEGORY_MAX},"[Too many midfielders]",'
        f'IF({all_att_expr}>{CATEGORY_MAX},"[Too many attackers]",""))))'
    )).font = Font(bold=True, color='C00000')

    # =======================================================================
    # MANUAL SUBSTITUTES
    # =======================================================================
    R_MSUBS_HDR = R_MXI_COUNTS + 2          # 100
    R_MSUBS_COL = R_MSUBS_HDR + 1           # 101
    R_MSUBS_FIRST = R_MSUBS_COL + 1         # 102
    R_MSUBS_LAST = R_MSUBS_FIRST + TACTICAL_MENTOR_SUBS_MANUAL_ROWS - 1  # 108

    mshh = ws.cell(row=R_MSUBS_HDR, column=1,
                   value='MANUAL SUBSTITUTES - 7 slots, player only (no position)')
    mshh.font = SECTION_FONT
    mshh.fill = SECTION_FILL
    ws.merge_cells(start_row=R_MSUBS_HDR, start_column=1,
                   end_row=R_MSUBS_HDR, end_column=10)

    ms_headers = ['#', 'Player', 'Best Pos', 'OVR', 'Eff B1', 'Class', 'Validation']
    for i, h in enumerate(ms_headers, 1):
        ws.cell(row=R_MSUBS_COL, column=i, value=h)
    head_row(ws, R_MSUBS_COL, len(ms_headers))

    msub_player_rng = f'$B${R_MSUBS_FIRST}:$B${R_MSUBS_LAST}'
    msub_dv = DataValidation(
        type='list',
        formula1=f'={squad_names_rng}',
        allow_blank=True,
    )
    ws.add_data_validation(msub_dv)
    msub_dv.add(f'B{R_MSUBS_FIRST}:B{R_MSUBS_LAST}')

    for i in range(TACTICAL_MENTOR_SUBS_MANUAL_ROWS):
        r = R_MSUBS_FIRST + i
        sub_num = i + 1
        player_ref = f'$B${r}'
        player_row_expr = f'IFERROR(MATCH({player_ref},{name_rng},0),"")'

        ws.cell(row=r, column=1, value=f'S{sub_num}').alignment = CENTER
        ws.cell(row=r, column=2, value=None).alignment = LEFT
        ws.cell(row=r, column=3, value=(
            f'=IF({player_ref}="","",IFERROR(INDEX({best_pos_rng},{player_row_expr}),""))'
        )).alignment = CENTER
        o = ws.cell(row=r, column=4, value=(
            f'=IF({player_ref}="","",IFERROR(INDEX({ovr_rng},{player_row_expr}),""))'
        ))
        o.alignment = CENTER
        o.number_format = '0.0'
        b = ws.cell(row=r, column=5, value=(
            f'=IF({player_ref}="","",IFERROR(INDEX({eff_b1_rng},{player_row_expr}),""))'
        ))
        b.alignment = CENTER
        b.number_format = '0.00'
        ws.cell(row=r, column=6, value=(
            f'=IF({player_ref}="","",IFERROR(INDEX({class_rng},{player_row_expr}),""))'
        )).alignment = CENTER

        # Sub validation: not in squad, duplicate sub, or duplicate with XI.
        val_formula = (
            f'=IF({player_ref}="","",'
            f'IF(COUNTIF({squad_names_rng},{player_ref})=0,"Not in squad",'
            f'IF(COUNTIF({msub_player_rng},{player_ref})>1,"Duplicate sub",'
            f'IF(COUNTIF({mxi_player_rng},{player_ref})>0,"Already in XI",'
            f'"OK"))))'
        )
        ws.cell(row=r, column=7, value=val_formula).alignment = CENTER

    apply_tier_coloring(ws, f'F{R_MSUBS_FIRST}:F{R_MSUBS_LAST}')

    # =======================================================================
    # TACTICS SELECTION (up to 4) + SLIDER SUGGESTIONS (0-20)
    # =======================================================================
    R_TAC_SEL_HDR = R_MSUBS_LAST + 2        # 110
    R_TAC_SEL_FIRST = R_TAC_SEL_HDR + 2     # 112
    R_TAC_SEL_LAST = R_TAC_SEL_FIRST + 3    # 115 (4 tactic slots)

    tsh = ws.cell(row=R_TAC_SEL_HDR, column=1,
                  value='TACTICS SELECTION - pick up to 4 PES6 team-style presets')
    tsh.font = SECTION_FONT
    tsh.fill = SECTION_FILL
    ws.merge_cells(start_row=R_TAC_SEL_HDR, start_column=1,
                   end_row=R_TAC_SEL_HDR, end_column=10)
    ws.cell(row=R_TAC_SEL_HDR + 1, column=1,
            value=('Leave blank to disable a slot. Each chosen tactic '
                   'shifts the slider values and the per-player B1 modifiers.')
            ).font = Font(italic=True, color='808080')
    ws.merge_cells(start_row=R_TAC_SEL_HDR + 1, start_column=1,
                   end_row=R_TAC_SEL_HDR + 1, end_column=10)

    tac_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(TACTIC_PRESETS) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(tac_dv)
    tac_dv.add(f'B{R_TAC_SEL_FIRST}:B{R_TAC_SEL_LAST}')

    for i in range(4):
        r = R_TAC_SEL_FIRST + i
        lbl = ws.cell(row=r, column=1, value=f'Tactic {i + 1}:')
        lbl.font = Font(bold=True)
        lbl.alignment = Alignment(horizontal='right')
        tc = ws.cell(row=r, column=2, value=None)
        tc.alignment = CENTER
        tc.fill = INPUT_FILL
        tc.font = Font(bold=True)

    tactic_cells = [f'$B${R_TAC_SEL_FIRST + i}' for i in range(4)]

    # --- Sliders ---
    R_SLIDER_HDR = R_TAC_SEL_LAST + 2       # 117
    R_SLIDER_COL = R_SLIDER_HDR + 1         # 118
    R_SLIDER_FIRST = R_SLIDER_COL + 1       # 119
    SLIDER_NAMES = ['Defensive Line', 'Pressing', 'Compactness',
                    'Width', 'Player Support', 'Position Switch']
    R_SLIDER_LAST = R_SLIDER_FIRST + len(SLIDER_NAMES) - 1  # 124

    slh = ws.cell(row=R_SLIDER_HDR, column=1,
                  value='SLIDER SUGGESTIONS (0-20 PES6 scale)')
    slh.font = SECTION_FONT
    slh.fill = SECTION_FILL
    ws.merge_cells(start_row=R_SLIDER_HDR, start_column=1,
                   end_row=R_SLIDER_HDR, end_column=10)

    slider_headers = ['Slider', 'Value (0-20)', 'Direction', 'Rationale']
    for i, h in enumerate(slider_headers, 1):
        ws.cell(row=R_SLIDER_COL, column=i, value=h)
    head_row(ws, R_SLIDER_COL, len(slider_headers))

    for idx, slider in enumerate(SLIDER_NAMES):
        r = R_SLIDER_FIRST + idx
        deltas = TACTIC_SLIDER_DELTAS.get(slider, {})
        # Build SUMPRODUCT-style delta: for each tactic slot, if it matches
        # a tactic with a delta, add that delta.
        delta_terms = []
        for tac, d in deltas.items():
            for tc in tactic_cells:
                delta_terms.append(f'({tc}="{tac}")*({d})')
        delta_sum = '(' + '+'.join(delta_terms) + ')' if delta_terms else '0'
        val_formula = f'=MAX(0,MIN(20,10+{delta_sum}))'

        ws.cell(row=r, column=1, value=slider).font = Font(bold=True)
        v = ws.cell(row=r, column=2, value=val_formula)
        v.alignment = CENTER
        v.number_format = '0'

        v_ref = f'B{r}'
        # Direction hint based on value vs 10 (neutral).
        dir_formula = (
            f'=IF({v_ref}>=14,"HIGH",IF({v_ref}>=8,"MEDIUM","LOW"))'
        )
        ws.cell(row=r, column=3, value=dir_formula).alignment = CENTER
        # Rationale lists contributing tactics and their deltas.
        # Generate text like "Frontline Pressure (+8), Counter Attack (-3)"
        rationale_parts = []
        for tac, d in deltas.items():
            sign = '+' if d > 0 else ''
            rationale_parts.append(
                f'IF(OR({",".join(f"{tc}=\"{tac}\"" for tc in tactic_cells)}),'
                f'"{tac} ({sign}{d})",'
                f'"")'
            )
        if rationale_parts:
            # Non-empty strings joined with ", ". TEXTJOIN is Excel 2019+ and
            # LibreOffice 5.2+, but openpyxl omits the _xlfn. prefix xlsx
            # expects - prefixing it explicitly keeps both apps happy.
            rat_formula = (
                f'=_xlfn.TEXTJOIN(", ",TRUE,'
                + ','.join(rationale_parts)
                + ')'
            )
        else:
            rat_formula = '=""'
        rat_cell = ws.cell(row=r, column=4, value=rat_formula)
        rat_cell.font = Font(italic=True, color='808080')
        rat_cell.alignment = LEFT

    # =======================================================================
    # EFFICIENCY PANEL
    # =======================================================================
    R_EFF_HDR = R_SLIDER_LAST + 2           # 126
    R_EFF_SUM_LBL = R_EFF_HDR + 1           # 127
    R_EFF_SUM_VAL = R_EFF_SUM_LBL + 1       # 128
    R_EFF_PP_HDR = R_EFF_SUM_VAL + 2        # 130
    R_EFF_PP_COL = R_EFF_PP_HDR + 1         # 131
    R_EFF_PP_FIRST = R_EFF_PP_COL + 1       # 132
    R_EFF_PP_LAST = R_EFF_PP_FIRST + 10     # 142

    eh = ws.cell(row=R_EFF_HDR, column=1,
                 value='EFFICIENCY PANEL - how your tactics change squad B1')
    eh.font = SECTION_FONT
    eh.fill = SECTION_FILL
    ws.merge_cells(start_row=R_EFF_HDR, start_column=1,
                   end_row=R_EFF_HDR, end_column=10)

    # Summary row labels
    for i, h in enumerate(['Baseline XI B1', 'With Tactics B1', 'Change %',
                           'Overall Rating'], 1):
        c = ws.cell(row=R_EFF_SUM_LBL, column=i, value=h)
        c.font = Font(bold=True)
        c.alignment = CENTER

    # Per-player impact table
    pp_headers = ['Slot', 'Player', 'Position', 'Category', 'Baseline B1',
                  'Total Mod %', 'Adjusted B1', 'Delta']
    for i, h in enumerate(pp_headers, 1):
        ws.cell(row=R_EFF_PP_COL, column=i, value=h)
    head_row(ws, R_EFF_PP_COL, len(pp_headers))

    # Reference ranges for the modifier table (defined below).
    R_MOD_HDR = R_EFF_PP_LAST + 2           # 144
    R_MOD_COL = R_MOD_HDR + 1               # 145
    R_MOD_FIRST = R_MOD_COL + 1             # 146
    R_MOD_LAST = R_MOD_FIRST + len(TACTIC_PRESETS) - 1  # 155

    mod_tac_names_rng = f'$A${R_MOD_FIRST}:$A${R_MOD_LAST}'
    # Category headers are in columns B..(B+11-1) = B..L at row R_MOD_COL
    mod_cat_headers_rng = (
        f'$B${R_MOD_COL}:$L${R_MOD_COL}'
    )
    mod_table_rng = f'$B${R_MOD_FIRST}:$L${R_MOD_LAST}'

    # Per-player impact rows reference XI rows directly.
    for i in range(11):
        r = R_EFF_PP_FIRST + i
        xi_r = R_MXI_FIRST + i
        player_ref = f'$B${xi_r}'
        pos_ref = f'$C${xi_r}'
        baseline_b1_ref = f'$D${xi_r}'

        ws.cell(row=r, column=1, value=i + 1).alignment = CENTER
        ws.cell(row=r, column=2, value=f'=IF({player_ref}="","",{player_ref})').alignment = LEFT
        ws.cell(row=r, column=3, value=f'=IF({pos_ref}="","",{pos_ref})').alignment = CENTER

        # Category from position.
        cat_formula = (
            f'=IF({pos_ref}="","",'
            + ''.join(
                f'IF({pos_ref}="{p}","{POSITION_TO_CATEGORY[p]}",'
                for p in POSITIONS_FULL
            )
            + '""'
            + ')' * len(POSITIONS_FULL)
            + ')'
        )
        ws.cell(row=r, column=4, value=cat_formula).alignment = CENTER

        baseline_cell = f'E{r}'
        ws.cell(row=r, column=5, value=(
            f'=IF({baseline_b1_ref}="","",{baseline_b1_ref})'
        )).alignment = CENTER
        ws.cell(row=r, column=5).number_format = '0.00'

        # Total modifier % = sum over 4 tactic slots of
        # INDEX(mod_table, MATCH(tactic, tac_names, 0), MATCH(cat, cat_headers, 0))
        cat_ref = f'$D${r}'
        mod_terms = []
        for tc in tactic_cells:
            term = (
                f'IFERROR(INDEX({mod_table_rng},'
                f'MATCH({tc},{mod_tac_names_rng},0),'
                f'MATCH({cat_ref},{mod_cat_headers_rng},0)),0)'
            )
            mod_terms.append(term)
        mod_sum_expr = '(' + '+'.join(mod_terms) + ')'
        mod_cell = ws.cell(row=r, column=6, value=(
            f'=IF(OR({player_ref}="",{pos_ref}=""),"",{mod_sum_expr})'
        ))
        mod_cell.alignment = CENTER
        mod_cell.number_format = '0'

        mod_ref = f'F{r}'
        adj_b1_formula = (
            f'=IF(OR({baseline_b1_ref}="",{mod_ref}=""),"",'
            f'{baseline_b1_ref}*(1+{mod_ref}/100))'
        )
        adj_cell = ws.cell(row=r, column=7, value=adj_b1_formula)
        adj_cell.alignment = CENTER
        adj_cell.number_format = '0.00'

        delta_cell = ws.cell(row=r, column=8, value=(
            f'=IF(OR({baseline_b1_ref}="",G{r}=""),"",G{r}-{baseline_b1_ref})'
        ))
        delta_cell.alignment = CENTER
        delta_cell.number_format = '+0.00;-0.00;0.00'

    # Summary row values
    pp_baseline_rng = f'$E${R_EFF_PP_FIRST}:$E${R_EFF_PP_LAST}'
    pp_adj_rng = f'$G${R_EFF_PP_FIRST}:$G${R_EFF_PP_LAST}'
    baseline_sum_formula = (
        f'=IFERROR(SUMPRODUCT(N(+{pp_baseline_rng})),0)'
    )
    adj_sum_formula = (
        f'=IFERROR(SUMPRODUCT(N(+{pp_adj_rng})),0)'
    )
    bsum = ws.cell(row=R_EFF_SUM_VAL, column=1, value=baseline_sum_formula)
    bsum.number_format = '0.00'
    bsum.alignment = CENTER
    asum = ws.cell(row=R_EFF_SUM_VAL, column=2, value=adj_sum_formula)
    asum.number_format = '0.00'
    asum.alignment = CENTER
    a_ref = f'A{R_EFF_SUM_VAL}'
    b_ref = f'B{R_EFF_SUM_VAL}'
    pct = ws.cell(row=R_EFF_SUM_VAL, column=3, value=(
        f'=IFERROR(({b_ref}-{a_ref})/{a_ref}*100,0)'
    ))
    pct.number_format = '+0.0"%";-0.0"%";0.0"%"'
    pct.alignment = CENTER
    rating_formula = (
        f'=IFERROR(IF({a_ref}=0,"-",'
        f'IF(({b_ref}-{a_ref})/{a_ref}*100>=5,"Excellent fit",'
        f'IF(({b_ref}-{a_ref})/{a_ref}*100>=2,"Good fit",'
        f'IF(({b_ref}-{a_ref})/{a_ref}*100>=-2,"Neutral",'
        f'IF(({b_ref}-{a_ref})/{a_ref}*100>=-5,"Slight misfit",'
        f'"Bad fit"))))),"-")'
    )
    rat = ws.cell(row=R_EFF_SUM_VAL, column=4, value=rating_formula)
    rat.font = Font(bold=True)
    rat.alignment = CENTER

    # =======================================================================
    # TACTIC MODIFIER TABLE (reference; editable)
    # =======================================================================
    mhh = ws.cell(row=R_MOD_HDR, column=1,
                  value='TACTIC MODIFIER TABLE (% applied to slot B1; admin-editable)')
    mhh.font = SECTION_FONT
    mhh.fill = SECTION_FILL
    ws.merge_cells(start_row=R_MOD_HDR, start_column=1,
                   end_row=R_MOD_HDR, end_column=12)

    mod_col_headers = ['Tactic'] + POSITION_CATEGORIES
    for i, h in enumerate(mod_col_headers, 1):
        ws.cell(row=R_MOD_COL, column=i, value=h)
    head_row(ws, R_MOD_COL, len(mod_col_headers))

    for i, tactic in enumerate(TACTIC_PRESETS):
        r = R_MOD_FIRST + i
        ws.cell(row=r, column=1, value=tactic).font = Font(bold=True)
        mods = TACTIC_MODIFIERS.get(tactic, {})
        for j, cat in enumerate(POSITION_CATEGORIES):
            val = mods.get(cat, 0)
            c = ws.cell(row=r, column=2 + j, value=val)
            c.alignment = CENTER
            # Positive green, negative red (visual cue).
            if isinstance(val, (int, float)) and val > 0:
                c.font = Font(color='375623', bold=True)
            elif isinstance(val, (int, float)) and val < 0:
                c.font = Font(color='C00000', bold=True)

    # =======================================================================
    # FREEZE + WIDTHS
    # =======================================================================
    ws.freeze_panes = f'A{R_SQUAD_FIRST}'

    widths = {
        1: 10, 2: 22, 3: 12, 4: 9, 5: 10, 6: 8, 7: 11, 8: 18, 9: 10,
        10: 7, 11: 9, 12: 8, 13: 8, 14: 8, 15: 8, 16: 8, 17: 8, 18: 8,
        19: 8, 20: 8,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def _extract_clubs_and_nations(df_source: pd.DataFrame) -> tuple[list[str], list[str]]:
    """Return (clubs, nations) sorted lists, filtered by minimum pool size."""
    club_series = df_source.iloc[:, SRC_CLUB_COL].dropna()
    nation_series = df_source.iloc[:, SRC_NATION_COL].dropna()

    club_counts = Counter(
        str(c) for c in club_series if str(c) != '0' and str(c).strip()
    )
    nation_counts = Counter(
        str(n) for n in nation_series if str(n) != 'Free Nationality' and str(n).strip()
    )

    clubs = sorted(c for c, n in club_counts.items() if n >= MIN_CLUB_POOL)
    nations = sorted(n for n, count in nation_counts.items() if count >= MIN_NATION_POOL)
    return clubs, nations


def main() -> int:
    print(f"Verifying {SRC}...")
    verify_source(SRC)

    OUT.parent.mkdir(parents=True, exist_ok=True)

    print(f"Loading {SRC}...")
    df_source = pd.read_excel(SRC, sheet_name='1. Paste cvs', header=0)
    print(f"  {len(df_source)} rows loaded.")

    clubs, nations = _extract_clubs_and_nations(df_source)
    print(f"  {len(clubs)} clubs (pool >= {MIN_CLUB_POOL}), "
          f"{len(nations)} nations (pool >= {MIN_NATION_POOL}).")

    wb = Workbook()
    build_readme(wb.active)
    wb.active.title = 'README'

    print("Building CSV Paste...")
    build_csv_paste(wb.create_sheet('CSV Paste'), df_source)

    print("Building Players (formulas)...")
    build_players(wb.create_sheet('Players'))

    print(f"Building Clubs ({len(clubs)} rows)...")
    build_clubs(wb.create_sheet('Clubs'), clubs)

    print(f"Building Nations ({len(nations)} rows)...")
    build_nations(wb.create_sheet('Nations'), nations)

    print(f"Building Team Composition ({len(clubs)} rows)...")
    build_team_composition(wb.create_sheet('Team Composition'), clubs)

    print("Building Draft Rules...")
    build_draft_rules(wb.create_sheet('Draft Rules'))

    print(f"Building Managers ({len(MANAGERS)} rows)...")
    build_managers(wb.create_sheet('Managers'))

    print(f"Building Competitions ({len(COMPETITIONS)} rows)...")
    build_competitions(wb.create_sheet('Competitions'))

    print("Building Expectations...")
    build_expectations(wb.create_sheet('Expectations'))

    print(f"Building Season Results ({SEASON_RESULTS_BLANK_ROWS} blank rows)...")
    build_season_results(wb.create_sheet('Season Results'))

    print("Building Tier Movement...")
    build_tier_movement(wb.create_sheet('Tier Movement'))

    print("Building Draft Order...")
    build_draft_order(wb.create_sheet('Draft Order'))

    print("Building Career Summary...")
    build_career_summary(wb.create_sheet('Career Summary'))

    print("Building Player Changes...")
    build_player_changes(wb.create_sheet('Player Changes'))

    print("Building Tactical Mentor...")
    build_tactical_mentor(wb.create_sheet('Tactical Mentor'))

    print(f"Saving {OUT}...")
    wb.save(OUT)
    print(f"Done. Sheets: {wb.sheetnames}")
    return 0


if __name__ == '__main__':
    sys.exit(main())
