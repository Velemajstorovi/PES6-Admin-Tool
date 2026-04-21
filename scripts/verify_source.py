"""Schema guardrail for data/source.xlsx.

Other scripts call verify_source() before reading source.xlsx so that a
re-exported or hand-edited file with a changed structure fails fast with a
message the user can act on, rather than producing silently-wrong output.

Can also be run directly:
    python scripts/verify_source.py                      # checks data/source.xlsx
    python scripts/verify_source.py path/to/other.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except AttributeError:
    pass

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SCRIPTS_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPTS_DIR.parent

# Actual column headers of source.xlsx's "1. Paste cvs" sheet, as produced by
# PES Fan Editor. build_admin.py reads this sheet positionally and copies
# values into the output workbook's "CSV Paste" sheet — so any change to the
# order or count of columns here breaks downstream formulas.
#
# Quirks preserved verbatim from the export (sic):
#   col 49:  'TACTIAL DRIBBLE'  (PES Fan Editor typo for TACTICAL)
#   col 81:  'CLUB TEAM'        (named CLUB in our internal admin_tool_builder)
#   cols 2-13: position IDs use double-spaces for single-digit codes
EXPECTED_CVS_HEADERS = [
    'NAME',
    'GK  0', 'CWP  2', 'CBT  3', 'SB  4', 'DMF  5', 'WB  6', 'CMF  7',
    'SMF  8', 'AMF  9', 'WF 10', 'SS  11', 'CF  12',
    'REGISTERED POSITION', 'HEIGHT', 'STRONG FOOT', 'FAVOURED SIDE',
    'WEAK FOOT ACCURACY', 'WEAK FOOT FREQUENCY',
    'ATTACK', 'DEFENSE', 'BALANCE', 'STAMINA', 'TOP SPEED', 'ACCELERATION',
    'RESPONSE', 'AGILITY', 'DRIBBLE ACCURACY', 'DRIBBLE SPEED',
    'SHORT PASS ACCURACY', 'SHORT PASS SPEED', 'LONG PASS ACCURACY',
    'LONG PASS SPEED', 'SHOT ACCURACY', 'SHOT POWER', 'SHOT TECHNIQUE',
    'FREE KICK ACCURACY', 'SWERVE', 'HEADING', 'JUMP', 'TECHNIQUE',
    'AGGRESSION', 'MENTALITY', 'GOAL KEEPING', 'TEAM WORK',
    'CONSISTENCY', 'CONDITION / FITNESS',
    'DRIBBLING', 'TACTIAL DRIBBLE', 'POSITIONING',
    'REACTION', 'PLAYMAKING', 'PASSING', 'SCORING', '1-1 SCORING',
    'POST PLAYER', 'LINES', 'MIDDLE SHOOTING', 'SIDE', 'CENTRE',
    'PENALTIES', '1-TOUCH PASS', 'OUTSIDE',
    'MARKING', 'SLIDING', 'COVERING', 'D-LINE CONTROL', 'PENALTY STOPPER',
    '1-ON-1 STOPPER', 'LONG THROW',
    'INJURY TOLERANCE', 'DRIBBLE STYLE', 'FREE KICK STYLE', 'PK STYLE',
    'DROP KICK STYLE',
    'AGE', 'WEIGHT', 'NATIONALITY', 'INTERNATIONAL NUMBER',
    'CLASSIC NUMBER', 'CLUB TEAM', 'CLUB NUMBER',
]

SHEET_SPECS = {
    '1. Paste cvs': {
        'positional': True,
        'columns': EXPECTED_CVS_HEADERS,
        'used_by': 'build_admin.py - rows are copied into the CSV Paste sheet by column position, so the order must match exactly.',
    },
}


class SourceSchemaError(Exception):
    """Raised when source.xlsx is structurally incompatible with the scripts."""


def _read_header(ws) -> list:
    row_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        return list(next(row_iter))
    except StopIteration:
        return []


def _check_positional(sheet_name: str, header: list, expected: list) -> list[str]:
    errors = []
    for i, exp in enumerate(expected):
        actual = header[i] if i < len(header) else None
        actual_norm = str(actual).strip() if actual is not None else None
        if actual_norm != exp:
            errors.append(
                f"  column {get_column_letter(i + 1)} (#{i + 1}): "
                f"expected {exp!r}, found {actual!r}"
            )
    return errors


def _check_set(sheet_name: str, header: list, expected: list) -> list[str]:
    found = {str(c).strip() for c in header if c is not None}
    missing = [c for c in expected if c not in found]
    if not missing:
        return []
    return [f"  missing columns: {missing}"]


def verify_source(path: str | Path) -> None:
    """Validate that the xlsx at *path* matches the expected schema.

    Raises SourceSchemaError with a consolidated, actionable message on failure.
    """
    path = Path(path)
    if not path.exists():
        raise SourceSchemaError(
            f"source.xlsx not found at: {path}\n"
            f"Place a PES Fan Editor Excel export at this path before running."
        )

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise SourceSchemaError(
            f"Could not open {path} as an Excel workbook: {e}\n"
            f"Re-export the patch from PES Fan Editor and try again."
        ) from e

    try:
        sheet_names = list(wb.sheetnames)
        problems: list[str] = []

        for sheet_name, spec in SHEET_SPECS.items():
            if sheet_name not in sheet_names:
                problems.append(
                    f"[{sheet_name}] sheet missing.\n"
                    f"  used_by: {spec['used_by']}\n"
                    f"  sheets found: {sheet_names}"
                )
                continue

            header = _read_header(wb[sheet_name])
            if spec['positional']:
                sub = _check_positional(sheet_name, header, spec['columns'])
            else:
                sub = _check_set(sheet_name, header, spec['columns'])

            if sub:
                problems.append(
                    f"[{sheet_name}] column mismatch.\n"
                    f"  used_by: {spec['used_by']}\n"
                    + '\n'.join(sub)
                )
    finally:
        wb.close()

    if problems:
        raise SourceSchemaError(
            f"source.xlsx at {path} is not compatible with this repo's scripts.\n\n"
            + '\n\n'.join(problems)
            + "\n\nFix options:\n"
              "  1. Re-export the patch from PES Fan Editor using the default layout.\n"
              "  2. If the patch format legitimately changed, update SHEET_SPECS in "
              "scripts/verify_source.py (and any affected column lists in the scripts)."
        )


def main() -> int:
    if len(sys.argv) > 1:
        target = Path(sys.argv[1])
    else:
        target = REPO_ROOT / 'data' / 'source.xlsx'

    try:
        verify_source(target)
    except SourceSchemaError as e:
        print(f"FAILED: {e}", file=sys.stderr)
        return 1
    print(f"OK: {target} matches expected schema.")
    return 0


if __name__ == '__main__':
    sys.exit(main())
